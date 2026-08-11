# -*- coding: utf-8 -*-
"""
Created on Mar 11 08:51:41 2024
@author:gcg
"""

import gurobipy as gp
from gurobipy import GRB
import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import math
import random
from pathlib import Path

N_t = 24 # time
Path("Figs").mkdir(exist_ok=True)
DATA_DIR = Path(__file__).resolve().parent
info = load_workbook(DATA_DIR / "settings.xlsx", data_only=True)
temdata = np.load(DATA_DIR / "tep2023.npy")
lmpdata = np.load(DATA_DIR / "lmp2023.npy")
tem = temdata[30*N_t:61*N_t]
lmp = lmpdata[30*N_t:61*N_t]
setting_table = info.worksheets[0]
Nsubfigs = 3
pri = np.zeros((Nsubfigs,N_t-1))
for i in range(Nsubfigs):
    for j in range(N_t-1):
        pri[i,j] = setting_table.cell(row=i+1, column=j+2).value
N_price = 2000
np.random.seed(42)  # set the random seed


def Fig2_d_g_j(SOC_ini,Cap,Pcmax,Pdmax,eta,pri,index):
    Res0= np.zeros((N_price)) 
    pri0 = np.linspace(min(pri)*0.8, max(pri)*1.2, N_price) # traverse the price
    
    #### min - abs min * 0.2
    
    
    for K in range(N_price):    
        # build the optimization
        model = gp.Model()    
        SOC = model.addVars( N_t+1, vtype=GRB.CONTINUOUS, lb=0.1, ub=1, name='SOC')
        Pc = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pcmax, name='Pc')
        Pd = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pdmax, name='Pd')
        '''storage'''
        model.addConstr((Cap*SOC[0] == Cap*SOC_ini), 'SOC_1')
        model.addConstrs((Cap*SOC[t] == Cap*SOC[t-1] - (Pd[t-1]/eta - eta*Pc[t-1]) for t in range(1, N_t+1)), 'SOC_t')
        model.addConstrs((Pc[t]*Pd[t]==0 for t in range(N_t)), 'PDPC')
        model.setObjective(((pri0[K]*((Pd[0]-Pc[0])) +(sum(pri[t-1]*((Pd[t]-Pc[t])) for t in range(1, N_t)))) ), GRB.MAXIMIZE)   
        # solve the model
        model.setParam('OutputFlag', 0)
        model.optimize()
        Res0[K] = Pd[0].X-Pc[0].X
    # plot the figure
    plt.figure(figsize=(12, 6))
    plt.plot(pri0,Res0.T,linewidth=3)
    plt.xlim((0.25,1.25))
    plt.xlabel('price(CNY/kWh)')
    plt.ylabel('discharge power(MW)')
    plt.savefig("Figs/"+index+".png",bbox_inches='tight',  transparent=True, dpi=600)

def Fig2_e_h_k(SOC_ini,Cap,Pcmax,Pdmax,eta,pri,index):
    Res0= np.zeros((N_price)) 
    pri0 = np.linspace(min(pri)*0.8, max(pri)*1.2, N_price) # traverse the price
    for K in range(N_price):    
        # build the optimization
        model = gp.Model()    
        SOC = model.addVars( N_t+1, vtype=GRB.CONTINUOUS, lb=0.1, ub=1, name='SOC')
        Pc = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pcmax, name='Pc')
        Pd = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pdmax, name='Pd')
        '''storage'''
        model.addConstr((Cap*SOC[0] == Cap*SOC_ini), 'SOC_1')
        model.addConstrs((Cap*SOC[t] == Cap*SOC[t-1] - (Pd[t-1]/eta - eta*Pc[t-1]) for t in range(1, N_t+1)), 'SOC_t')
        model.addConstrs((Pc[t]*Pd[t]==0 for t in range(N_t)), 'PDPC')
        model.setObjective(((pri0[K]*((Pd[0]-Pc[0])) +(sum(pri[t-1]*((Pd[t]-Pc[t])) for t in range(1,N_t)))) ), GRB.MAXIMIZE)   
        # solve the model
        model.setParam('OutputFlag', 0)
        model.optimize()
        Res0[K] = Pd[0].X-Pc[0].X ### MILP
    # plot the figure
    plt.figure(figsize=(12, 6))
    plt.plot(pri0/1000,Res0.T,linewidth=3)
    plt.xlim((0.06,0.08))
    plt.xlabel('price(USD/kWh)')
    plt.ylabel('discharge power(MW)')
    plt.savefig("Figs/"+index+".png",bbox_inches='tight',  transparent=True, dpi=600)

        
def Fig2_f_i_l(SOC_ini,Cap,Pcmax,Pdmax,eta,pri,index):
    Res0= np.zeros((N_price)) 
    pri0 = np.linspace(min(pri)*1.2, max(pri)*1.2, N_price) # traverse the price
    for K in range(N_price):    
        # build the optimization
        model = gp.Model()    
        SOC = model.addVars( N_t+1, vtype=GRB.CONTINUOUS, lb=0.1, ub=1, name='SOC')
        Pc = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pcmax, name='Pc')
        Pd = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pdmax, name='Pd')
        '''storage'''
        model.addConstr((Cap*SOC[0] == Cap*SOC_ini), 'SOC_1')
        model.addConstrs((Cap*SOC[t] == Cap*SOC[t-1] - (Pd[t-1]/eta - eta*Pc[t-1]) for t in range(1, N_t+1)), 'SOC_t')
        model.addConstrs((Pc[t]*Pd[t]==0 for t in range(N_t)), 'state')
        model.setObjective(((pri0[K]*((Pd[0]-Pc[0])) +(sum(pri[t-1]*((Pd[t]-Pc[t])) for t in range(1,N_t)))) ), GRB.MAXIMIZE)   
        # solve the model
        model.setParam('OutputFlag', 0)
        model.optimize()
        Res0[K] = Pd[0].X-Pc[0].X
    # plot the figure
    plt.figure(figsize=(12, 6))
    plt.plot(pri0/1000,Res0.T,linewidth=3)
    plt.xlim((0.02,0.04))
    plt.xlabel('price(USD/kWh)')
    plt.ylabel('discharge power(MW)')
    plt.savefig("Figs/"+index+".png",bbox_inches='tight',  transparent=True, dpi=600)    
    
def Fig2_m(SOC_ini,Cap,Pcmax,Pdmax,eta,pri,index):
    N_price = 1000
    alphaset = [0,0.01,0.02]
    Res0= np.zeros((len(alphaset),N_price)) 
    pri0 = np.linspace(0.6,0.68, N_price) # traverse the price
    for a in range(len(alphaset)):
        alpha = alphaset[a]
        for K in range(N_price):    
            # build the optimization
            model = gp.Model()    
            SOC = model.addVars( N_t+1, vtype=GRB.CONTINUOUS, lb=0.1, ub=1, name='SOC')
            Pc = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pcmax, name='Pc')
            Pd = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pdmax, name='Pd')
            '''storage'''
            model.addConstr((Cap*SOC[0] == Cap*SOC_ini), 'SOC_1')
            model.addConstrs((Cap*SOC[t] == (1-alpha)*Cap*SOC[t-1] - (Pd[t-1]/eta - eta*Pc[t-1]) for t in range(1, N_t+1)), 'SOC')
            model.setObjective(((pri0[K]*((Pd[0]-Pc[0])) +(sum(pri[t-1]*((Pd[t]-Pc[t])) for t in range(1,N_t)))) ), GRB.MAXIMIZE)   
            # solve the model
            model.setParam('OutputFlag', 0)
            model.optimize()
            Res0[a,K] = Pd[0].X-Pc[0].X
    # plot the figure
    plt.figure(figsize=(12, 6))
    for k in range(len(alphaset)):
        plt.plot(pri0,Res0[k,:].T,label='dissipation rate='+str(alphaset[k]),linewidth=3)
    plt.xlabel('price(CNY/kWh)')
    plt.ylabel('discharge power(MW)')
    plt.xlim((0.6,0.68))
    plt.ylim((-0.65,0.65))
    plt.legend(loc=4)
    plt.savefig("Figs/"+index+".png",bbox_inches='tight',  transparent=True, dpi=600)

    
def Fig2_n(SOC_ini,Cap,Pcmax,Pdmax,eta,pri,index):
    N_price = 1000
    alphaset = [0,0.01,0.02]
    Res0= np.zeros((len(alphaset),N_price)) 
    pri0 = np.linspace(55,75, N_price) # traverse the price
    for a in range(len(alphaset)):
        alpha = alphaset[a]
        for K in range(N_price):    
            # build the optimization
            model = gp.Model()    
            SOC = model.addVars( N_t+1, vtype=GRB.CONTINUOUS, lb=0.1, ub=1, name='SOC')
            Pc = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pcmax, name='Pc')
            Pd = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pdmax, name='Pd')
            '''storage'''
            model.addConstr((Cap*SOC[0] == Cap*SOC_ini), 'SOC_1')
            model.addConstrs((Cap*SOC[t] == (1-alpha)*Cap*SOC[t-1] - (Pd[t-1]/eta - eta*Pc[t-1]) for t in range(1, N_t+1)), 'SOC')
            model.setObjective(((pri0[K]*((Pd[0]-Pc[0])) +(sum(pri[t-1]*((Pd[t]-Pc[t])) for t in range(1,N_t)))) ), GRB.MAXIMIZE)   
            # solve the model
            model.setParam('OutputFlag', 0)
            model.optimize()
            Res0[a,K] = Pd[0].X-Pc[0].X
    # plot the figure
    plt.figure(figsize=(12, 6))
    for k in range(len(alphaset)):
        plt.plot(pri0/1000,Res0[k,:].T,label='dissipation rate='+str(alphaset[k]),linewidth=3)
    plt.xlabel('price(USD/kWh)')
    plt.ylabel('discharge power(MW)')
    plt.xlim((0.055,0.075))
    plt.ylim((-0.65,0.65))
    plt.legend(loc=4)
    plt.savefig("Figs/"+index+".png",bbox_inches='tight',  transparent=True, dpi=600)    

def Fig2_o(SOC_ini,Cap,Pcmax,Pdmax,eta,pri,index):
    N_price = 1000
    alphaset = [0,0.01,0.02]
    Res0= np.zeros((len(alphaset),N_price)) 
    pri0 = np.linspace(22,30, N_price) # traverse the price
    for a in range(len(alphaset)):
        alpha = alphaset[a]
        for K in range(N_price):    
            # build the optimization model
            model = gp.Model()    
            SOC = model.addVars( N_t+1, vtype=GRB.CONTINUOUS, lb=0.1, ub=1, name='SOC')
            Pc = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pcmax, name='Pc')
            Pd = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pdmax, name='Pd')
            '''storage'''
            model.addConstr((Cap*SOC[0] == Cap*SOC_ini), 'SOC_1')
            model.addConstrs((Cap*SOC[t] == (1-alpha)*Cap*SOC[t-1] - (Pd[t-1]/eta - eta*Pc[t-1]) for t in range(1, N_t+1)), 'SOC')
            model.addConstrs((Pc[t]*Pd[t]==0 for t in range(N_t)), 'PDPC')
            model.setObjective(((pri0[K]*((Pd[0]-Pc[0])) +(sum(pri[t-1]*((Pd[t]-Pc[t])) for t in range(1,N_t)))) ), GRB.MAXIMIZE)   
            # solve the model
            model.setParam('OutputFlag', 0)
            model.optimize()
            Res0[a,K] = Pd[0].X-Pc[0].X
    # plot the figure
    plt.figure(figsize=(12, 6))
    for k in range(len(alphaset)):
        plt.plot(pri0/1000,Res0[k,:].T,label='dissipation rate='+str(alphaset[k]),linewidth=3)
    plt.xlabel('price(USD/kWh)')
    plt.ylabel('discharge power(MW)')
    plt.xlim((0.022,0.030))
    plt.ylim((-0.65,0.65))
    plt.legend(loc=4)
    plt.savefig("Figs/"+index+".png",bbox_inches='tight',  transparent=True, dpi=600)    


def Fig2_q(Ti_ini,Tmax,Tmin,Pmax,Ra,C,eta,index):
    Res0= np.zeros((N_price)) 
    a = math.exp(-1/(Ra*C))
    To = tem[0*N_t+10:1*N_t+10]
    pri = lmp[0*N_t+10:1*N_t+10]
    print(pri)
    print(To)
    pri0 = np.linspace(15,30, N_price) # traverse the price
    for K in range(N_price):        
        # build the optimization model 
        model = gp.Model()    
        P = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pmax, name='Pd')
        Ti = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=Tmin, ub=Tmax, name='Ti')
        model.addConstr((Ti[0] == a*Ti_ini +(1-a)*(To[0] - Ra*eta*P[0]) ), 'Ti_0')  
        model.addConstrs((Ti[t] == a*Ti[t-1] +(1-a)*(To[t] - Ra*eta*P[t]) for t in range(1, N_t)), 'T_2-24')         
        model.setObjective(( pri0[K]*(P[0]) + sum(pri[t]*(P[t+1])  for t in range(0,N_t-1))), GRB.MINIMIZE)  
        # solve the model
        model.setParam('OutputFlag', 0)
        model.optimize()
        Res0[K] = P[0].X
    # plot the figure
    plt.figure(figsize=(12, 6))
    plt.plot(pri0/1000,Res0.T,linewidth=3)
    plt.xlim((0.015,0.030))
    plt.xlabel('price(USD/kWh)')
    plt.ylabel('demand power(MW)')
    plt.savefig("Figs/"+index+".png",bbox_inches='tight',  transparent=True, dpi=600)    

def Fig2_EV(Nev,pri,eta,index): 
    Emax = 1
    Emin = 0.1
    Emaxset = np.zeros((Nev,N_t+1))
    Eminset = np.zeros((Nev,N_t+1))
    Pdmaxset = np.zeros((Nev,N_t))

    SINIset = np.random.rand((Nev))*0.6+0.1
    Pseries = np.random.normal(7,0,size = Nev )
    for t in range(N_t):
        for s in range(Nev):
            Pdmaxset[s,t] = Pseries[s]
    Pcmaxset = Pdmaxset
    Capset = np.random.normal(50,0,size = Nev ) # Capacity 
    etaset = np.random.normal(eta,0,size = Nev ) # eta
    TArrset = np.random.normal(0,1,size = Nev ) #
    TDepset = np.random.normal(10,1,size = Nev )
    Ressum = np.zeros((N_price))
    
    for i in range(Nev):
        if etaset[i]>=1:
            etaset[i] = 1
        if TArrset[i]<0:
            TArrset[i]=0
        for t in range(N_t+1):
            if t<=TArrset[i]:
                Emaxset[i,t] = SINIset[i]
                Eminset[i,t] = SINIset[i]
            elif t>=TDepset[i]:
                Emaxset[i,t] = Emax
                Eminset[i,t] = Emax   
            else:
                Emaxset[i,t] = Emax
                Eminset[i,t] = Emin   
    pri0 = np.linspace(60,90, N_price) 
    for K in range(N_price):        
        # build the model 
        model = gp.Model()    
        SOC = model.addVars( Nev,N_t+1, vtype=GRB.CONTINUOUS, lb=Eminset, ub=Emaxset, name='SOC')
        Pd = model.addVars( Nev,N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pdmaxset, name='Pd')
        Pc = model.addVars( Nev,N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pcmaxset, name='Pc')
        model.addConstrs((Capset[s]*SOC[s,t+1] == Capset[s]*SOC[s,t] -Pd[s,t]/etaset[s] 
                        + Pc[s,t]*etaset[s] for s in range(Nev) for t in range(N_t)), 'SOC')
        model.setObjective(((pri0[K]*(sum(Pd[s,0] for s in range(Nev))-sum(Pc[s,0] for s in range(Nev)))
                           +(sum(pri[t-1]*(sum(Pd[s,t] for s in range(Nev))-sum(Pc[s,t] for s in range(Nev)) )
                                   for t in range(1,N_t))))  ), GRB.MAXIMIZE)  
        model.setParam('OutputFlag', 0)
        model.optimize()
        Ressum[K] = sum(Pd[s,0].X for s in range(Nev))-sum(Pc[s,0].X for s in range(Nev))

    plt.figure(figsize=(8, 4))
    x=pri0/1000
    plt.plot(x,Ressum,linewidth=3)
    plt.xlim(0.06,0.09)
    plt.xlabel('Price(USD/kWh)')
    plt.ylabel('Discharge Power(kW)')
    plt.savefig("Figs/"+index+".png",bbox_inches='tight',  transparent=True, dpi=600)    
    return TArrset

def Fig3_d(SOC_ini,Cap,Pcmax,Pdmax,eta,pri,index):
    N_price = 5000
    priset = np.zeros((3,N_t-1))
    priset[0,:] = pri
    priset[1,:] = pri * 2
    priset[2,:] = pri + 2
    Res0= np.zeros((3,N_price)) 
    
    for i in range(3):
        price = priset[i,:]
        pri0 = np.linspace(0,4, N_price) # traverse the price
        for K in range(N_price):    
            # build the optimization model
            model = gp.Model()    
            SOC = model.addVars( N_t+1, vtype=GRB.CONTINUOUS, lb=0.1, ub=1, name='SOC')
            Pc = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pcmax, name='Pc')
            Pd = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pdmax, name='Pd')
            '''storage'''
            model.addConstr((Cap*SOC[0] == Cap*SOC_ini), 'SOC_1')
            model.addConstrs((Cap*SOC[t] == Cap*SOC[t-1] - (Pd[t-1]/eta - eta*Pc[t-1]) for t in range(1, N_t+1)), 'SOC')
            model.setObjective(((pri0[K]*((Pd[0]-Pc[0])) +(sum(price[t-1]*((Pd[t]-Pc[t])) for t in range(1,N_t)))) ), GRB.MAXIMIZE)   
            # solve the model
            model.setParam('OutputFlag', 0)
            model.optimize()
            Res0[i,K] = Pd[0].X-Pc[0].X
    # plot the figure
    plt.figure(figsize=(3, 2))
    for k in range(3):
        plt.plot(pri0,Res0[k,:].T,label='price'+str(k),linewidth=3)
    plt.xlabel('price(CNY/kWh)')
    plt.ylabel('discharge power(MW)')
    plt.xlim((0,4))
    plt.ylim((-0.65,0.65))
    # plt.legend(loc=4)
    plt.savefig("Figs/3-"+index+".png",bbox_inches='tight',  transparent=True, dpi=600)    


def Fig3_e(SOC_ini,Cap,Pcmax,Pdmax,eta,pri,index):
    N_price = 5000
    priset = np.zeros((3,N_t-1))
    priset[0,:] = pri
    priset[1,:] = pri * 2
    priset[2,:] = pri + 80
    Res0= np.zeros((3,N_price)) 
    
    for i in range(3):
        price = priset[i,:]
        pri0 = np.linspace(0,230, N_price) # traverse the price
        for K in range(N_price):    
            # build the optimization model
            model = gp.Model()    
            SOC = model.addVars( N_t+1, vtype=GRB.CONTINUOUS, lb=0.1, ub=1, name='SOC')
            Pc = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pcmax, name='Pc')
            Pd = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pdmax, name='Pd')
            '''storage'''
            model.addConstr((Cap*SOC[0] == Cap*SOC_ini), 'SOC_1')
            model.addConstrs((Cap*SOC[t] == Cap*SOC[t-1] - (Pd[t-1]/eta - eta*Pc[t-1]) for t in range(1, N_t+1)), 'SOC')
            model.setObjective(((pri0[K]*((Pd[0]-Pc[0])) +(sum(price[t-1]*((Pd[t]-Pc[t])) for t in range(1,N_t)))) ), GRB.MAXIMIZE)   
            # solve the model
            model.setParam('OutputFlag', 0)
            model.optimize()
            Res0[i,K] = Pd[0].X-Pc[0].X
    # plot the figure
    plt.figure(figsize=(3, 2))
    for k in range(3):
        plt.plot(pri0/1000,Res0[k,:].T,label='price'+str(k),linewidth=3)
    plt.xlabel('price(USD/kWh)')
    plt.ylabel('discharge power(MW)')
    plt.xlim((0,0.23))
    plt.ylim((-0.65,0.65))
    # plt.legend(loc=4)
    plt.savefig("Figs/3-"+index+".png",bbox_inches='tight',  transparent=True, dpi=600)    

def Fig3_f(SOC_ini,Cap,Pcmax,Pdmax,eta,pri,index):
    N_price = 5000
    priset = np.zeros((3,N_t-1))
    priset[0,:] = pri
    priset[1,:] = pri * 2
    priset[2,:] = pri + 80
    Res0= np.zeros((3,N_price)) 
    
    for i in range(3):
        price = priset[i,:]
        pri0 = np.linspace(0,130, N_price) # traverse the price
        for K in range(N_price):    
            # build the optimization model
            model = gp.Model()    
            SOC = model.addVars( N_t+1, vtype=GRB.CONTINUOUS, lb=0.1, ub=1, name='SOC')
            Pc = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pcmax, name='Pc')
            Pd = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pdmax, name='Pd')
            '''storage'''
            model.addConstr((Cap*SOC[0] == Cap*SOC_ini), 'SOC_1')
            model.addConstrs((Cap*SOC[t] == Cap*SOC[t-1] - (Pd[t-1]/eta - eta*Pc[t-1]) for t in range(1, N_t+1)), 'SOC')
            model.addConstrs((Pc[t]*Pd[t]==0 for t in range(N_t)), 'PDPC')
            model.setObjective(((pri0[K]*((Pd[0]-Pc[0])) +(sum(price[t-1]*((Pd[t]-Pc[t])) for t in range(1,N_t)))) ), GRB.MAXIMIZE)   
            # solve the model
            model.setParam('OutputFlag', 0)
            model.optimize()
            Res0[i,K] = Pd[0].X-Pc[0].X
    # plot the figure
    plt.figure(figsize=(3, 2))
    for k in range(3):
        plt.plot(pri0/1000,Res0[k,:].T,label='price'+str(k),linewidth=3)
    plt.xlabel('price(USD/kWh)')
    plt.ylabel('discharge power(MW)')
    plt.xlim((0,0.13))
    plt.ylim((-0.65,0.65))
    # plt.legend(loc=4)
    plt.savefig("Figs/3-"+index+".png",bbox_inches='tight',  transparent=True, dpi=600)    


def Fig3_j(SOC_ini,Cap,Pcmax,Pdmax,eta,pri,index):
    N_price = 1000
    Siniset = [0.2,0.8,0.5]
    Res0= np.zeros((3,N_price)) 
    pri0 = np.linspace(0.2,1.2, N_price) # traverse the price
    for i in range(3):
        SOC_ini = Siniset[i]
        for K in range(N_price):    
            # build the optimization model
            model = gp.Model()    
            SOC = model.addVars( N_t+1, vtype=GRB.CONTINUOUS, lb=0.1, ub=1, name='SOC')
            Pc = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pcmax, name='Pc')
            Pd = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pdmax, name='Pd')
            '''storage'''
            model.addConstr((Cap*SOC[0] == Cap*SOC_ini), 'SOC_1')
            model.addConstrs((Cap*SOC[t] == Cap*SOC[t-1] - (Pd[t-1]/eta - eta*Pc[t-1]) for t in range(1, N_t+1)), 'SOC')
            model.setObjective(((pri0[K]*((Pd[0]-Pc[0])) +(sum(pri[t-1]*((Pd[t]-Pc[t])) for t in range(1,N_t)))) ), GRB.MAXIMIZE)   
            # solve the model
            model.setParam('OutputFlag', 0)
            model.optimize()
            Res0[i,K] = Pd[0].X-Pc[0].X
    # plot the figure
    plt.figure(figsize=(3, 2))
    for k in range(3):
        plt.plot(pri0,Res0[k,:].T,label='SOCini='+str(Siniset[k]),linewidth=3)
    plt.xlabel('price(CNY/kWh)')
    plt.ylabel('discharge power(MW)')
    plt.xlim((0.2,1.2))
    plt.ylim((-0.65,0.65))
    # plt.legend(loc=4)
    plt.savefig("Figs/3-"+index+".png",bbox_inches='tight',  transparent=True, dpi=600)    

def Fig3_k(SOC_ini,Cap,Pcmax,Pdmax,eta,pri,index):
    N_price = 1000
    Siniset = [0.2,0.8,0.5]
    Res0= np.zeros((3,N_price)) 
    pri0 = np.linspace(40,120, N_price) # traverse the price
    for i in range(3):
        SOC_ini = Siniset[i]
        for K in range(N_price):    
            # build the optimization model
            model = gp.Model()    
            SOC = model.addVars( N_t+1, vtype=GRB.CONTINUOUS, lb=0.1, ub=1, name='SOC')
            Pc = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pcmax, name='Pc')
            Pd = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pdmax, name='Pd')
            '''storage'''
            model.addConstr((Cap*SOC[0] == Cap*SOC_ini), 'SOC_1')
            model.addConstrs((Cap*SOC[t] == Cap*SOC[t-1] - (Pd[t-1]/eta - eta*Pc[t-1]) for t in range(1, N_t+1)), 'SOC')
            model.setObjective(((pri0[K]*((Pd[0]-Pc[0])) +(sum(pri[t-1]*((Pd[t]-Pc[t])) for t in range(1,N_t)))) ), GRB.MAXIMIZE)   
            # solve the model
            model.setParam('OutputFlag', 0)
            model.optimize()
            Res0[i,K] = Pd[0].X-Pc[0].X
    # plot the figure
    plt.figure(figsize=(3, 2))
    for k in range(3):
        plt.plot(pri0/1000,Res0[k,:].T,label='SOCini='+str(Siniset[k]),linewidth=3)
    plt.xlabel('price(USD/kWh)')
    plt.ylabel('discharge power(MW)')
    plt.xlim((0.04,0.12))
    plt.ylim((-0.65,0.65))
    # plt.legend(loc=4)
    plt.savefig("Figs/3-"+index+".png",bbox_inches='tight',  transparent=True, dpi=600)   
    
def Fig3_l(SOC_ini,Cap,Pcmax,Pdmax,eta,pri,index):
    N_price = 1000
    Siniset = [0.2,0.8,0.5]
    Res0= np.zeros((3,N_price)) 
    pri0 = np.linspace(20,35, N_price) # traverse the price
    for i in range(3):
        SOC_ini = Siniset[i]
        for K in range(N_price):    
            # build the optimization model
            model = gp.Model()    
            SOC = model.addVars( N_t+1, vtype=GRB.CONTINUOUS, lb=0.1, ub=1, name='SOC')
            Pc = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pcmax, name='Pc')
            Pd = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pdmax, name='Pd')
            '''storage'''
            model.addConstr((Cap*SOC[0] == Cap*SOC_ini), 'SOC_1')
            model.addConstrs((Cap*SOC[t] == Cap*SOC[t-1] - (Pd[t-1]/eta - eta*Pc[t-1]) for t in range(1, N_t+1)), 'SOC')
            model.addConstrs((Pc[t]*Pd[t]==0 for t in range(N_t)), 'PDPC')
            model.setObjective(((pri0[K]*((Pd[0]-Pc[0])) +(sum(pri[t-1]*((Pd[t]-Pc[t])) for t in range(1,N_t)))) ), GRB.MAXIMIZE)   
            # solve the model
            model.setParam('OutputFlag', 0)
            model.optimize()
            Res0[i,K] = Pd[0].X-Pc[0].X
    # plot the figure
    plt.figure(figsize=(3, 2))
    for k in range(3):
        plt.plot(pri0/1000,Res0[k,:].T,label='SOCini='+str(Siniset[k]),linewidth=3)
    plt.xlabel('price(USD/kWh)')
    plt.ylabel('discharge power(MW)')
    plt.xlim((0.02,0.035))
    plt.ylim((-0.65,0.65))
    # plt.legend(loc=4)
    plt.savefig("Figs/3-"+index+".png",bbox_inches='tight',  transparent=True, dpi=600)   
    
    
# def Fig2_f_i_l_new(SOC_ini,Cap,Pcmax,Pdmax,eta,index):
#     Res0= np.zeros((N_price))
#     pri1 = random.uniform(-100, 0)
#     pri = np.random.uniform(low=pri1/eta/eta, high=pri1, size=N_t-1)
#     pri0 = np.linspace(min(pri)*1.2, max(pri)*1.2, N_price) # traverse the price
#     for K in range(N_price):    
#         # build the optimization
#         model = gp.Model()    
#         SOC = model.addVars( N_t+1, vtype=GRB.CONTINUOUS, lb=0.1, ub=1, name='SOC')
#         Pc = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pcmax, name='Pc')
#         Pd = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pdmax, name='Pd')
#         '''storage'''
#         model.addConstr((Cap*SOC[0] == Cap*SOC_ini), 'SOC_1')
#         model.addConstrs((Cap*SOC[t] == Cap*SOC[t-1] - (Pd[t-1]/eta - eta*Pc[t-1]) for t in range(1, N_t+1)), 'SOC_t')
#         model.addConstrs((Pc[t]*Pd[t]==0 for t in range(N_t)), 'state')
#         model.setObjective(((pri0[K]*((Pd[0]-Pc[0])) +(sum(pri[t-1]*((Pd[t]-Pc[t])) for t in range(1,N_t)))) ), GRB.MAXIMIZE)   
#         # solve the model
#         model.setParam('OutputFlag', 0)
#         model.optimize()
#         Res0[K] = Pd[0].X-Pc[0].X
#     # plot the figure
#     plt.figure(figsize=(12, 6))
#     plt.plot(pri0/1000,Res0.T,linewidth=3)
#     # plt.xlim((0.02,0.04))
#     plt.xlabel('price(USD/kWh)')
#     plt.ylabel('discharge power(MW)')
#     plt.savefig("Figs/"+index+".png",bbox_inches='tight',  transparent=True, dpi=600)        
    
    
Fig2_d_g_j(0.5,2,0.6,0.6,1,pri[0,:],"d")
Fig2_e_h_k(0.5,2,0.6,0.6,1,pri[1,:],"e")
Fig2_f_i_l(0.5,2,0.6,0.6,1,pri[2,:],"f")
Fig2_d_g_j(0.5,2,0.6,0.6,0.98,pri[0,:],"g")
Fig2_e_h_k(0.5,2,0.6,0.6,0.98,pri[1,:],"h")
Fig2_f_i_l(0.5,2,0.6,0.6,0.98,pri[2,:],"i")
Fig2_d_g_j(0.5,2,0.6,0.6,0.9,pri[0,:],"j")
Fig2_e_h_k(0.5,2,0.6,0.6,0.9,pri[1,:],"k")
Fig2_f_i_l(0.5,2,0.6,0.6,0.9,pri[2,:],"l")
Fig2_m(0.5,2,0.6,0.6,1,pri[0,:],"m")
Fig2_n(0.5,2,0.6,0.6,1,pri[1,:],"n")
Fig2_o(0.5,2,0.6,0.6,1,pri[2,:],"o")
Fig2_q(23,24,21,12,3,5,1,"p")
Fig2_q(23,24,21,12,3,5,0.98,"q")
Fig2_q(23,24,21,12,3,5,0.95,"r")
Fig2_EV(20,pri[1,:],1,"s")
Fig2_EV(20,pri[1,:],0.98,"t")
Fig2_EV(20,pri[1,:],0.95,"u")

Fig3_d(0.5,2,0.6,0.6,1,pri[0,:],"d")
Fig3_e(0.5,2,0.6,0.6,1,pri[1,:],"e")
Fig3_f(0.5,2,0.6,0.6,1,pri[2,:],"f")
Fig3_d(0.5,2,0.6,0.6,0.9,pri[0,:],"g")
Fig3_e(0.5,2,0.6,0.6,0.9,pri[1,:],"h")
Fig3_f(0.5,2,0.6,0.6,0.9,pri[2,:],"i")
Fig3_j(0.5,2,0.6,0.6,1,pri[0,:],"j")
Fig3_k(0.5,2,0.6,0.6,1,pri[1,:],"k")
Fig3_l(0.5,2,0.6,0.6,1,pri[2,:],"l")

# for i in range(5):
#     Fig2_f_i_l_new(0.5,2,0.6,0.6,0.8,str(i))
