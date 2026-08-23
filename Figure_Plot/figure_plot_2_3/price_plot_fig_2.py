# -*- coding: utf-8 -*-
"""
Created on Aug 14 2026
Price curves for the three price series used in Fig. 2 / Fig. 3.

Plots pri[0,:], pri[1,:], pri[2,:] (the three rows of settings.xlsx)
as three separate figures, 11 cm x 4.5 cm each.
"""

import numpy as np
import matplotlib.pyplot as plt
from matplotlib.ticker import FormatStrFormatter
from openpyxl import load_workbook
from pathlib import Path

# ----- global style: Times New Roman -----
plt.rcParams['font.family'] = 'serif'
plt.rcParams['font.serif'] = ['Times New Roman', 'Times', 'DejaVu Serif']
plt.rcParams['font.size'] = 15

N_forecasts = 23  # Forecast prices 1..23; the original price 24 is omitted.
N_t = N_forecasts + 1  # Include x=0 for the binding-interval price.
Path("Figs").mkdir(exist_ok=True)
DATA_DIR = Path(__file__).resolve().parent

info = load_workbook(DATA_DIR / "settings.xlsx", data_only=True)
setting_table = info.worksheets[0]

Nsubfigs = 3
pri = np.zeros((Nsubfigs, N_forecasts))
for i in range(Nsubfigs):
    for j in range(N_forecasts):
        pri[i, j] = setting_table.cell(row=i + 1, column=j + 2).value

t = np.arange(N_t)  # x=0 is binding-interval price; x=1..23 are forecasts.
figsize = (11 / 2.54, 4.5 / 2.54)  # 11 cm x 4.5 cm


def add_binding_interval_price(price):
    """Prepend the x=0 price by linearly extrapolating forecast points 1 and 2."""
    binding_price = 2 * price[0] - price[1]
    return np.concatenate(([binding_price], price[:N_forecasts]))

# (price data, ylabel, yticks) for figures 1..3
configs = [
    (add_binding_interval_price(pri[0, :]),        'Price (CNY/kWh)', [0.25, 0.50, 0.75, 1.00]),
    (add_binding_interval_price(pri[1, :] / 1000), 'Price (USD/kWh)', [0.06, 0.07, 0.08]),
    (add_binding_interval_price(pri[2, :] / 1000), 'Price (USD/kWh)', [0.00, 0.02, 0.04]),
]

for idx, (price, ylabel, yticks) in enumerate(configs, start=1):
    fig, ax = plt.subplots(figsize=figsize)
    ax.plot(t, price, linewidth=3)
    ax.set_xlim(0, 23)
    ax.set_xticks([0, 1, 5, 10, 15, 20, 23])
    ax.set_xticklabels(['', '1', '5', '10', '15', '20', '23'])
    ax.annotate('(binding)', xy=(0, price[0]), xytext=(0, -0.24),
                textcoords='axes fraction', ha='center', va='top',
                arrowprops=dict(arrowstyle='-', linestyle='--', lw=1.2,
                                color='k'))
    ax.set_xlabel('Time (hours)')
    ax.set_ylabel(ylabel)
    ax.set_yticks(yticks)
    ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
    fig.savefig("Figs/price_%d.png" % idx, bbox_inches='tight',
                transparent=True, dpi=600)
    plt.close(fig)
