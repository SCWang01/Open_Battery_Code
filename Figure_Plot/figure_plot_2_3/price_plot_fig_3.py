# -*- coding: utf-8 -*-
"""Plot the three price-series sets used by Fig3_d, Fig3_e and Fig3_f."""

from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import FormatStrFormatter
from openpyxl import load_workbook


# ----- global style: Times New Roman -----
plt.rcParams["font.family"] = "Times New Roman"
plt.rcParams["font.serif"] = ["Times New Roman"]
plt.rcParams["font.size"] = 14
plt.rcParams["axes.labelsize"] = 14
plt.rcParams["xtick.labelsize"] = 14
plt.rcParams["ytick.labelsize"] = 14
plt.rcParams["legend.fontsize"] = 14

N_forecasts = 23  # Forecast prices 1..23; the original price 24 is omitted.
N_t = N_forecasts + 1  # Include x=0 for the binding-interval price.
Path("Figs").mkdir(exist_ok=True)
DATA_DIR = Path(__file__).resolve().parent

info = load_workbook(DATA_DIR / "settings.xlsx", data_only=True)
setting_table = info.worksheets[0]

pri = np.zeros((3, N_forecasts))
for i in range(3):
    for j in range(N_forecasts):
        pri[i, j] = setting_table.cell(row=i + 1, column=j + 2).value

t = np.arange(N_t)  # x=0 is binding-interval price; x=1..23 are forecasts.
figsize = (10 / 2.54, 5/ 2.54)


def add_binding_interval_price(price):
    """Prepend the x=0 price by linearly extrapolating forecast points 1 and 2."""
    binding_price = 2 * price[:, 0] - price[:, 1]
    return np.concatenate((binding_price[:, None], price[:, :N_forecasts]), axis=1)

configs = [
    (
        "d",
        add_binding_interval_price(
            np.vstack((pri[0, :], pri[0, :] * 2, pri[0, :] + 2))
        ),
        "Price (CNY/kWh)",
        [1, 2, 3],
        "%.0f",
    ),
    (
        "e",
        add_binding_interval_price(
            np.vstack((pri[1, :], pri[1, :] * 2, pri[1, :] + 80)) / 1000
        ),
        "Price (USD/kWh)",
        [0.05, 0.10, 0.15],
        "%.2f",
    ),
    (
        "f",
        add_binding_interval_price(
            np.vstack((pri[2, :], pri[2, :] * 2, pri[2, :] + 80)) / 1000
        ),
        "Price (USD/kWh)",
        [0.00, 0.05, 0.10],
        "%.2f",
    ),
]

for index, prices, ylabel, yticks, yformat in configs:
    fig, ax = plt.subplots(figsize=figsize)
    for price in prices:
        ax.plot(t, price, linewidth=3)

    ax.set_xlim(0, 23)
    ax.set_xticks([0, 1, 5, 10, 15, 20, 23])
    ax.set_xticklabels(["", "1", "5", "10", "15", "20", "23"])
    ax.annotate(
        "(binding)",
        xy=(0, np.max(prices[:, 0])),
        xytext=(0, -0.24),
        textcoords="axes fraction",
        ha="center",
        va="top",
        arrowprops=dict(arrowstyle="-", linestyle="--", lw=1.2, color="k"),
    )
    ax.set_xlabel("Time (hours)")
    ax.set_ylabel(ylabel)
    ax.set_yticks(yticks)
    ax.yaxis.set_major_formatter(FormatStrFormatter(yformat))

    fig.savefig(
        f"Figs/price_3_{index}.png",
        bbox_inches="tight",
        transparent=True,
        dpi=600,
    )
    plt.close(fig)
