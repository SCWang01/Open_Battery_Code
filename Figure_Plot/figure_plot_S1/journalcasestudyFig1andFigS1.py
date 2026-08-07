# -*- coding: utf-8 -*-
"""
Created on Mar 11 08:51:41 2024
@author:gcg
"""
#%% Global settings

import gurobipy as gp
from gurobipy import GRB
import numpy as np
from openpyxl import load_workbook

#%% Global settings
N_t = 24 # time
info = load_workbook('settings_sequence.xlsx', data_only=True)# datatable
setting_table_positive = info.worksheets[0] # positive price data for Fig1
setting_table_negative = info.worksheets[1] # negative price data for FigS1
N_price = 6 # six stairs
pri_pos = np.zeros((N_price,N_t)) # initialize the price matrix
pri_neg = np.zeros((N_price,N_t)) # initialize the price matrix
# read price data
for i in range(N_price):
    for j in range(N_t):
        pri_pos[i,j] = setting_table_positive.cell(row=i+2, column=j+2).value  # load the positve-case price of Fig 1
        pri_neg[i,j] = setting_table_negative.cell(row=i+2, column=j+2).value # load the negatve-case price of Fig S1

#%% function
def Fig1(SOC_ini,SOC_min,Cap,Pcmax,Pdmax,eta,pri):
    Res= np.zeros((N_price,N_t)) # Initialize the P matrix
    ResE = np.zeros((N_price,N_t+1)) # Initialize the E matrix
    for K in range(N_price):    # for each stair example
        # build the optimization
        model = gp.Model()    
        # set variables
        SOC = model.addVars( N_t+1, vtype=GRB.CONTINUOUS, lb=SOC_min, ub=1, name='SOC') # State of charge
        Pc = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pcmax, name='Pc') # charge power
        Pd = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pdmax, name='Pd') # discharge power
        # set constraints
        model.addConstr((Cap*SOC[0] == Cap*SOC_ini), 'SOC_1')
        model.addConstrs((Cap*SOC[t] == Cap*SOC[t-1] - (Pd[t-1]/eta - eta*Pc[t-1]) for t in range(1, N_t+1)), 'SOC_t')
        model.addConstrs((Pc[t]*Pd[t]==0 for t in range(N_t)), 'PDPC')
        model.setObjective(((sum(pri[K,t]*((Pd[t]-Pc[t])) for t in range(0, N_t))) ), GRB.MAXIMIZE)   
        # solve the model
        model.setParam('OutputFlag', 0)
        model.optimize()
        # save the results 
        ResE[K,0] = Cap*SOC_ini
        for t in range(N_t):
            Res[K,t] = Pd[t].X-Pc[t].X
            if t > 0:
                ResE[K,t] = Cap*SOC[t].X
        ResE[K,N_t] = Cap*SOC[N_t].X
    return Res, ResE

#%% main
#Fig1
#Res_pos, ResE_pos = Fig1(14/22,0,22,5,5,1,pri_pos)# sequence is Res_pos

#Fig S1
Res_neg, ResE_neg = Fig1(0.5,0.1,2,0.6,0.6,0.9,pri_neg)# sequence is Res_pos

# %%
