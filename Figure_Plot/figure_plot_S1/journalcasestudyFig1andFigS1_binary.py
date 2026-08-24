# -*- coding: utf-8 -*-
"""
Created on Mar 11 08:51:41 2024
@author:gcg

Charge/discharge mutual exclusion follows the paper's two-status formulation:
    Pc[t] <= Pcmax * sc[t]
    Pd[t] <= Pdmax * sd[t]
    sc[t] + sd[t] <= 1
where sc[t], sd[t] ∈ {0,1}; (0,0) is the explicit idle mode.
"""
#%% Global settings

import gurobipy as gp
from gurobipy import GRB
import numpy as np
from openpyxl import load_workbook

#%% Global settings
N_t = 24 # time
WORKBOOK_PATH = 'settings_sequence.xlsx'
info = load_workbook(WORKBOOK_PATH, data_only=True) # datatable
setting_table_positive = info.worksheets[0] # positive price data for Fig1
setting_table_negative = info.worksheets[1] # negative price data for FigS1
N_price = 6 # six stairs
pri_pos = np.zeros((N_price,N_t)) # initialize the price matrix
pri_neg = np.zeros((N_price,N_t)) # initialize the price matrix
# read price data
for i in range(N_price):
    for j in range(N_t):
        pri_pos[i,j] = setting_table_positive.cell(row=i+2, column=j+2).value  # load the positve-case price of Fig 1
        pri_neg[i,j] = setting_table_negative.cell(row=i+2, column=j+2).value  # load the negatve-case price of Fig S1
info.close()

#%% function
def Fig1(SOC_ini,SOC_min,Cap,Pcmax,Pdmax,eta,pri):
    Res= np.zeros((N_price,N_t)) # Initialize the P matrix
    ResE = np.zeros((N_price,N_t+1)) # Initialize the E matrix
    for K in range(N_price):    # for each stair example
        # build the optimization
        model = gp.Model()
        # set variables
        SOC          = model.addVars(N_t+1, vtype=GRB.CONTINUOUS, lb=SOC_min, ub=1,     name='SOC') # State of charge
        Pc           = model.addVars(N_t,   vtype=GRB.CONTINUOUS, lb=0,       ub=Pcmax, name='Pc')  # charge power
        Pd           = model.addVars(N_t,   vtype=GRB.CONTINUOUS, lb=0,       ub=Pdmax, name='Pd')  # discharge power
        sc = model.addVars(N_t, vtype=GRB.BINARY, name='sc')
        sd = model.addVars(N_t, vtype=GRB.BINARY, name='sd')
        # set constraints
        model.addConstr((Cap*SOC[0] == Cap*SOC_ini), 'SOC_1')
        model.addConstrs((Cap*SOC[t] == Cap*SOC[t-1] - (Pd[t-1]/eta - eta*Pc[t-1]) for t in range(1, N_t+1)), 'SOC_t')
        model.addConstrs((Pc[t] <= Pcmax * sc[t] for t in range(N_t)), 'Pc_mode')
        model.addConstrs((Pd[t] <= Pdmax * sd[t] for t in range(N_t)), 'Pd_mode')
        model.addConstrs((sc[t] + sd[t] <= 1 for t in range(N_t)), 'mutually_exclusive_state')
        model.setObjective(((sum(pri[K,t]*((Pd[t]-Pc[t])) for t in range(0, N_t))) ), GRB.MAXIMIZE)
        # solve the model
        model.setParam('OutputFlag', 0)
        model.optimize()
        # save the results
        ResE[K,0] = Cap*SOC_ini
        for t in range(N_t):
            Res[K,t] = Pd[t].X-Pc[t].X
            if t > 0:
                ResE[K,t] = Cap*SOC[t].X
        ResE[K,N_t] = Cap*SOC[N_t].X
    return Res, ResE


def save_and_verify_results(workbook_path, power, soc):
    """Write power and SOC results and verify the Excel round trip."""
    expected_shapes = {'Power': (N_price, N_t), 'SOC': (N_price, N_t + 1)}
    results = {'Power': power, 'SOC': soc}

    for sheet_name, expected_shape in expected_shapes.items():
        if results[sheet_name].shape != expected_shape:
            raise ValueError(
                f'{sheet_name} result has shape {results[sheet_name].shape}; '
                f'expected {expected_shape}'
            )

    workbook = load_workbook(workbook_path)
    missing_sheets = [name for name in results if name not in workbook.sheetnames]
    if missing_sheets:
        workbook.close()
        raise KeyError(f'Missing result sheets: {missing_sheets}')

    for sheet_name, values in results.items():
        worksheet = workbook[sheet_name]
        for row_index in range(values.shape[0]):
            for column_index in range(values.shape[1]):
                worksheet.cell(
                    row=row_index + 2,
                    column=column_index + 2,
                    value=float(values[row_index, column_index]),
                )

    workbook.save(workbook_path)
    workbook.close()

    saved_workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        for sheet_name, expected in results.items():
            worksheet = saved_workbook[sheet_name]
            actual = np.array([
                [
                    worksheet.cell(row=row_index + 2, column=column_index + 2).value
                    for column_index in range(expected.shape[1])
                ]
                for row_index in range(expected.shape[0])
            ], dtype=float)
            # Excel serializes floating-point values as decimal text. Allow only
            # sub-picounit round-trip noise while requiring every cell to match.
            np.testing.assert_allclose(actual, expected, rtol=0, atol=1e-12)
    finally:
        saved_workbook.close()

#%% main
#Fig1
# Res_pos, ResE_pos = Fig1(14/22,0,22,5,5,1,pri_pos)# sequence is Res_pos

#Fig S1
Res_neg, ResE_neg = Fig1(0.5,0.1,2,0.6,0.6,0.9,pri_neg)# sequence is Res_neg
save_and_verify_results(WORKBOOK_PATH, Res_neg, ResE_neg)

# %%
