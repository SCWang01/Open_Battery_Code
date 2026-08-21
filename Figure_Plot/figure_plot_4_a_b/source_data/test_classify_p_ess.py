from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from classify_p_ess import (
    MAX_RATE_ABS_TOL,
    classify_p_ess,
    classify_workbook,
    resolve_power_limits,
)


def ds_row(power: float, seventh: float = 1.0) -> list[float]:
    """Return one valid eight-column dsfunction segment."""
    return [0.0, 1.0, power, 0.0, 0.0, 0.0, seventh, 0.5]


class PowerLimitTests(unittest.TestCase):
    def test_asymmetric_explicit_limits(self) -> None:
        self.assertEqual(
            classify_p_ess(
                1860.0,
                (),
                2,
                pcmax=1540.0,
                pdmax=1860.0,
            ),
            "MD",
        )
        self.assertEqual(
            classify_p_ess(
                -1540.0,
                (),
                2,
                pcmax=1540.0,
                pdmax=1860.0,
            ),
            "MC",
        )

    def test_near_limit_values_use_rounding_aware_tolerance(self) -> None:
        offset = MAX_RATE_ABS_TOL * 0.8
        self.assertEqual(
            classify_p_ess(
                1860.0 - offset,
                (),
                2,
                pcmax=1540.0,
                pdmax=1860.0,
            ),
            "MD",
        )
        self.assertEqual(
            classify_p_ess(
                -1540.0 + offset,
                (),
                2,
                pcmax=1540.0,
                pdmax=1860.0,
            ),
            "MC",
        )

    def test_infers_asymmetric_limits_from_all_dsfunction_stairs(self) -> None:
        values = [
            (
                repr(
                    [
                        ds_row(-1540.0),
                        ds_row(0.0),
                        ds_row(1200.0),
                    ]
                ),
                2,
            ),
            (
                repr(
                    [
                        ds_row(-800.0),
                        ds_row(1860.0),
                    ]
                ),
                3,
            ),
        ]

        self.assertEqual(resolve_power_limits(values), (1540.0, 1860.0))

    def test_explicit_limits_override_inferred_extrema(self) -> None:
        stale_values = [
            (
                repr(
                    [
                        ds_row(-1860.0),
                        ds_row(1860.0),
                    ]
                ),
                2,
            )
        ]

        self.assertEqual(
            resolve_power_limits(
                stale_values,
                pcmax=1540.0,
                pdmax=1700.0,
            ),
            (1540.0, 1700.0),
        )

    def test_workbook_uses_explicit_asymmetric_limits(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            input_path = Path(directory) / "input.xlsx"
            output_path = Path(directory) / "output.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["time", "dsfunction", "P_ESS", "SOC"])
            worksheet.append(
                [
                    "2025-05-01 00:00:00",
                    repr(
                        [
                            ds_row(-1860.0),
                            ds_row(-1540.0),
                            ds_row(1860.0),
                        ]
                    ),
                    -1540.0,
                    0.5,
                ]
            )
            worksheet.append(
                [
                    "2025-05-01 01:00:00",
                    repr(
                        [
                            ds_row(-1860.0),
                            ds_row(1860.0),
                        ]
                    ),
                    1860.0,
                    0.5,
                ]
            )
            workbook.save(input_path)

            counts = classify_workbook(
                input_path,
                output_path,
                None,
                pcmax=1540.0,
                pdmax=1860.0,
            )

            self.assertEqual(counts["MC"], 1)
            self.assertEqual(counts["MD"], 1)
            classified = load_workbook(
                output_path,
                read_only=True,
                data_only=True,
            )
            try:
                rows = list(classified.active.iter_rows(values_only=True))
            finally:
                classified.close()
            self.assertEqual(rows[0][3], "P_ESS_state")
            self.assertEqual(rows[1][3], "MC")
            self.assertEqual(rows[2][3], "MD")


if __name__ == "__main__":
    unittest.main()
