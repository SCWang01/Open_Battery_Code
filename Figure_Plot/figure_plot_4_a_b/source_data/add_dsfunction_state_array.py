"""Add an hourly dsfunction state interval-width array.

Marginal-boundary (mB) classification is local to the derived width output:
an embedded dsfunction row is mB exactly when its eighth value — the one-hour
post-action SOC appended by Marginal_Check in V5_Case_Study.py — is within
``MB_TOLERANCE`` of either 0 or 1. The worksheet's ``P_ESS_state`` column is
read-only and is never used as an output target.
"""

from __future__ import annotations

import argparse
import ast
from collections import Counter
from copy import copy
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from classify_p_ess import (
    as_float,
    classify_p_ess,
    resolve_power_limits,
    values_match,
)


SOURCE = Path("dsfunction_May2025_exact_V5_k20_classified.xlsx")
STATE_ORDER = ("CC", "CD", "NU", "DC", "DD")
OUTPUT_HEADER = "dsfunction_state_width"
LEGACY_OUTPUT_HEADERS = ("dsfunction_state_share", "dsfunction_state_array")
MB_TOLERANCE = 1e-5
DSFUNCTION_COLUMN_COUNT = 8


def find_column(headers: list[object], header: str) -> int:
    """Return the unique one-based column index for ``header``."""
    matches = [index for index, value in enumerate(headers, start=1) if value == header]
    if len(matches) != 1:
        raise ValueError(
            f"Header {header!r} appears {len(matches)} times; exactly one is required."
        )
    return matches[0]


def format_width(value: float) -> str:
    """Format one interval width compactly while retaining analysis-grade precision."""
    if abs(value) < 5e-13:
        return "0"
    if abs(value - 1.0) < 5e-13:
        return "1"
    return f"{value:.10f}".rstrip("0").rstrip(".")


def parse_dsfunction(dsfunction_value: object, excel_row: int) -> Sequence[Sequence[object]]:
    """Parse one dsfunction cell and require eight values per embedded row."""
    if not isinstance(dsfunction_value, str):
        raise ValueError(f"Row {excel_row}: dsfunction is not text.")
    try:
        dsfunction = ast.literal_eval(dsfunction_value)
    except (SyntaxError, ValueError) as exc:
        raise ValueError(f"Row {excel_row}: dsfunction is not a valid literal.") from exc

    if not isinstance(dsfunction, (list, tuple)) or not dsfunction:
        raise ValueError(f"Row {excel_row}: dsfunction must be a non-empty 2-D array.")
    for ds_row_number, ds_row in enumerate(dsfunction, start=1):
        if not isinstance(ds_row, (list, tuple)) or len(ds_row) != DSFUNCTION_COLUMN_COUNT:
            raise ValueError(
                f"Row {excel_row}: dsfunction segment {ds_row_number} must contain "
                f"exactly {DSFUNCTION_COLUMN_COUNT} values."
            )
    return dsfunction


def is_marginal_boundary(eighth: float) -> bool:
    """Return whether column 8 is strictly within tolerance of 0 or 1."""
    return abs(eighth) < MB_TOLERANCE or abs(eighth - 1.0) < MB_TOLERANCE


def state_width_array(
    dsfunction_value: object,
    excel_row: int,
    *,
    pcmax: float,
    pdmax: float,
) -> tuple[str, list[str], list[float]]:
    """Classify embedded rows and return widths in ``STATE_ORDER``.

    mB rows remain recorded in the returned state labels but intentionally do
    not occupy a slot or contribute width to the five-element output array.
    """
    dsfunction = parse_dsfunction(dsfunction_value, excel_row)
    states: list[str] = []
    widths = Counter({state: 0.0 for state in STATE_ORDER})

    for ds_row_number, ds_row in enumerate(dsfunction, start=1):
        values = [
            as_float(
                value,
                name=f"dsfunction segment {ds_row_number}, column {column_number}",
                excel_row=excel_row,
            )
            for column_number, value in enumerate(ds_row, start=1)
        ]
        lower, upper, third, fourth, fifth, _sixth, seventh, eighth = values

        if is_marginal_boundary(eighth):
            state = "mB"
        else:
            state = classify_p_ess(
                third,
                ds_row,
                excel_row,
                pcmax=pcmax,
                pdmax=pdmax,
            )
            if values_match(seventh, 0.0) and values_match(fourth - fifth, 0.0):
                if third > 0:
                    state = "DD"
                elif third < 0:
                    state = "CC"

        width = upper - lower
        if width <= 0:
            raise ValueError(
                f"Row {excel_row}: dsfunction segment {ds_row_number} must have a "
                "positive interval width."
            )
        states.append(state)
        if state in STATE_ORDER:
            widths[state] += width

    state_widths = [widths[state] for state in STATE_ORDER]
    array = "[" + ",".join(format_width(width) for width in state_widths) + "]"
    return array, states, state_widths


def build_parser() -> argparse.ArgumentParser:
    """Build command-line options for explicit or inferred power limits."""
    parser = argparse.ArgumentParser(
        description="Add dsfunction state widths using asymmetric power limits."
    )
    parser.add_argument(
        "--pcmax",
        type=float,
        help="Positive controlled-unit charge limit (default: infer globally)",
    )
    parser.add_argument(
        "--pdmax",
        type=float,
        help="Positive controlled-unit discharge limit (default: infer globally)",
    )
    return parser


def main(*, pcmax: float | None = None, pdmax: float | None = None) -> None:
    """Update only the derived width column in ``SOURCE``."""
    if not SOURCE.is_file():
        raise FileNotFoundError(f"Input file not found: {SOURCE}")

    workbook = load_workbook(SOURCE, data_only=False)
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    dsfunction_col = find_column(headers, "dsfunction")
    resolved_pcmax, resolved_pdmax = resolve_power_limits(
        (
            (worksheet.cell(row, dsfunction_col).value, row)
            for row in range(2, worksheet.max_row + 1)
        ),
        pcmax=pcmax,
        pdmax=pdmax,
    )

    # Column 6 is the established derived-output position. P_ESS_state remains
    # untouched in its existing column.
    output_col = 6
    existing_headers = [
        index
        for index, value in enumerate(headers, start=1)
        if value in (OUTPUT_HEADER, *LEGACY_OUTPUT_HEADERS)
    ]
    if existing_headers and existing_headers != [output_col]:
        raise ValueError(
            f"Existing {OUTPUT_HEADER!r} column is not in column {output_col}; "
            "cannot update safely."
        )
    if worksheet.max_column >= output_col:
        current_header = worksheet.cell(1, output_col).value
        if current_header not in (None, OUTPUT_HEADER, *LEGACY_OUTPUT_HEADERS):
            worksheet.insert_cols(output_col, amount=1)

    style_source_col = 4 if worksheet.max_column >= 4 else 3
    for excel_row in range(1, worksheet.max_row + 1):
        source_cell = worksheet.cell(excel_row, style_source_col)
        target_cell = worksheet.cell(excel_row, output_col)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.hyperlink:
            target_cell.hyperlink = copy(source_cell.hyperlink)

    worksheet.cell(1, output_col).value = OUTPUT_HEADER
    output_letter = get_column_letter(output_col)
    worksheet.column_dimensions[output_letter].width = 60

    segment_state_counts: Counter[str] = Counter()
    mb_rows: list[int] = []
    for excel_row in range(2, worksheet.max_row + 1):
        value, states, _ = state_width_array(
            worksheet.cell(excel_row, dsfunction_col).value,
            excel_row,
            pcmax=resolved_pcmax,
            pdmax=resolved_pdmax,
        )
        cell = worksheet.cell(excel_row, output_col)
        cell.value = value
        cell.number_format = "@"
        segment_state_counts.update(states)
        if "mB" in states:
            mb_rows.append(excel_row)

    temporary = SOURCE.with_name(f"{SOURCE.stem}.tmp{SOURCE.suffix}")
    workbook.save(temporary)
    temporary.replace(SOURCE)

    print(f"Updated: {SOURCE}")
    print(f"Records: {worksheet.max_row - 1}")
    print(f"Power limits: Pcmax={resolved_pcmax:g}, Pdmax={resolved_pdmax:g}")
    print(f"Output state order: {STATE_ORDER}")
    print(f"Segment state counts: {dict(segment_state_counts)}")
    print(f"Worksheet rows containing mB: {mb_rows}")


if __name__ == "__main__":
    try:
        args = build_parser().parse_args()
        main(pcmax=args.pcmax, pdmax=args.pdmax)
    except Exception as exc:
        raise SystemExit(f"Error: {exc}") from None
