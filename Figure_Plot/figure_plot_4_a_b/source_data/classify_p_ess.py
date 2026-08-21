"""Classify P_ESS values using the row-specific dsfunction array.

The script keeps the source workbook unchanged and writes a classified copy.
"""

from __future__ import annotations

import argparse
import ast
import math
from collections import Counter
from copy import copy
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DEFAULT_INPUT = Path("dsfunction_May2025_exact_V5_k20.xlsx")
STATUS_HEADER = "P_ESS_state"
REL_TOL = 1e-9
ABS_TOL = 1e-6
MAX_RATE_ABS_TOL = 5e-4
SUMMARY_ORDER = ("MC", "MD", "NU", "DC", "DD", "CC", "CD", "NC")


def values_match(left: float, right: float) -> bool:
    """Compare numeric values while allowing small spreadsheet round-off."""
    return math.isclose(left, right, rel_tol=REL_TOL, abs_tol=ABS_TOL)


def max_rate_values_match(left: float, right: float) -> bool:
    """Compare power limits after the upstream four-decimal rounding."""
    return math.isclose(
        left,
        right,
        rel_tol=REL_TOL,
        abs_tol=MAX_RATE_ABS_TOL,
    )


def parse_dsfunction(value: Any, excel_row: int) -> Sequence[Sequence[Any]]:
    """Parse one dsfunction cell with 7 core columns and optional SOC."""
    if isinstance(value, str):
        try:
            value = ast.literal_eval(value)
        except (SyntaxError, ValueError) as exc:
            raise ValueError(
                f"Row {excel_row} in the worksheet: dsfunction cannot be parsed: {exc}"
            ) from exc

    if not isinstance(value, (list, tuple)) or not value:
        raise ValueError(f"Row {excel_row}: dsfunction must be a non-empty 2-D array.")

    for ds_row_number, ds_row in enumerate(value, start=1):
        if not isinstance(ds_row, (list, tuple)) or len(ds_row) not in (7, 8):
            raise ValueError(
                f"Row {excel_row}: dsfunction segment {ds_row_number} "
                "must contain 7 core values and may contain an eighth SOC value."
            )
    return value


def as_float(value: Any, *, name: str, excel_row: int) -> float:
    """Convert a cell or dsfunction element to a finite float."""
    try:
        result = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Row {excel_row}: {name} is not a valid number: {value!r}") from exc

    if not math.isfinite(result):
        raise ValueError(f"Row {excel_row}: {name} is not a finite number: {value!r}")
    return result


def validate_power_limit(value: Any, *, name: str) -> float:
    """Normalize one positive Pcmax/Pdmax magnitude."""
    result = as_float(value, name=name, excel_row=1)
    if result <= 0:
        raise ValueError(f"{name} must be a strictly positive magnitude.")
    return result


def resolve_power_limits(
    dsfunction_values: Iterable[tuple[Any, int]],
    *,
    pcmax: Any | None = None,
    pdmax: Any | None = None,
) -> tuple[float, float]:
    """Resolve positive Pcmax/Pdmax from overrides or workbook-wide candidates."""
    resolved_pcmax = (
        validate_power_limit(pcmax, name="Pcmax") if pcmax is not None else None
    )
    resolved_pdmax = (
        validate_power_limit(pdmax, name="Pdmax") if pdmax is not None else None
    )
    if resolved_pcmax is not None and resolved_pdmax is not None:
        return resolved_pcmax, resolved_pdmax

    minimum_power: float | None = None
    maximum_power: float | None = None
    for dsfunction_value, excel_row in dsfunction_values:
        dsfunction = parse_dsfunction(dsfunction_value, excel_row)
        for ds_row_number, ds_row in enumerate(dsfunction, start=1):
            power = as_float(
                ds_row[2],
                name=f"dsfunction segment {ds_row_number}, column 3",
                excel_row=excel_row,
            )
            minimum_power = (
                power if minimum_power is None else min(minimum_power, power)
            )
            maximum_power = (
                power if maximum_power is None else max(maximum_power, power)
            )

    if resolved_pcmax is None:
        if minimum_power is None or minimum_power >= 0:
            raise ValueError(
                "Cannot infer Pcmax: dsfunction column 3 has no negative power."
            )
        resolved_pcmax = validate_power_limit(
            -minimum_power,
            name="inferred Pcmax",
        )
    if resolved_pdmax is None:
        if maximum_power is None or maximum_power <= 0:
            raise ValueError(
                "Cannot infer Pdmax: dsfunction column 3 has no positive power."
            )
        resolved_pdmax = validate_power_limit(
            maximum_power,
            name="inferred Pdmax",
        )

    return resolved_pcmax, resolved_pdmax


def classify_p_ess(
    p_ess_value: Any,
    ds_row: Sequence[Any],
    excel_row: int,
    *,
    pcmax: float,
    pdmax: float,
) -> str:
    """Return the state for one dsfunction segment.

    The caller supplies the segment being classified.  This is important
    because column 3 may legitimately contain the same P_ESS value in more
    than one segment; classification must use each segment's own columns 4-7.
    """
    p_ess = as_float(p_ess_value, name="P_ESS", excel_row=excel_row)

    if max_rate_values_match(p_ess, pdmax):
        return "MD"
    if max_rate_values_match(p_ess, -pcmax):
        return "MC"
    if values_match(p_ess, 0.0):
        return "NU"

    seventh = as_float(
        ds_row[6],
        name="dsfunction segment column 7",
        excel_row=excel_row,
    )

    if values_match(seventh, 1.0):
        return "DC" if p_ess > 0 else "CC"
    if values_match(seventh, -1.0):
        return "DD" if p_ess > 0 else "CD"

    if values_match(seventh, 0.0):
        fourth = as_float(
            ds_row[3],
            name="dsfunction segment column 4",
            excel_row=excel_row,
        )
        fifth = as_float(
            ds_row[4],
            name="dsfunction segment column 5",
            excel_row=excel_row,
        )

        if values_match(fourth, fifth):
            return "DD" if p_ess > 0 else "CC"
        if p_ess > 0:
            return "DC" if fourth > fifth else "DD"
        return "CC" if fourth > fifth else "CD"

    return "NC"


def classify_p_ess_from_dsfunction(
    p_ess_value: Any,
    dsfunction_value: Any,
    excel_row: int,
    *,
    pcmax: float,
    pdmax: float,
) -> str:
    """Match P_ESS to a dsfunction segment, then classify by its first 7 columns."""
    p_ess = as_float(p_ess_value, name="P_ESS", excel_row=excel_row)

    # These three states do not depend on a dsfunction segment.  In particular,
    # their candidate power may occur more than once in the table.
    if (
        max_rate_values_match(p_ess, pdmax)
        or max_rate_values_match(p_ess, -pcmax)
        or values_match(p_ess, 0.0)
    ):
        return classify_p_ess(
            p_ess,
            (),
            excel_row,
            pcmax=pcmax,
            pdmax=pdmax,
        )

    dsfunction = parse_dsfunction(dsfunction_value, excel_row)
    matches = []
    for ds_row_number, ds_row in enumerate(dsfunction, start=1):
        candidate = as_float(
            ds_row[2],
            name=f"dsfunction segment {ds_row_number}, column 3",
            excel_row=excel_row,
        )
        if values_match(p_ess, candidate):
            matches.append((ds_row_number, ds_row))

    if len(matches) != 1:
        raise ValueError(
            f"Row {excel_row}: P_ESS={p_ess:g} matched {len(matches)} rows in "
            "dsfunction column 3; exactly 1 match expected."
        )

    return classify_p_ess(
        p_ess,
        matches[0][1],
        excel_row,
        pcmax=pcmax,
        pdmax=pdmax,
    )


def find_column(headers: list[Any], header_name: str) -> int:
    """Find a unique 1-based worksheet column by its row-1 header."""
    matches = [index for index, value in enumerate(headers, start=1) if value == header_name]
    if len(matches) != 1:
        raise ValueError(
            f"Worksheet header {header_name!r} appears {len(matches)} times; "
            f"exactly one is required."
        )
    return matches[0]


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_classified{input_path.suffix}")


def classify_workbook(
    input_path: Path,
    output_path: Path,
    sheet_name: str | None,
    *,
    pcmax: float | None = None,
    pdmax: float | None = None,
) -> Counter:
    """Add/update P_ESS_state and save a classified workbook copy."""
    if input_path.resolve() == output_path.resolve():
        raise ValueError("Output file must differ from the input file to avoid overwriting the source data.")
    if not input_path.is_file():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    workbook = load_workbook(input_path, data_only=False)
    if sheet_name is None:
        worksheet = workbook.active
    elif sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        raise ValueError(
            f"Worksheet {sheet_name!r} not found; available worksheets: "
            f"{', '.join(workbook.sheetnames)}"
        )

    headers = [cell.value for cell in worksheet[1]]
    dsfunction_col = find_column(headers, "dsfunction")
    p_ess_col = find_column(headers, "P_ESS")
    resolved_pcmax, resolved_pdmax = resolve_power_limits(
        (
            (worksheet.cell(row, dsfunction_col).value, row)
            for row in range(2, worksheet.max_row + 1)
        ),
        pcmax=pcmax,
        pdmax=pdmax,
    )
    status_col = p_ess_col + 1

    existing_status_columns = [
        index for index, value in enumerate(headers, start=1) if value == STATUS_HEADER
    ]
    if existing_status_columns:
        if existing_status_columns != [status_col]:
            raise ValueError(
                f"Existing {STATUS_HEADER!r} column is not immediately to the "
                f"right of P_ESS; cannot update safely."
            )
    else:
        worksheet.insert_cols(status_col, amount=1)

    # Match the adjacent P_ESS column's style without altering source data.
    for excel_row in range(1, worksheet.max_row + 1):
        source_cell = worksheet.cell(row=excel_row, column=p_ess_col)
        target_cell = worksheet.cell(row=excel_row, column=status_col)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.hyperlink:
            target_cell.hyperlink = copy(source_cell.hyperlink)

    source_letter = get_column_letter(p_ess_col)
    status_letter = get_column_letter(status_col)
    source_width = worksheet.column_dimensions[source_letter].width or 13.0
    worksheet.column_dimensions[status_letter].width = max(
        source_width, float(len(STATUS_HEADER) + 2)
    )
    worksheet.cell(row=1, column=status_col).value = STATUS_HEADER

    counts: Counter = Counter()
    for excel_row in range(2, worksheet.max_row + 1):
        state = classify_p_ess_from_dsfunction(
            worksheet.cell(row=excel_row, column=p_ess_col).value,
            worksheet.cell(row=excel_row, column=dsfunction_col).value,
            excel_row,
            pcmax=resolved_pcmax,
            pdmax=resolved_pdmax,
        )
        target_cell = worksheet.cell(row=excel_row, column=status_col)
        target_cell.value = state
        target_cell.number_format = "@"
        counts[state] += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return counts


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Classify each row's P_ESS using its dsfunction array and "
        "write P_ESS_state to the right of P_ESS."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"Input Excel workbook (default: {DEFAULT_INPUT})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Output Excel workbook (default: input name suffixed with _classified)",
    )
    parser.add_argument(
        "--sheet",
        help="Worksheet to process (default: active worksheet)",
    )
    parser.add_argument(
        "--pcmax",
        type=float,
        help="Positive controlled-unit charge limit (default: infer globally)",
    )
    parser.add_argument(
        "--pdmax",
        type=float,
        help="Positive controlled-unit discharge limit (default: infer globally)",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    output_path = args.output or default_output_path(args.input)
    counts = classify_workbook(
        args.input,
        output_path,
        args.sheet,
        pcmax=args.pcmax,
        pdmax=args.pdmax,
    )

    print(f"Done: {output_path}")
    print(f"Total records: {sum(counts.values())}")
    for state in SUMMARY_ORDER:
        print(f"P_ESS_state={state}: {counts[state]}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        raise SystemExit(f"Error: {exc}") from None
