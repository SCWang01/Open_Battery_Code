"""Create the May 2025 dsfunction/P_ESS state-distribution pie figure.

All deliverables are written beside this script so that the complete folder can
be moved as one reproducible figure package.  Panel (a) sums the hourly raw
``dsfunction_state_width`` arrays vertically and normalizes CC/CD/NU/DC/DD to
100%.  Panel (b) counts the hourly ``P_ESS_state`` labels.
"""

from __future__ import annotations

import ast
import csv
import math
from collections import Counter
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from openpyxl import load_workbook


PACKAGE_ROOT = Path(__file__).resolve().parent.parent
SOURCE_WORKBOOK = (
    PACKAGE_ROOT / "source_data" / "dsfunction_May2025_exact_V5_k20_classified.xlsx"
)
SOURCE_SHEET = "May2025"
OUTPUT_DIR = Path(__file__).resolve().parent
OUTPUT_BASENAME = "state_distribution_pies_May2025"
SOURCE_DATA_NAME = f"{OUTPUT_BASENAME}_source_data.csv"

P_ESS_ORDER = ("MC", "CC", "CD", "NU", "DC", "DD", "MD", "NC")
LEFT_ORDER = ("CC", "CD", "NU", "DC", "DD")
LEGEND_ORDER = ("MC", "CC", "CD", "NU", "DC", "DD", "MD")

LEGEND_LABELS = {
    "MC": "Max-rate charge",
    "CC": "Charge-for-charge",
    "CD": "Charge-for-discharge",
    "NU": "Null segment",
    "DC": "Discharge-for-charge",
    "DD": "Discharge-for-discharge",
    "MD": "Max-rate discharge",
}

COLORS = {
    "MC": "#C90000",
    "CC": "#FF6B6B",
    "CD": "#F6B4B4",
    "NU": "#A6A6A6",
    "DC": "#9CC5E5",
    "DD": "#4F95D5",
    "MD": "#1D4B73",
}

FIGURE_WIDTH_MM = 180.0
FIGURE_HEIGHT_MM = 70.0
PNG_DPI = 600


def configure_matplotlib() -> None:
    """Apply the agreed publication typography and editable-vector settings."""
    mpl.rcParams.update(
        {
            "font.family": ["Times New Roman", "Arial", "DejaVu Serif"],
            "font.serif": ["Times New Roman", "Times", "DejaVu Serif"],
            "font.size": 9,
            "legend.frameon": False,
            "svg.fonttype": "none",
            "pdf.fonttype": 42,
            "axes.unicode_minus": False,
        }
    )


def find_unique_header(headers: tuple[object, ...], name: str) -> int:
    """Return a zero-based column index for a unique worksheet header."""
    matches = [index for index, value in enumerate(headers) if value == name]
    if len(matches) != 1:
        raise ValueError(
            f"Header {name!r} occurs {len(matches)} times; exactly one is required."
        )
    return matches[0]


def read_source_data() -> tuple[Counter[str], dict[str, float], int]:
    """Read and validate the two distributions from the source workbook."""
    if not SOURCE_WORKBOOK.is_file():
        raise FileNotFoundError(f"Source workbook not found: {SOURCE_WORKBOOK}")

    workbook = load_workbook(SOURCE_WORKBOOK, read_only=True, data_only=True)
    try:
        if SOURCE_SHEET not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet {SOURCE_SHEET!r} not found; available sheets: "
                f"{', '.join(workbook.sheetnames)}"
            )
        worksheet = workbook[SOURCE_SHEET]
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows)
        state_col = find_unique_header(headers, "P_ESS_state")
        width_col = find_unique_header(headers, "dsfunction_state_width")

        p_ess_counts: Counter[str] = Counter()
        width_sums = {state: 0.0 for state in LEFT_ORDER}
        record_count = 0

        for excel_row, row in enumerate(rows, start=2):
            if all(value is None for value in row):
                continue

            p_ess_state = row[state_col]
            if p_ess_state not in P_ESS_ORDER:
                raise ValueError(
                    f"Excel row {excel_row}: unknown P_ESS_state {p_ess_state!r}."
                )
            p_ess_counts[p_ess_state] += 1

            raw_array = row[width_col]
            if not isinstance(raw_array, str):
                raise ValueError(
                    f"Excel row {excel_row}: dsfunction_state_width must be text."
                )
            try:
                widths = ast.literal_eval(raw_array)
            except (SyntaxError, ValueError) as exc:
                raise ValueError(
                    f"Excel row {excel_row}: invalid dsfunction_state_width: {exc}"
                ) from exc

            if not isinstance(widths, (list, tuple)) or len(widths) != len(LEFT_ORDER):
                raise ValueError(
                    f"Excel row {excel_row}: expected {len(LEFT_ORDER)} widths."
                )
            numeric_widths = [float(value) for value in widths]
            if not all(math.isfinite(value) and value >= 0 for value in numeric_widths):
                raise ValueError(
                    f"Excel row {excel_row}: widths must be finite and non-negative."
                )

            for state, value in zip(LEFT_ORDER, numeric_widths):
                width_sums[state] += value
            record_count += 1
    finally:
        workbook.close()

    if record_count != 744:
        raise ValueError(f"Expected 744 May 2025 records; found {record_count}.")
    if sum(p_ess_counts.values()) != record_count:
        raise ValueError("P_ESS counts do not reconcile to the worksheet record count.")
    if math.isclose(
        sum(width_sums.values()), 0.0, rel_tol=0, abs_tol=1e-12
    ):
        raise ValueError("The five plotted dsfunction width totals are all zero.")
    if p_ess_counts["NC"] != 0:
        raise ValueError("NC occurs in P_ESS_state and cannot be silently omitted.")

    return p_ess_counts, width_sums, record_count


def source_rows(
    p_ess_counts: Counter[str], width_sums: dict[str, float], record_count: int
) -> list[dict[str, object]]:
    """Build one auditable table containing raw totals and plotted percentages."""
    left_total = sum(width_sums.values())
    rows: list[dict[str, object]] = []
    for state in P_ESS_ORDER:
        in_left = state in LEFT_ORDER
        in_legend = state in LEGEND_ORDER
        rows.append(
            {
                "state": state,
                "legend_label": LEGEND_LABELS.get(state, "not classified"),
                "color_hex": COLORS.get(state, ""),
                "dsfunction_vertical_width_sum": (
                    f"{width_sums[state]:.12f}" if in_left else ""
                ),
                "included_in_left_pie": str(in_left).lower(),
                "dsfunction_left_normalized_percent": (
                    f"{100.0 * width_sums[state] / left_total:.8f}" if in_left else ""
                ),
                "p_ess_count": p_ess_counts[state],
                "p_ess_percent": f"{100.0 * p_ess_counts[state] / record_count:.8f}",
                "included_in_shared_legend": str(in_legend).lower(),
            }
        )
    return rows


def write_source_csv(rows: list[dict[str, object]]) -> Path:
    """Write the exact figure-source table beside all other deliverables."""
    output_path = OUTPUT_DIR / SOURCE_DATA_NAME
    fieldnames = list(rows[0])
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    return output_path


def percent_text(percent: float) -> str:
    return f"{percent:.2f}%"


def draw_left_pie(ax: plt.Axes, width_sums: dict[str, float]) -> None:
    values = [width_sums[state] for state in LEFT_ORDER]
    ax.pie(
        values,
        colors=[COLORS[state] for state in LEFT_ORDER],
        startangle=90,
        counterclock=False,
        radius=0.88,
        autopct=percent_text,
        pctdistance=0.68,
        textprops={"fontsize": 9, "color": "#222222"},
        wedgeprops={"linewidth": 0, "edgecolor": "none"},
        normalize=True,
    )
    ax.set_aspect("equal")
    ax.set_axis_off()


def draw_right_pie(
    ax: plt.Axes, p_ess_counts: Counter[str], record_count: int
) -> None:
    radius = 0.82
    values = [p_ess_counts[state] for state in LEGEND_ORDER]
    wedges, _ = ax.pie(
        values,
        colors=[COLORS[state] for state in LEGEND_ORDER],
        startangle=90,
        counterclock=False,
        radius=radius,
        wedgeprops={"linewidth": 0, "edgecolor": "none"},
        normalize=True,
    )

    outside_positions = {
        "CC": (1.10, -0.46),
        "CD": (1.16, -0.75),
        "DC": (-1.16, -0.75),
        "DD": (-1.10, -0.46),
    }

    for state, wedge in zip(LEGEND_ORDER, wedges):
        percent = 100.0 * p_ess_counts[state] / record_count
        angle = math.radians((wedge.theta1 + wedge.theta2) / 2.0)
        if state in {"MC", "NU", "MD"}:
            label_radius = 0.48 if state == "NU" else 0.52
            ax.text(
                label_radius * math.cos(angle),
                label_radius * math.sin(angle),
                percent_text(percent),
                ha="center",
                va="center",
                fontsize=9,
                color="#222222" if state == "NU" else "#FFFFFF",
            )
            continue

        text_x, text_y = outside_positions[state]
        side = 1 if text_x > 0 else -1
        anchor_x = radius * 0.98 * math.cos(angle)
        anchor_y = radius * 0.98 * math.sin(angle)
        elbow_x = radius * 1.10 * math.cos(angle)
        elbow_y = radius * 1.10 * math.sin(angle)
        line_end_x = text_x - side * 0.035
        ax.plot(
            [anchor_x, elbow_x, line_end_x],
            [anchor_y, elbow_y, text_y],
            color="#B7B7B7",
            linewidth=0.65,
            solid_capstyle="round",
            clip_on=False,
            zorder=1,
        )
        ax.text(
            text_x,
            text_y,
            percent_text(percent),
            ha="left" if side > 0 else "right",
            va="center",
            fontsize=9,
            color="#222222",
            clip_on=False,
            zorder=2,
        )

    # Use asymmetric horizontal limits to reserve space for the longer labels
    # on the right without shrinking the pie or changing the 180 mm canvas.
    ax.set_xlim(-1.36, 1.58)
    ax.set_ylim(-1.07, 1.07)
    ax.set_aspect("equal")
    ax.set_axis_off()


def add_aligned_panel_labels(
    fig: plt.Figure, left_ax: plt.Axes, right_ax: plt.Axes
) -> None:
    """Place both panel labels on one exact figure-coordinate baseline."""
    inverse_figure = fig.transFigure.inverted()
    left_center_x = inverse_figure.transform(left_ax.transData.transform((0, 0)))[0]
    right_center_x = inverse_figure.transform(right_ax.transData.transform((0, 0)))[0]
    panel_y = 0.055


def draw_legend(ax: plt.Axes) -> None:
    handles = [
        Line2D(
            [0],
            [0],
            marker="s",
            linestyle="None",
            markersize=7.2,
            markerfacecolor=COLORS[state],
            markeredgecolor="none",
        )
        for state in LEGEND_ORDER
    ]
    labels = [LEGEND_LABELS[state] for state in LEGEND_ORDER]
    ax.legend(
        handles,
        labels,
        loc="center",
        fontsize=9,
        handlelength=0.8,
        handletextpad=0.45,
        labelspacing=0.53,
        borderaxespad=0,
        frameon=False,
    )
    ax.set_axis_off()


def build_figure(
    p_ess_counts: Counter[str], width_sums: dict[str, float], record_count: int
) -> plt.Figure:
    configure_matplotlib()
    mm_per_inch = 25.4
    fig = plt.figure(
        figsize=(FIGURE_WIDTH_MM / mm_per_inch, FIGURE_HEIGHT_MM / mm_per_inch),
        facecolor="none",
    )
    grid = fig.add_gridspec(
        1,
        3,
        width_ratios=(1.10, 1.03, 1.30),
        left=0.018,
        right=0.988,
        bottom=0.135,
        top=0.985,
        wspace=0.015,
    )
    left_ax = fig.add_subplot(grid[0, 0])
    legend_ax = fig.add_subplot(grid[0, 1])
    right_ax = fig.add_subplot(grid[0, 2])

    draw_left_pie(left_ax, width_sums)
    draw_legend(legend_ax)
    draw_right_pie(right_ax, p_ess_counts, record_count)
    add_aligned_panel_labels(fig, left_ax, right_ax)
    return fig


def save_figure(fig: plt.Figure) -> list[Path]:
    """Export editable vectors and a high-resolution transparent preview."""
    base = OUTPUT_DIR / OUTPUT_BASENAME
    svg_path = base.with_suffix(".svg")
    pdf_path = base.with_suffix(".pdf")
    png_path = base.with_suffix(".png")

    fig.savefig(
        svg_path,
        transparent=True,
        metadata={"Creator": "Matplotlib", "Description": "May 2025 state distributions"},
    )
    fig.savefig(
        pdf_path,
        transparent=True,
        metadata={"Creator": "Matplotlib", "Title": "May 2025 state distributions"},
    )
    fig.savefig(
        png_path,
        dpi=PNG_DPI,
        transparent=True,
        metadata={"Title": "May 2025 state distributions"},
    )
    return [svg_path, pdf_path, png_path]


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    p_ess_counts, width_sums, record_count = read_source_data()
    rows = source_rows(p_ess_counts, width_sums, record_count)
    csv_path = write_source_csv(rows)
    fig = build_figure(p_ess_counts, width_sums, record_count)
    figure_paths = save_figure(fig)
    plt.close(fig)

    left_total = sum(width_sums.values())
    print(f"Source workbook: {SOURCE_WORKBOOK}")
    print(f"Records: {record_count}")
    print(
        "Panel (a): "
        + ", ".join(
            f"{state}={100.0 * width_sums[state] / left_total:.2f}%"
            for state in LEFT_ORDER
        )
    )
    print(
        "Panel (b): "
        + ", ".join(
            f"{state}={100.0 * p_ess_counts[state] / record_count:.2f}%"
            for state in LEGEND_ORDER
        )
    )
    print(f"Source data: {csv_path}")
    for path in figure_paths:
        print(f"Figure: {path}")


if __name__ == "__main__":
    main()
