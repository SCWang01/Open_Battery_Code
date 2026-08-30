"""Create dsfunction/P_ESS state-distribution pie figures from an arbitrary
selection of calendar months across 2023, 2024 and 2025.

Month selection is driven by the ``SELECTION_MATRIX`` constant defined at the
top of this file: 12 rows (January..December) x 3 columns (the years 2023,
2024, 2025).  A cell set to ``1`` includes that (month, year) in the
aggregation; ``0`` skips it.  Edit the matrix and run ``python <script>`` --
no command-line arguments are used.  This lets you request non-contiguous
month sets such as "Q1 of each of the three years" (rows 1-3 all ones, the
rest zeros) which a single continuous --start/--end range cannot express.
When the selected months match a meteorological season, the season name is
also added to the output basename, e.g. ``..._2023_3_2024_3_2025_3_Winter``.

Panel (a) sums the hourly raw ``dsfunction_state_width`` arrays vertically
across every selected month and normalizes CC/CD/NU/DC/DD to 100%.  Panel (b)
counts the hourly ``P_ESS_state`` labels across every selected month.  No
per-month averaging is performed: hours are pooled directly, so longer months
naturally contribute more.

Source workbooks are read from ``Results/Bidding`` (resolved relative to this
script's location: ``<checkout>/Results/Bidding``).  Each selected month must
provide a ``dsfunction_<MonthName><Year>_exact_V5_k20_classified_width.xlsx``
workbook whose single sheet is named ``<MonthName><Year>``.  If a selected
month's file is absent (or malformed) the script stops with a clear error
naming the expected path, so the missing file can be supplied.

All deliverables (SVG, PDF, PNG and the auditable source-data CSV) are written
into the ``Figs`` folder beside this script.  The basename encodes the
per-year selection counts as ``2023_A_2024_B_2025_C``, where A/B/C are the
number of selected months in each year (zero is shown, not omitted).  Because
the counts do not record *which* months were selected, two different matrices
that yield the same A/B/C produce the same basename and the later run
overwrites the earlier -- this is intentional for the "edit matrix, re-run"
workflow.  Recognized seasonal selections receive an additional suffix such
as ``_Winter`` or ``_Spring``.  Rename by hand if you need to keep two
otherwise-identical selections.

Usage:
    python plot_state_distribution_pies_May2025.py
"""

from __future__ import annotations

import ast
import calendar
import csv
import math
from collections import Counter
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
BIDDING_DIR = PROJECT_ROOT / "Results" / "Bidding"
OUTPUT_DIR = Path(__file__).resolve().parent / "Figs"

# Years covered by the selection matrix, left to right.
SELECTION_YEARS = (2023, 2024, 2025)

# 12 rows (January..December) x 3 columns (2023, 2024, 2025).
# Set a cell to 1 to include that (month, year) in the aggregation, 0 to skip.
# To plot, e.g., Q1 of all three years, set rows 0-2 to all ones and the rest
# to zeros.  The default matrix below selects all 36 months.
SELECTION_MATRIX: list[list[int]] = [
    [0, 0, 1],  # January
    [0, 0, 1],  # February
    [0, 0, 1],  # March
    [0, 0, 1],  # April
    [0, 0, 1],  # May
    [0, 0, 1],  # June
    [0, 0, 1],  # July
    [0, 0, 1],  # August
    [0, 0, 1],  # September
    [0, 0, 1],  # October
    [0, 0, 1],  # November
    [0, 0, 1],  # December
]

# Month sets used for automatic filename labels.  A season is recognized when
# every year has exactly the corresponding three selected months.  For a
# non-standard or cross-year selection, the count-based basename is retained.
SEASON_MONTHS = {
    "Winter": frozenset((12, 1, 2)),
    "Spring": frozenset((3, 4, 5)),
    "Summer": frozenset((6, 7, 8)),
    "Autumn": frozenset((9, 10, 11)),
}

# When True, one run creates all four seasonal figures.  Set to False to use
# the full SELECTION_MATRIX.  With the default all-one matrix, this creates
# one figure pooled across all 36 months from 2023 through 2025.
PLOT_SEASONS = False

P_ESS_ORDER = ("MC", "CC", "CD", "NU", "DC", "DD", "MD", "NC")
LEFT_ORDER = ("CC", "CD", "NU", "DC", "DD")
LEGEND_ORDER = ("MC", "CC", "CD", "NU", "DC", "DD", "MD")

LEGEND_LABELS = {
    "MC": "Max-rate charge",
    "CC": "Charge-for-charge",
    "CD": "Charge-for-discharge",
    "NU": "Null segment",
    "DC": "Discharge-for-charge",
    "DD": "Discharge-for-discharge",
    "MD": "Max-rate discharge",
}

COLORS = {
    "MC": "#C90000",
    "CC": "#FF6B6B",
    "CD": "#F6B4B4",
    "NU": "#A6A6A6",
    "DC": "#9CC5E5",
    "DD": "#4F95D5",
    "MD": "#1D4B73",
}

FIGURE_WIDTH_MM = 180.0
FIGURE_HEIGHT_MM = 70.0
PNG_DPI = 600


# --------------------------------------------------------------------------- #
# Month selection
# --------------------------------------------------------------------------- #
def select_months(
    matrix: list[list[int]] = SELECTION_MATRIX,
    years: tuple[int, ...] = SELECTION_YEARS,
) -> list[tuple[int, int]]:
    """Expand the selection matrix into an ordered list of (year, month) tuples.

    Rows are months 1..12 (January..December); columns are the years in
    ``years``.  A cell equal to 1 includes that (month, year).  The matrix
    shape is validated: exactly 12 rows, each with one entry per year, and
    each entry 0 or 1.  At least one month must be selected.
    """
    if len(matrix) != 12:
        raise ValueError(
            f"SELECTION_MATRIX must have 12 rows (Jan..Dec); found {len(matrix)}."
        )
    for row_index, row in enumerate(matrix):
        if len(row) != len(years):
            raise ValueError(
                f"SELECTION_MATRIX row {row_index + 1} ({calendar.month_name[row_index + 1]}): "
                f"expected {len(years)} columns, found {len(row)}."
            )
        for col_index, value in enumerate(row):
            if value not in (0, 1):
                raise ValueError(
                    f"SELECTION_MATRIX row {row_index + 1} "
                    f"({calendar.month_name[row_index + 1]}), column "
                    f"{years[col_index]}: entries must be 0 or 1, found {value!r}."
                )

    months = [
        (years[col], month)
        for month in range(1, 13)
        for col in range(len(years))
        if matrix[month - 1][col] == 1
    ]
    if not months:
        raise ValueError("SELECTION_MATRIX selects no months; set at least one cell to 1.")
    return months


def per_year_counts(
    months: list[tuple[int, int]], years: tuple[int, ...] = SELECTION_YEARS
) -> dict[int, int]:
    """Count how many selected months fall in each year (0 for years with none)."""
    counts = {year: 0 for year in years}
    for year, _month in months:
        counts[year] += 1
    return counts


def month_name(year: int, month: int) -> str:
    """Worksheet/sheet name for a month, e.g. ``May2024``."""
    return f"{calendar.month_name[month]}{year}"


def workbook_path(year: int, month: int) -> Path:
    """Expected source workbook path for a month."""
    return BIDDING_DIR / f"dsfunction_{month_name(year, month)}_exact_V5_k20_classified_width.xlsx"


def hours_in_month(year: int, month: int) -> int:
    """Number of hourly records expected for a calendar month."""
    return calendar.monthrange(year, month)[1] * 24


def period_basename_suffix(
    months: list[tuple[int, int]], years: tuple[int, ...] = SELECTION_YEARS
) -> str:
    """Filename-safe suffix from per-year selection counts and season label."""
    counts = per_year_counts(months, years)
    suffix = "_".join(f"{year}_{counts[year]}" for year in years)
    season = season_label(months, years)
    return f"{suffix}_{season}" if season is not None else suffix


def season_label(
    months: list[tuple[int, int]], years: tuple[int, ...] = SELECTION_YEARS
) -> str | None:
    """Return a season name when every year selects exactly one season."""
    selected_by_year = {
        year: {month for selected_year, month in months if selected_year == year}
        for year in years
    }
    for label, season_months in SEASON_MONTHS.items():
        if all(selected_by_year[year] == season_months for year in years):
            return label
    return None


def season_selection_matrix(
    season: str, years: tuple[int, ...] = SELECTION_YEARS
) -> list[list[int]]:
    """Build a 12-by-year selection matrix for one meteorological season."""
    try:
        season_months = SEASON_MONTHS[season]
    except KeyError as exc:
        available = ", ".join(SEASON_MONTHS)
        raise ValueError(f"Unknown season {season!r}; choose from {available}.") from exc
    return [
        [1 if month in season_months else 0 for _year in years]
        for month in range(1, 13)
    ]


def period_description(
    months: list[tuple[int, int]], years: tuple[int, ...] = SELECTION_YEARS
) -> str:
    """Human-readable description for metadata and console output.

    Reports the per-year selection counts and lists the selected months, so the
    ``A/B/C`` filename token is unambiguous in the textual output even though
    it records only counts.
    """
    counts = per_year_counts(months, years)
    year_part = ", ".join(f"{year}: {counts[year]} month(s)" for year in years)
    month_list = ", ".join(f"{year}-{month:02d}" for year, month in months)
    return f"{len(months)} selected months ({year_part}) -> [{month_list}]"


# --------------------------------------------------------------------------- #
# Workbook reading
# --------------------------------------------------------------------------- #
def configure_matplotlib() -> None:
    """Apply the agreed publication typography and editable-vector settings."""
    mpl.rcParams.update(
        {
            "font.family": ["Times New Roman", "Arial", "DejaVu Serif"],
            "font.serif": ["Times New Roman", "Times", "DejaVu Serif"],
            "font.size": 9,
            "legend.frameon": False,
            "svg.fonttype": "none",
            "pdf.fonttype": 42,
            "axes.unicode_minus": False,
        }
    )


def find_unique_header(headers: tuple[object, ...], name: str) -> int:
    """Return a zero-based column index for a unique worksheet header."""
    matches = [index for index, value in enumerate(headers) if value == name]
    if len(matches) != 1:
        raise ValueError(
            f"Header {name!r} occurs {len(matches)} times; exactly one is required."
        )
    return matches[0]


def read_one_month(
    year: int, month: int, p_ess_counts: Counter[str], width_sums: dict[str, float]
) -> int:
    """Read a single month's workbook, accumulate into the shared tallies.

    Returns the number of valid records read for the month.
    """
    path = workbook_path(year, month)
    if not path.is_file():
        raise FileNotFoundError(
            f"Source workbook not found for {year}-{month:02d}: {path}"
        )

    sheet = month_name(year, month)
    workbook = load_workbook(path, read_only=True, data_only=True)
    total_rows = 0
    valid_records = 0
    try:
        if sheet not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet {sheet!r} not found in {path.name}; "
                f"available sheets: {', '.join(workbook.sheetnames)}"
            )
        worksheet = workbook[sheet]
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows)
        try:
            state_col = find_unique_header(headers, "P_ESS_state")
            width_col = find_unique_header(headers, "dsfunction_state_width")
        except ValueError as exc:
            raise ValueError(
                f"{path.name}: the classified-width columns are missing or "
                f"duplicated. {exc} Available headers: "
                f"{', '.join(str(h) for h in headers)}"
            ) from exc

        for excel_row, row in enumerate(rows, start=2):
            if all(value is None for value in row):
                continue
            # Every non-empty row represents one hour; this count is used to
            # verify the workbook is complete for the calendar month, before
            # any Pcmax=0 / Pdmax=0 hours are excluded from the tallies below.
            total_rows += 1

            p_ess_state = row[state_col]
            raw_array = row[width_col]
            if p_ess_state is None and raw_array is None:
                # Pcmax=0 or Pdmax=0 rows are intentionally excluded from
                # characteristic analysis by the Program pipeline.
                continue
            if p_ess_state not in P_ESS_ORDER:
                raise ValueError(
                    f"{path.name} row {excel_row}: unknown P_ESS_state {p_ess_state!r}."
                )
            p_ess_counts[p_ess_state] += 1

            if not isinstance(raw_array, str):
                raise ValueError(
                    f"{path.name} row {excel_row}: dsfunction_state_width must be text."
                )
            try:
                widths = ast.literal_eval(raw_array)
            except (SyntaxError, ValueError) as exc:
                raise ValueError(
                    f"{path.name} row {excel_row}: invalid dsfunction_state_width: {exc}"
                ) from exc

            if not isinstance(widths, (list, tuple)) or len(widths) != len(LEFT_ORDER):
                raise ValueError(
                    f"{path.name} row {excel_row}: expected {len(LEFT_ORDER)} widths."
                )
            numeric_widths = [float(value) for value in widths]
            if not all(math.isfinite(value) and value >= 0 for value in numeric_widths):
                raise ValueError(
                    f"{path.name} row {excel_row}: widths must be finite and non-negative."
                )

            for state, value in zip(LEFT_ORDER, numeric_widths):
                width_sums[state] += value
            valid_records += 1
    finally:
        workbook.close()

    expected = hours_in_month(year, month)
    if total_rows != expected:
        raise ValueError(
            f"Expected {expected} hourly rows for {calendar.month_name[month]} {year} "
            f"({year}-{month:02d}); found {total_rows} in {path.name}."
        )
    return valid_records


def read_source_data(
    months: list[tuple[int, int]]
) -> tuple[Counter[str], dict[str, float], int]:
    """Read and validate the pooled distributions for every selected month."""
    p_ess_counts: Counter[str] = Counter()
    width_sums = {state: 0.0 for state in LEFT_ORDER}
    total_records = 0

    for year, month in months:
        total_records += read_one_month(year, month, p_ess_counts, width_sums)

    if total_records == 0:
        raise ValueError("No records read for the requested period.")
    if sum(p_ess_counts.values()) != total_records:
        raise ValueError("P_ESS counts do not reconcile to the total record count.")
    if math.isclose(sum(width_sums.values()), 0.0, rel_tol=0, abs_tol=1e-12):
        raise ValueError("The five plotted dsfunction width totals are all zero.")
    if p_ess_counts["NC"] != 0:
        raise ValueError("NC occurs in P_ESS_state and cannot be silently omitted.")

    return p_ess_counts, width_sums, total_records


# --------------------------------------------------------------------------- #
# Source-data CSV
# --------------------------------------------------------------------------- #
def source_rows(
    p_ess_counts: Counter[str], width_sums: dict[str, float], record_count: int
) -> list[dict[str, object]]:
    """Build one auditable table containing raw totals and plotted percentages."""
    left_total = sum(width_sums.values())
    rows: list[dict[str, object]] = []
    for state in P_ESS_ORDER:
        in_left = state in LEFT_ORDER
        in_legend = state in LEGEND_ORDER
        rows.append(
            {
                "state": state,
                "legend_label": LEGEND_LABELS.get(state, "not classified"),
                "color_hex": COLORS.get(state, ""),
                "dsfunction_vertical_width_sum": (
                    f"{width_sums[state]:.12f}" if in_left else ""
                ),
                "included_in_left_pie": str(in_left).lower(),
                "dsfunction_left_normalized_percent": (
                    f"{100.0 * width_sums[state] / left_total:.8f}" if in_left else ""
                ),
                "p_ess_count": p_ess_counts[state],
                "p_ess_percent": f"{100.0 * p_ess_counts[state] / record_count:.8f}",
                "included_in_shared_legend": str(in_legend).lower(),
            }
        )
    return rows


def write_source_csv(rows: list[dict[str, object]], basename: str) -> Path:
    """Write the exact figure-source table beside all other deliverables."""
    output_path = OUTPUT_DIR / f"{basename}_source_data.csv"
    fieldnames = list(rows[0])
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    return output_path


# --------------------------------------------------------------------------- #
# Figure
# --------------------------------------------------------------------------- #
def percent_text(percent: float) -> str:
    return f"{percent:.2f}%"


def draw_left_pie(ax: plt.Axes, width_sums: dict[str, float]) -> None:
    plotted_states = [state for state in LEFT_ORDER if width_sums[state] > 0]
    values = [width_sums[state] for state in plotted_states]
    ax.pie(
        values,
        colors=[COLORS[state] for state in plotted_states],
        startangle=90,
        counterclock=False,
        radius=0.88,
        autopct=percent_text,
        pctdistance=0.68,
        textprops={"fontsize": 9, "color": "#222222"},
        wedgeprops={"linewidth": 0, "edgecolor": "none"},
        normalize=True,
    )
    ax.set_aspect("equal")
    ax.set_axis_off()


def draw_right_pie(
    ax: plt.Axes, p_ess_counts: Counter[str], record_count: int
) -> None:
    radius = 0.82
    values = [p_ess_counts[state] for state in LEGEND_ORDER]
    wedges, _ = ax.pie(
        values,
        colors=[COLORS[state] for state in LEGEND_ORDER],
        startangle=90,
        counterclock=False,
        radius=radius,
        wedgeprops={"linewidth": 0, "edgecolor": "none"},
        normalize=True,
    )

    outside_positions = {
        "CC": (1.10, -0.46),
        "CD": (1.16, -0.75),
        "DC": (-1.16, -0.75),
        "DD": (-1.10, -0.46),
    }

    for state, wedge in zip(LEGEND_ORDER, wedges):
        count = p_ess_counts[state]
        if count == 0:
            # No wedge exists for this state; skip its label and leader line
            # so degenerate geometry is never drawn for an absent state.
            continue
        percent = 100.0 * count / record_count
        angle = math.radians((wedge.theta1 + wedge.theta2) / 2.0)
        if state in {"MC", "NU", "MD"}:
            label_radius = 0.48 if state == "NU" else 0.52
            ax.text(
                label_radius * math.cos(angle),
                label_radius * math.sin(angle),
                percent_text(percent),
                ha="center",
                va="center",
                fontsize=9,
                color="#222222" if state == "NU" else "#FFFFFF",
            )
            continue

        text_x, text_y = outside_positions[state]
        side = 1 if text_x > 0 else -1
        anchor_x = radius * 0.98 * math.cos(angle)
        anchor_y = radius * 0.98 * math.sin(angle)
        elbow_x = radius * 1.10 * math.cos(angle)
        elbow_y = radius * 1.10 * math.sin(angle)
        line_end_x = text_x - side * 0.035
        ax.plot(
            [anchor_x, elbow_x, line_end_x],
            [anchor_y, elbow_y, text_y],
            color="#B7B7B7",
            linewidth=0.65,
            solid_capstyle="round",
            clip_on=False,
            zorder=1,
        )
        ax.text(
            text_x,
            text_y,
            percent_text(percent),
            ha="left" if side > 0 else "right",
            va="center",
            fontsize=9,
            color="#222222",
            clip_on=False,
            zorder=2,
        )

    # Use asymmetric horizontal limits to reserve space for the longer labels
    # on the right without shrinking the pie or changing the 180 mm canvas.
    ax.set_xlim(-1.36, 1.58)
    ax.set_ylim(-1.07, 1.07)
    ax.set_aspect("equal")
    ax.set_axis_off()


def add_aligned_panel_labels(
    fig: plt.Figure, left_ax: plt.Axes, right_ax: plt.Axes
) -> None:
    """Add panel labels aligned to the horizontal centres of both pie axes."""
    inverse_figure = fig.transFigure.inverted()
    left_center_x = inverse_figure.transform(left_ax.transData.transform((0, 0)))[0]
    right_center_x = inverse_figure.transform(right_ax.transData.transform((0, 0)))[0]
    panel_y = 0.055

    for label, center_x in (("(a)", left_center_x), ("(b)", right_center_x)):
        fig.text(
            center_x,
            panel_y,
            label,
            ha="center",
            va="center",
            fontsize=9,
            color="#222222",
        )


def draw_legend(ax: plt.Axes) -> None:
    handles = [
        Line2D(
            [0],
            [0],
            marker="s",
            linestyle="None",
            markersize=7.2,
            markerfacecolor=COLORS[state],
            markeredgecolor="none",
        )
        for state in LEGEND_ORDER
    ]
    labels = [LEGEND_LABELS[state] for state in LEGEND_ORDER]
    ax.legend(
        handles,
        labels,
        loc="center",
        fontsize=9,
        handlelength=0.8,
        handletextpad=0.45,
        labelspacing=0.53,
        borderaxespad=0,
        frameon=False,
    )
    ax.set_axis_off()


def build_figure(
    p_ess_counts: Counter[str], width_sums: dict[str, float], record_count: int
) -> plt.Figure:
    configure_matplotlib()
    mm_per_inch = 25.4
    fig = plt.figure(
        figsize=(FIGURE_WIDTH_MM / mm_per_inch, FIGURE_HEIGHT_MM / mm_per_inch),
        facecolor="none",
    )
    grid = fig.add_gridspec(
        1,
        3,
        width_ratios=(1.10, 1.03, 1.30),
        left=0.018,
        right=0.988,
        bottom=0.135,
        top=0.985,
        wspace=0.015,
    )
    left_ax = fig.add_subplot(grid[0, 0])
    legend_ax = fig.add_subplot(grid[0, 1])
    right_ax = fig.add_subplot(grid[0, 2])

    draw_left_pie(left_ax, width_sums)
    draw_legend(legend_ax)
    draw_right_pie(right_ax, p_ess_counts, record_count)
    add_aligned_panel_labels(fig, left_ax, right_ax)
    return fig


def save_figure(fig: plt.Figure, basename: str, description: str) -> list[Path]:
    """Export editable vectors and a high-resolution transparent preview."""
    base = OUTPUT_DIR / basename
    svg_path = base.with_suffix(".svg")
    pdf_path = base.with_suffix(".pdf")
    png_path = base.with_suffix(".png")

    fig.savefig(
        svg_path,
        transparent=True,
        metadata={
            "Creator": "Matplotlib",
            "Description": f"{description} state distributions",
        },
    )
    fig.savefig(
        pdf_path,
        transparent=True,
        metadata={
            "Creator": "Matplotlib",
            "Title": f"{description} state distributions",
        },
    )
    fig.savefig(
        png_path,
        dpi=PNG_DPI,
        transparent=True,
        metadata={"Title": f"{description} state distributions"},
    )
    return [svg_path, pdf_path, png_path]


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if PLOT_SEASONS:
        periods = [
            (season, select_months(season_selection_matrix(season)))
            for season in SEASON_MONTHS
        ]
    else:
        periods = [("Custom", select_months())]

    for requested_label, months in periods:
        description = period_description(months)
        basename = f"state_distribution_pies_{period_basename_suffix(months)}"
        p_ess_counts, width_sums, record_count = read_source_data(months)
        rows = source_rows(p_ess_counts, width_sums, record_count)
        csv_path = write_source_csv(rows, basename)
        fig = build_figure(p_ess_counts, width_sums, record_count)
        figure_paths = save_figure(fig, basename, description)
        plt.close(fig)

        left_total = sum(width_sums.values())
        print(f"Period: {requested_label} ({description})")
        print(f"Months: {len(months)}")
        print(f"Records: {record_count}")
        print(
            "Panel (a): "
            + ", ".join(
                f"{state}={100.0 * width_sums[state] / left_total:.2f}%"
                for state in LEFT_ORDER
            )
        )
        print(
            "Panel (b): "
            + ", ".join(
                f"{state}={100.0 * p_ess_counts[state] / record_count:.2f}%"
                for state in LEGEND_ORDER
            )
        )
        print(f"Source data: {csv_path}")
        for path in figure_paths:
            print(f"Figure: {path}")


if __name__ == "__main__":
    main()
