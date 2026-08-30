"""Create combined and standalone radial figures for the study rates.

Figure contract
---------------
Core conclusion:
    The three side-by-side radial panels compare monthly and annual variation
    in profit increment, carbon reduction, and generation-cost reduction.
Figure archetype:
    Three-panel quantitative polar comparison.
Field mapping:
    ``rate carbon`` -> monthly carbon reduction rate -> fraction -> percent
    ``cost reduction`` -> monthly cost reduction rate -> fraction -> percent
    ``carbon reduction rate`` -> annual carbon rate -> fraction -> percent
    ``cost reduction rate`` -> annual cost rate -> fraction -> percent
Integrity:
    All 36 monthly rows and the three matching annual rows are retained for
    every metric. Carbon and cost values below zero extend outward from 0%.
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
import re

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

from plot_profit_increment_radial import (
    ANNUAL_TEXT_SPAN_DEGREES,
    GRID_COLOR,
    MONTHS,
    OVERALL_PERIOD_LABEL,
    OVERALL_SUBTITLE_FONTSIZE,
    SCENARIO_WORKBOOKS,
    TEXT_COLOR,
    TITLE_FONTSIZE,
    YEARS,
    ProfitIncrementData,
    _draw_curved_text,
    _draw_grid_interval_legend,
    _upright_radial_rotation,
    _upright_tangent_rotation,
    draw_profit_increment_panel,
    load_profit_increment_data,
    print_input_workbook_info,
)


OUTPUT_ROOT = Path(__file__).with_name("outputs")
OUTPUT_DIRECTORY = "combined_radial"
OUTPUT_STEM = "profit_carbon_cost_radial_2023_2025"
STANDALONE_OUTPUTS = {
    "carbon": (
        "carbon_reduction_radial",
        "carbon_reduction_rate_radial_2023_2025",
        "Carbon reduction rate",
    ),
    "cost": (
        "cost_reduction_radial",
        "cost_reduction_rate_radial_2023_2025",
        "Generation-cost reduction rate",
    ),
}
FIGURE_WIDTH_MM = 250
FIGURE_HEIGHT_MM = 90
STANDALONE_FIGURE_WIDTH_MM = 90
STANDALONE_FIGURE_HEIGHT_MM = 90
PNG_DPI = 720
INNER_RADIUS = 10.0
DATA_OUTER_RADIUS = 50.0
DATA_RADIAL_SPAN = DATA_OUTER_RADIUS - INNER_RADIUS
MONTH_LABEL_RADIUS = 56.0
ANNULUS_INNER_RADIUS = 62.0
ANNULUS_OUTER_RADIUS = 76.0
PLOT_LIMIT = 78.0
ANNUAL_LABEL_FONTSIZE = 10.2
REDUCTION_SCALE_LABEL_FONTSIZE = 9.4


@dataclass(frozen=True)
class MetricSpec:
    """Data-column, scale, color, and export contract for one rate."""

    key: str
    title: str
    monthly_header: str
    annual_header: str
    axis_min: float
    axis_max: float
    ticks: tuple[float, ...]
    label_ticks: tuple[float, ...]
    label_tick_angles: dict[float, float]
    bar_color: str
    annulus_color: str


@dataclass(frozen=True)
class MetricData:
    """Validated ordered monthly and annual percentage values."""

    months: tuple[str, ...]
    monthly_percent: np.ndarray
    annual_percent: dict[int, float]
    overall_percent: float


METRICS = (
    MetricSpec(
        key="carbon",
        title="Carbon-emission reduction rate",
        monthly_header="rate carbon",
        annual_header="carbon reduction rate",
        axis_min=-0.5,
        axis_max=8.0,
        ticks=(0.0, 2.0, 4.0, 6.0, 8.0),
        label_ticks=(2.0, 4.0, 6.0, 8.0),
        label_tick_angles={2.0: 120.0, 4.0: 120.0, 6.0: 120.0, 8.0: 120.0},
        bar_color="#55A868",
        annulus_color="#DCEFE3",
    ),
    MetricSpec(
        key="cost",
        title="Natural-gas fuel-cost reduction rate",
        monthly_header="cost reduction",
        annual_header="cost reduction rate",
        axis_min=-0.5,
        axis_max=8.0,
        ticks=(0.0, 2.0, 4.0, 6.0, 8.0),
        label_ticks=(2.0, 4.0, 6.0, 8.0),
        label_tick_angles={2.0: 120.0, 4.0: 120.0, 6.0: 120.0, 8.0: 120.0},
        bar_color="#E69F45",
        annulus_color="#F9E6CF",
    ),
)


def configure_matplotlib() -> None:
    """Set explicit, publication-safe and editable export typography."""
    mpl.rcParams.update(
        {
            "font.family": ["Times New Roman", "Arial", "serif"],
            "font.serif": ["Times New Roman", "Times", "DejaVu Serif"],
            "font.size": 9,
            "svg.fonttype": "none",
            "pdf.fonttype": 42,
            "axes.unicode_minus": True,
            "figure.facecolor": "none",
            "axes.facecolor": "none",
            "savefig.facecolor": "none",
            "savefig.edgecolor": "none",
            "savefig.transparent": True,
        }
    )


def _header_map(header_row: tuple[object, ...]) -> dict[str, int]:
    return {
        str(value).strip().lower(): index
        for index, value in enumerate(header_row)
        if value is not None
    }


def _month_code(value: object) -> str:
    if isinstance(value, (int, float)):
        return str(int(value))
    return str(value).strip()


def load_metric_data(workbook_path: Path, spec: MetricSpec) -> MetricData:
    """Load one metric using its explicit monthly and annual field mapping."""
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        monthly_sheet = workbook["Monthly Analysis"]
        annual_sheet = workbook["Annual Summary"]

        monthly_rows = monthly_sheet.iter_rows(values_only=True)
        monthly_headers = _header_map(next(monthly_rows))
        required_monthly = {"month", spec.monthly_header}
        missing_monthly = required_monthly.difference(monthly_headers)
        if missing_monthly:
            raise ValueError(
                f"Missing Monthly Analysis columns: {sorted(missing_monthly)}"
            )

        monthly_values: dict[str, float] = {}
        for row in monthly_rows:
            code = _month_code(row[monthly_headers["month"]])
            if not re.fullmatch(r"202[3-5](0[1-9]|1[0-2])", code):
                continue
            raw_value = row[monthly_headers[spec.monthly_header]]
            if raw_value is None:
                raise ValueError(f"Missing {spec.key} value for {code}")
            monthly_values[code] = float(raw_value) * 100.0

        expected_months = tuple(
            f"{year}{month:02d}" for year in YEARS for month in MONTHS
        )
        missing_months = [code for code in expected_months if code not in monthly_values]
        if missing_months:
            raise ValueError(f"Missing monthly observations: {missing_months}")

        annual_rows = annual_sheet.iter_rows(values_only=True)
        annual_headers = _header_map(next(annual_rows))
        required_annual = {"annual", spec.annual_header}
        missing_annual = required_annual.difference(annual_headers)
        if missing_annual:
            raise ValueError(
                f"Missing Annual Summary columns: {sorted(missing_annual)}"
            )

        annual_values: dict[int, float] = {}
        overall_value: float | None = None
        for row in annual_rows:
            period = str(row[annual_headers["annual"]]).strip()
            raw_value = row[annual_headers[spec.annual_header]]
            if period == OVERALL_PERIOD_LABEL:
                if raw_value is None:
                    raise ValueError(f"Missing overall {spec.key} rate")
                overall_value = float(raw_value) * 100.0
                continue
            match = re.fullmatch(r"(202[3-5])-(\d{4})", period)
            if match is None:
                continue
            year = int(match.group(1))
            if raw_value is None:
                raise ValueError(f"Missing annual {spec.key} rate for {year}")
            annual_values[year] = float(raw_value) * 100.0

        missing_years = [year for year in YEARS if year not in annual_values]
        if missing_years:
            raise ValueError(f"Missing annual summary values: {missing_years}")
        if overall_value is None:
            raise ValueError(
                f"Missing overall summary row: {OVERALL_PERIOD_LABEL}"
            )

        ordered_monthly = np.array(
            [monthly_values[code] for code in expected_months], dtype=float
        )
        if not np.isfinite(ordered_monthly).all():
            raise ValueError(f"{spec.key} data contain non-finite values")
        if (
            ordered_monthly.min() < spec.axis_min
            or ordered_monthly.max() > spec.axis_max
        ):
            raise ValueError(
                f"{spec.key} values fall outside the confirmed axis range "
                f"{spec.axis_min:g}% to {spec.axis_max:g}%"
            )

        return MetricData(
            months=expected_months,
            monthly_percent=ordered_monthly,
            annual_percent=annual_values,
            overall_percent=overall_value,
        )
    finally:
        workbook.close()


def _value_to_radius(value: float | np.ndarray, spec: MetricSpec) -> float | np.ndarray:
    """Map a percentage to a fixed visual radius on its independent scale."""
    fraction_from_max = (spec.axis_max - value) / (spec.axis_max - spec.axis_min)
    return INNER_RADIUS + fraction_from_max * DATA_RADIAL_SPAN


def _format_tick(value: float) -> str:
    return f"{value:g}%"


def _draw_scale(ax: mpl.axes.Axes, spec: MetricSpec) -> None:
    """Draw transparent percentage rings without opaque label backgrounds."""
    theta = np.linspace(0.0, 2.0 * np.pi, 721)

    for tick in spec.ticks:
        radius = float(_value_to_radius(tick, spec))
        is_zero = abs(tick) < 1e-12
        ax.plot(
            theta,
            np.full_like(theta, radius),
            color=GRID_COLOR,
            linewidth=2.0 if is_zero else 1.8,
            linestyle="-" if is_zero else (0, (3.2, 2.4)),
            alpha=0.78,
            zorder=0,
        )

        # Omit labels that would crowd the month ring or the center boundary;
        # the remaining values sit in the annual-sector gap and stay legible.
        if tick in spec.label_ticks:
            ax.text(
                np.deg2rad(spec.label_tick_angles[tick]),
                radius,
                _format_tick(tick),
                ha="center",
                va="center",
                fontsize=REDUCTION_SCALE_LABEL_FONTSIZE,
                fontfamily="Times New Roman",
                fontweight="semibold",
                color=TEXT_COLOR,
                zorder=7,
            )


def _draw_monthly_bars(
    ax: mpl.axes.Axes, data: MetricData, spec: MetricSpec
) -> None:
    """Draw positive values inward and negative values outward from 0%."""
    count = len(data.monthly_percent)
    sector_width = 2.0 * np.pi / count
    angles = np.arange(count, dtype=float) * sector_width + sector_width / 2.0
    bar_width = sector_width * 0.78

    zero_radius = float(_value_to_radius(0.0, spec))
    tips = np.asarray(_value_to_radius(data.monthly_percent, spec), dtype=float)
    bottoms = np.minimum(tips, zero_radius)
    heights = np.abs(tips - zero_radius)

    ax.bar(
        angles,
        heights,
        width=bar_width,
        bottom=bottoms,
        align="center",
        color=spec.bar_color,
        edgecolor="white",
        linewidth=0.35,
        alpha=0.94,
        zorder=3,
    )

    month_labels = ("JAN", "FEB", "MAR", "APR", "MAY", "JUN",
                    "JUL", "AUG", "SEP", "OCT", "NOV", "DEC")
    for index, (angle, code) in enumerate(zip(angles, data.months)):
        angle_degrees = float(np.rad2deg(angle))
        ax.text(
            angle,
            MONTH_LABEL_RADIUS,
            month_labels[index % 12],
            ha="center",
            va="center",
            rotation=_upright_radial_rotation(angle_degrees),
            rotation_mode="anchor",
            fontsize=7.2,
            fontweight="semibold",
            color=TEXT_COLOR,
            zorder=8,
        )


def _draw_annual_annulus(
    ax: mpl.axes.Axes, data: MetricData, spec: MetricSpec
) -> None:
    """Draw the three-part annual ring with editable curved labels."""
    sector_width = 2.0 * np.pi / 3.0
    ring_height = ANNULUS_OUTER_RADIUS - ANNULUS_INNER_RADIUS
    centers = np.arange(3, dtype=float) * sector_width + sector_width / 2.0

    ax.bar(
        centers,
        np.full(3, ring_height),
        width=np.full(3, sector_width),
        bottom=np.full(3, ANNULUS_INNER_RADIUS),
        align="center",
        color=spec.annulus_color,
        edgecolor="white",
        linewidth=1.35,
        zorder=2,
    )

    label_radius = (ANNULUS_INNER_RADIUS + ANNULUS_OUTER_RADIUS) / 2.0
    label_layout = ((60.0, 1.0), (180.0, -1.0), (300.0, 1.0))
    for year_index, ((center_degrees, reading_direction), year) in enumerate(
        zip(label_layout, YEARS)
    ):
        annual_months = data.monthly_percent[year_index * 12 : (year_index + 1) * 12]
        annual_label = (
            f"{year}: {annual_months.min():.2f}–{annual_months.max():.2f}%, "
            f"Annual. {data.annual_percent[year]:.2f}%"
        )
        _draw_curved_text(
            ax=ax,
            text=annual_label,
            center_angle_degrees=center_degrees,
            radius=label_radius,
            fontsize=ANNUAL_LABEL_FONTSIZE,
            color=TEXT_COLOR,
            span_degrees=ANNUAL_TEXT_SPAN_DEGREES,
            fontweight="bold",
            direction=reading_direction,
            outward_facing=year in (2023, 2025),
        )

    theta = np.linspace(0.0, 2.0 * np.pi, 721)
    for radius in (ANNULUS_INNER_RADIUS, ANNULUS_OUTER_RADIUS):
        ax.plot(
            theta,
            np.full_like(theta, radius),
            color=spec.bar_color,
            linewidth=0.7,
            zorder=4,
        )

    for boundary in np.arange(0.0, 2.0 * np.pi, sector_width):
        ax.plot(
            [boundary, boundary],
            [INNER_RADIUS, ANNULUS_OUTER_RADIUS],
            color="white",
            linewidth=1.35,
            solid_capstyle="butt",
            zorder=6,
        )


def draw_metric_panel(
    ax: mpl.axes.Axes,
    data: MetricData,
    spec: MetricSpec,
    title: str | None = None,
) -> None:
    """Draw one reduction-rate radial chart on an existing polar axis."""
    ax.set_theta_zero_location("N")
    ax.set_theta_direction(-1)
    ax.set_ylim(0.0, PLOT_LIMIT)
    ax.set_xticks([])
    ax.set_yticks([])
    ax.grid(False)
    ax.spines["polar"].set_visible(False)
    ax.set_facecolor("none")

    _draw_scale(ax, spec)
    _draw_monthly_bars(ax, data, spec)
    _draw_annual_annulus(ax, data, spec)
    grid_interval = spec.ticks[1] - spec.ticks[0]
    _draw_grid_interval_legend(ax, grid_interval)

    center = plt.Circle(
        (0.5, 0.5),
        INNER_RADIUS / PLOT_LIMIT / 2.0,
        transform=ax.transAxes,
        facecolor="none",
        edgecolor=GRID_COLOR,
        linewidth=1.1,
        alpha=0.75,
        zorder=5,
    )
    ax.add_artist(center)

    if title is not None:
        ax.set_title(
            title,
            fontsize=TITLE_FONTSIZE,
            fontweight="bold",
            color=TEXT_COLOR,
            pad=16,
        )
        ax.text(
            0.5,
            1.003,
            f"(Overall: {data.overall_percent:.2f}%)",
            transform=ax.transAxes,
            ha="center",
            va="bottom",
            fontsize=OVERALL_SUBTITLE_FONTSIZE,
            fontweight="semibold",
            color=TEXT_COLOR,
            clip_on=False,
        )


def create_combined_figure(
    profit_data: ProfitIncrementData,
    metric_data: dict[str, MetricData],
) -> mpl.figure.Figure:
    """Create the final 250 mm × 90 mm three-panel figure."""
    fig = plt.figure(
        figsize=(FIGURE_WIDTH_MM / 25.4, FIGURE_HEIGHT_MM / 25.4),
        facecolor="none",
    )
    axes = [
        fig.add_axes([0.000, 0.005, 0.334, 0.86], projection="polar"),
        fig.add_axes([0.333, 0.005, 0.334, 0.86], projection="polar"),
        fig.add_axes([0.666, 0.005, 0.334, 0.86], projection="polar"),
    ]

    draw_profit_increment_panel(axes[0], profit_data, "Profit increment rate")
    for ax, spec in zip(axes[1:], METRICS):
        draw_metric_panel(ax, metric_data[spec.key], spec, spec.title)

    return fig


def create_metric_figure(
    data: MetricData,
    spec: MetricSpec,
    title: str,
) -> mpl.figure.Figure:
    """Create a journal-column-width standalone reduction-rate figure."""
    fig = plt.figure(
        figsize=(
            STANDALONE_FIGURE_WIDTH_MM / 25.4,
            STANDALONE_FIGURE_HEIGHT_MM / 25.4,
        ),
        facecolor="none",
    )
    ax = fig.add_axes([0.02, 0.005, 0.96, 0.86], projection="polar")
    draw_metric_panel(ax, data, spec, title)
    return fig


def export_figure(
    fig: mpl.figure.Figure,
    output_directory: str = OUTPUT_DIRECTORY,
    output_stem: str = OUTPUT_STEM,
    output_root: Path = OUTPUT_ROOT,
) -> Path:
    """Export editable vectors plus 720 dpi PNG and compressed TIFF."""
    output_dir = output_root / output_directory
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = output_dir / output_stem
    fig.savefig(stem.with_suffix(".svg"), transparent=True)
    fig.savefig(stem.with_suffix(".png"), dpi=PNG_DPI, transparent=True)
    fig.savefig(
        stem.with_suffix(".tiff"),
        dpi=PNG_DPI,
        transparent=True,
        pil_kwargs={"compression": "tiff_lzw"},
    )
    try:
        fig.savefig(stem.with_suffix(".pdf"), transparent=True)
    except PermissionError:
        print(f"Warning: PDF is open and could not be replaced: {stem.with_suffix('.pdf')}")
    return output_dir


def build_parser() -> argparse.ArgumentParser:
    """Build a CLI that defaults to both study scenarios."""
    parser = argparse.ArgumentParser(
        description="Plot combined and standalone Figure 4d radial rates."
    )
    parser.add_argument(
        "--input",
        type=Path,
        help="Process one custom workbook instead of both default scenarios.",
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        help="Output root for --input (defaults to the legacy output root).",
    )
    return parser


def _process_workbook(
    scenario: str,
    workbook_path: Path,
    output_root: Path,
) -> None:
    """Create the combined figure and two reduction figures for one scenario."""
    print_input_workbook_info(scenario, workbook_path)
    profit_data = load_profit_increment_data(workbook_path)
    metric_data = {
        spec.key: load_metric_data(workbook_path, spec) for spec in METRICS
    }
    figure = create_combined_figure(profit_data, metric_data)
    output_dir = export_figure(figure, output_root=output_root)
    plt.close(figure)

    standalone_dirs = []
    for spec in METRICS:
        directory, stem, title = STANDALONE_OUTPUTS[spec.key]
        figure = create_metric_figure(metric_data[spec.key], spec, title)
        standalone_dirs.append(
            export_figure(figure, directory, stem, output_root=output_root)
        )
        plt.close(figure)

    print(
        f"Profit increment rate [{scenario}]: "
        f"{len(profit_data.monthly_percent)} months, "
        f"range {profit_data.monthly_percent.min():.2f}% to "
        f"{profit_data.monthly_percent.max():.2f}%"
    )
    for spec in METRICS:
        data = metric_data[spec.key]
        print(
            f"{spec.title} [{scenario}]: {len(data.monthly_percent)} months, "
            f"range {data.monthly_percent.min():.2f}% to "
            f"{data.monthly_percent.max():.2f}%"
        )
        print(
            f"Annual rates [{scenario}]: "
            + ", ".join(
                f"{year}={data.annual_percent[year]:.2f}%" for year in YEARS
            )
        )
    print(f"Created combined figures [{scenario}] in: {output_dir}")
    print(
        f"Created standalone figures [{scenario}] in: "
        + ", ".join(str(path) for path in standalone_dirs)
    )


def main(
    workbook_path: Path | None = None,
    output_root: Path | None = None,
) -> None:
    configure_matplotlib()
    if workbook_path is not None:
        if not workbook_path.is_file():
            print(f"Warning [Custom]: input workbook not found; skipped: {workbook_path}")
            return
        _process_workbook(
            "Custom",
            workbook_path,
            output_root if output_root is not None else OUTPUT_ROOT,
        )
        return

    for scenario, scenario_workbook in SCENARIO_WORKBOOKS.items():
        if not scenario_workbook.is_file():
            print(
                f"Warning [{scenario}]: input workbook not found; skipped: "
                f"{scenario_workbook}"
            )
            continue
        _process_workbook(
            scenario,
            scenario_workbook,
            OUTPUT_ROOT / scenario,
        )


if __name__ == "__main__":
    cli_arguments = build_parser().parse_args()
    main(cli_arguments.input, cli_arguments.output_root)
