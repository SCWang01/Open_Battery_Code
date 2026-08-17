"""Plot 2023-2025 profit increment rates as inward radial bars.

Figure contract
---------------
Core conclusion:
    Monthly profit increment rates vary across 2023-2025, while the outer
    annulus provides the corresponding annual rates for direct comparison.
Figure archetype:
    Quantitative polar panel that can be rendered alone or in a three-panel
    comparison with the carbon- and cost-reduction rates.
Evidence hierarchy:
    The 36 monthly observations are the primary evidence; the three annual
    summary values are direct contextual labels in the outer annulus.
Integrity:
    All 36 monthly observations and all three matching annual summary rows
    are used without filtering or aggregation.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re

import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties
from matplotlib.textpath import TextPath
import numpy as np
from openpyxl import load_workbook


SOURCE_WORKBOOK = Path(__file__).with_name("analysis_202301_202512.xlsx")
OUTPUT_DIR = Path(__file__).with_name("outputs") / "profit_increment_radial"
OUTPUT_STEM = "profit_increment_rate_radial_2023_2025"

YEARS = (2023, 2024, 2025)
MONTHS = tuple(range(1, 13))

FIGURE_WIDTH_MM = 90
FIGURE_HEIGHT_MM = 90
PNG_DPI = 720

# The workbook currently reaches about 209%.  Keep headroom so the chart does
# not reject valid observations when a later workbook contains a slightly
# larger monthly increment.
MAX_PERCENT = 250.0
INNER_RADIUS = 50.0
BAR_OUTER_RADIUS = INNER_RADIUS + MAX_PERCENT
MONTH_LABEL_RADIUS = 280.0
ANNULUS_INNER_RADIUS = 310.0
ANNULUS_OUTER_RADIUS = 380.0
PLOT_LIMIT = 390.0
ANNUAL_TEXT_SPAN_DEGREES = 104.0
SCALE_LABEL_ANGLE_DEGREES = 240.0
SCALE_LABEL_FONTSIZE = 9.4

BLUE = "#4C8FD3"
ANNULUS_BLUE = "#DCEAF7"
GRID_COLOR = "#B7B7B7"
TEXT_COLOR = "#333333"
BAR_EDGE_COLOR = "#FFFFFF"
ANNUAL_LABEL_FONTSIZE = 10.2
TITLE_FONTSIZE = 15.5


@dataclass(frozen=True)
class ProfitIncrementData:
    """Validated monthly and annual percentage values."""

    months: tuple[str, ...]
    monthly_percent: np.ndarray
    annual_percent: dict[int, float]


def configure_matplotlib() -> None:
    """Set publication-oriented typography and editable vector text."""
    mpl.rcParams.update(
        {
            # Times New Roman is the requested primary face; Arial and serif
            # remain explicit fallbacks for portable rendering.
            "font.family": ["Times New Roman", "Arial", "serif"],
            "font.serif": ["Times New Roman", "Times", "DejaVu Serif"],
            "font.size": 9,
            "svg.fonttype": "none",
            "pdf.fonttype": 42,
            "axes.unicode_minus": True,
            "figure.facecolor": "none",
            "axes.facecolor": "none",
            "savefig.facecolor": "none",
            "savefig.edgecolor": "none",
            "savefig.transparent": True,
        }
    )


def _header_map(header_row: tuple[object, ...]) -> dict[str, int]:
    return {
        str(value).strip().lower(): index
        for index, value in enumerate(header_row)
        if value is not None
    }


def _month_code(value: object) -> str:
    if isinstance(value, (int, float)):
        return str(int(value))
    return str(value).strip()


def load_profit_increment_data(workbook_path: Path) -> ProfitIncrementData:
    """Load the 36 monthly rates and the three annual rates from Sheet2."""
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        monthly_sheet = workbook["Monthly Analysis"]
        annual_sheet = workbook.worksheets[1]

        monthly_rows = monthly_sheet.iter_rows(values_only=True)
        monthly_headers = _header_map(next(monthly_rows))
        required_monthly = {"month", "profit_increment_k20"}
        missing_monthly = required_monthly.difference(monthly_headers)
        if missing_monthly:
            raise ValueError(
                f"Missing Monthly Analysis columns: {sorted(missing_monthly)}"
            )

        monthly_values: dict[str, float] = {}
        for row in monthly_rows:
            code = _month_code(row[monthly_headers["month"]])
            if not re.fullmatch(r"202[3-5](0[1-9]|1[0-2])", code):
                continue
            raw_value = row[monthly_headers["profit_increment_k20"]]
            if raw_value is None:
                raise ValueError(f"Missing profit increment value for {code}")
            monthly_values[code] = float(raw_value) * 100.0

        expected_months = tuple(
            f"{year}{month:02d}" for year in YEARS for month in MONTHS
        )
        missing_months = [code for code in expected_months if code not in monthly_values]
        if missing_months:
            raise ValueError(f"Missing monthly observations: {missing_months}")

        annual_rows = annual_sheet.iter_rows(values_only=True)
        annual_headers = _header_map(next(annual_rows))
        required_annual = {"annual", "profit_increment_rate_k20"}
        missing_annual = required_annual.difference(annual_headers)
        if missing_annual:
            raise ValueError(
                f"Missing Sheet2 annual columns: {sorted(missing_annual)}"
            )

        annual_values: dict[int, float] = {}
        for row in annual_rows:
            period = str(row[annual_headers["annual"]]).strip()
            match = re.fullmatch(r"(202[3-5])-(\d{4})", period)
            if match is None:
                continue
            year = int(match.group(1))
            raw_value = row[annual_headers["profit_increment_rate_k20"]]
            if raw_value is None:
                raise ValueError(f"Missing annual profit increment rate for {year}")
            annual_values[year] = float(raw_value) * 100.0

        missing_years = [year for year in YEARS if year not in annual_values]
        if missing_years:
            raise ValueError(f"Missing annual summary values: {missing_years}")

        ordered_monthly = np.array(
            [monthly_values[code] for code in expected_months], dtype=float
        )
        if not np.isfinite(ordered_monthly).all():
            raise ValueError("Monthly profit increment data contain non-finite values")
        if (ordered_monthly < 0).any() or (ordered_monthly > MAX_PERCENT).any():
            raise ValueError(
                f"Monthly values must lie between 0% and {MAX_PERCENT:.0f}%"
            )

        return ProfitIncrementData(
            months=expected_months,
            monthly_percent=ordered_monthly,
            annual_percent=annual_values,
        )
    finally:
        workbook.close()


def _upright_tangent_rotation(angle_degrees: float) -> float:
    """Return a tangent rotation constrained to an upright readable range."""
    rotation = ((-angle_degrees + 180.0) % 360.0) - 180.0
    if rotation > 90.0:
        rotation -= 180.0
    elif rotation < -90.0:
        rotation += 180.0
    return rotation


def _upright_radial_rotation(angle_degrees: float) -> float:
    """Return a radial rotation while keeping the month text upright."""
    rotation = _upright_tangent_rotation(angle_degrees) + 90.0
    if rotation > 90.0:
        rotation -= 180.0
    elif rotation < -90.0:
        rotation += 180.0
    return rotation


def _draw_curved_text(
    ax: mpl.axes.Axes,
    text: str,
    center_angle_degrees: float,
    radius: float,
    fontsize: float,
    color: str,
    span_degrees: float,
    fontweight: str = "normal",
    direction: float | None = None,
    outward_facing: bool = False,
) -> None:
    """Draw editable characters continuously along a circular arc."""
    font = FontProperties(
        family=["Times New Roman", "Arial", "serif"],
        size=fontsize,
        weight=fontweight,
    )
    advances: list[float] = []
    for character in text:
        if character.isspace():
            advance = fontsize * 0.34
        else:
            bounds = TextPath(
                (0, 0), character, prop=font, size=fontsize
            ).get_extents()
            advance = max(float(bounds.width), fontsize * 0.22)
        # Keep the annual summaries compact.  TextPath already includes each
        # glyph's natural Times New Roman width, so only a very small tracking
        # allowance is needed between characters.
        advances.append(advance + fontsize * 0.045)

    advance_array = np.asarray(advances, dtype=float)
    centers = np.cumsum(advance_array) - advance_array / 2.0
    normalized = centers / advance_array.sum() - 0.5

    # Text on the lower half runs against increasing polar angle so that the
    # complete phrase still reads left-to-right after tangent rotation.
    reading_direction = (
        direction
        if direction is not None
        else (-1.0 if 90.0 < center_angle_degrees < 270.0 else 1.0)
    )
    character_angles = (
        center_angle_degrees + reading_direction * normalized * span_degrees
    )

    for character, angle_degrees in zip(text, character_angles):
        if character.isspace():
            continue
        if outward_facing:
            # For clockwise polar coordinates, a baseline angle of -theta
            # makes the glyph's vertical axis point from the center outward.
            character_rotation = ((-float(angle_degrees) + 180.0) % 360.0) - 180.0
        else:
            character_rotation = _upright_tangent_rotation(float(angle_degrees))
        ax.text(
            np.deg2rad(angle_degrees),
            radius,
            character,
            ha="center",
            va="center",
            rotation=character_rotation,
            rotation_mode="anchor",
            fontsize=fontsize,
            fontproperties=font,
            color=color,
            zorder=9,
        )


def _draw_scale(ax: mpl.axes.Axes) -> None:
    """Draw reversed percentage rings: 0% outside and MAX_PERCENT inward."""
    theta = np.linspace(0.0, 2.0 * np.pi, 721)
    scale_values = (0, 50, 100, 150, 200)
    # Place the 100% and 150% labels near the requested month axes.  The
    # 200% label remains in its established position.
    # Keep the 150% label slightly to the right of its previous midpoint.
    label_angles = {100: 245.0, 150: 222.0, 200: 185.0}

    for value in scale_values:
        radius = BAR_OUTER_RADIUS - value
        linewidth = 2.0 if value == 0 else 1.8
        linestyle = "-" if value == 0 else (0, (3.2, 2.4))
        ax.plot(
            theta,
            np.full_like(theta, radius),
            color=GRID_COLOR,
            linewidth=linewidth,
            linestyle=linestyle,
            alpha=0.78,
            zorder=0,
        )

        # The 240-degree ray follows the gap between annual sectors.  Values
        # near the outer monthly bars are intentionally omitted when the
        # horizontal label would overlap those bars or their month labels.
        # Place each horizontal label in a different lower-half gap.  The
        # chosen locations avoid the local monthly bars at the corresponding
        # radius and form a clean outer-to-inner staircase.
        if value in label_angles:
            ax.text(
                np.deg2rad(label_angles[value]),
                radius,
                f"{value:g}%",
                ha="center",
                va="center",
                fontsize=SCALE_LABEL_FONTSIZE,
                fontfamily="Times New Roman",
                fontweight="semibold",
                color=TEXT_COLOR,
                zorder=7,
            )


def _draw_grid_interval_legend(ax: mpl.axes.Axes, interval: float) -> None:
    """Place a compact scale key at the lower-left of an individual panel."""
    ax.plot(
        [0.035, 0.135], [0.040, 0.040],
        transform=ax.transAxes,
        color=GRID_COLOR,
        linewidth=2.0,
        linestyle=(0, (3.2, 2.4)),
        alpha=0.95,
        clip_on=False,
        zorder=12,
    )
    for endpoint in (0.035, 0.135):
        ax.plot(
            [endpoint, endpoint],
            [0.040, 0.070],
            transform=ax.transAxes,
            color=GRID_COLOR,
            linewidth=2.0,
            alpha=0.95,
            solid_capstyle="butt",
            clip_on=False,
            zorder=12,
        )
    ax.text(
        0.15, 0.055, f"{interval:g}%",
        transform=ax.transAxes,
        ha="left", va="center",
        fontsize=10.0,
        fontweight="bold",
        color=TEXT_COLOR,
        clip_on=False,
        zorder=12,
    )


def _draw_monthly_bars(ax: mpl.axes.Axes, data: ProfitIncrementData) -> None:
    """Draw 36 bars whose values extend inward from the 0% outer baseline."""
    count = len(data.monthly_percent)
    sector_width = 2.0 * np.pi / count
    angles = np.arange(count, dtype=float) * sector_width + sector_width / 2.0
    bar_width = sector_width * 0.78
    bottoms = BAR_OUTER_RADIUS - data.monthly_percent

    ax.bar(
        angles,
        data.monthly_percent,
        width=bar_width,
        bottom=bottoms,
        align="center",
        color=BLUE,
        edgecolor=BAR_EDGE_COLOR,
        linewidth=0.35,
        alpha=0.94,
        zorder=3,
    )

    month_labels = ("JAN", "FEB", "MAR", "APR", "MAY", "JUN",
                    "JUL", "AUG", "SEP", "OCT", "NOV", "DEC")
    for index, (angle, code) in enumerate(zip(angles, data.months)):
        angle_degrees = float(np.rad2deg(angle))
        ax.text(
            angle,
            MONTH_LABEL_RADIUS,
            month_labels[index % 12],
            ha="center",
            va="center",
            rotation=_upright_radial_rotation(angle_degrees),
            rotation_mode="anchor",
            fontsize=7.2,
            fontweight="semibold",
            color=TEXT_COLOR,
            zorder=8,
        )


def _draw_annual_annulus(ax: mpl.axes.Axes, data: ProfitIncrementData) -> None:
    """Draw the complete three-part annual summary ring and direct labels."""
    sector_width = 2.0 * np.pi / 3.0
    ring_height = ANNULUS_OUTER_RADIUS - ANNULUS_INNER_RADIUS
    centers = np.arange(3, dtype=float) * sector_width + sector_width / 2.0

    ax.bar(
        centers,
        np.full(3, ring_height),
        width=np.full(3, sector_width),
        bottom=np.full(3, ANNULUS_INNER_RADIUS),
        align="center",
        color=ANNULUS_BLUE,
        edgecolor="white",
        linewidth=1.35,
        zorder=2,
    )

    label_radius = (ANNULUS_INNER_RADIUS + ANNULUS_OUTER_RADIUS) / 2.0
    label_layout = ((60.0, 1.0), (180.0, -1.0), (300.0, 1.0))
    for year_index, ((center_degrees, reading_direction), year) in enumerate(
        zip(label_layout, YEARS)
    ):
        annual_months = data.monthly_percent[year_index * 12 : (year_index + 1) * 12]
        annual_label = (
            f"{year}: {annual_months.min():.2f}–{annual_months.max():.2f}%, "
            f"Annual. {data.annual_percent[year]:.2f}%"
        )
        _draw_curved_text(
            ax=ax,
            text=annual_label,
            center_angle_degrees=center_degrees,
            radius=label_radius,
            fontsize=ANNUAL_LABEL_FONTSIZE,
            color=TEXT_COLOR,
            span_degrees=ANNUAL_TEXT_SPAN_DEGREES,
            fontweight="bold",
            direction=reading_direction,
            outward_facing=year in (2023, 2025),
        )

    theta = np.linspace(0.0, 2.0 * np.pi, 721)
    for radius in (ANNULUS_INNER_RADIUS, ANNULUS_OUTER_RADIUS):
        ax.plot(theta, np.full_like(theta, radius), color=BLUE, lw=0.7, zorder=4)

    for boundary in np.arange(0.0, 2.0 * np.pi, sector_width):
        ax.plot(
            [boundary, boundary],
            [INNER_RADIUS, ANNULUS_OUTER_RADIUS],
            color="white",
            linewidth=1.35,
            solid_capstyle="butt",
            zorder=6,
        )


def draw_profit_increment_panel(
    ax: mpl.axes.Axes,
    data: ProfitIncrementData,
    title: str | None = "Profit increment rate",
) -> None:
    """Draw the profit radial chart on an existing polar axis."""
    ax.set_theta_zero_location("N")
    ax.set_theta_direction(-1)
    ax.set_ylim(0.0, PLOT_LIMIT)
    ax.set_xticks([])
    ax.set_yticks([])
    ax.grid(False)
    ax.spines["polar"].set_visible(False)
    ax.set_facecolor("none")

    _draw_scale(ax)
    _draw_monthly_bars(ax, data)
    _draw_annual_annulus(ax, data)
    _draw_grid_interval_legend(ax, 50.0)

    # Keep a crisp, uncluttered center while retaining the 200% boundary ring.
    center = plt.Circle(
        (0.5, 0.5),
        INNER_RADIUS / PLOT_LIMIT / 2.0,
        transform=ax.transAxes,
        facecolor="none",
        edgecolor=GRID_COLOR,
        linewidth=1.1,
        zorder=5,
    )
    ax.add_artist(center)

    if title is not None:
        ax.set_title(
            title,
            fontsize=TITLE_FONTSIZE,
            fontweight="bold",
            color=TEXT_COLOR,
            pad=8,
        )


def create_figure(data: ProfitIncrementData) -> mpl.figure.Figure:
    """Create the standalone 90 mm square radial figure."""
    width_inches = FIGURE_WIDTH_MM / 25.4
    height_inches = FIGURE_HEIGHT_MM / 25.4
    fig = plt.figure(figsize=(width_inches, height_inches), facecolor="none")
    ax = fig.add_axes([0.035, 0.005, 0.93, 0.86], projection="polar")
    draw_profit_increment_panel(ax, data)
    return fig


def export_figure(fig: mpl.figure.Figure, output_dir: Path) -> None:
    """Export editable vectors plus 720 dpi PNG and compressed TIFF."""
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = output_dir / OUTPUT_STEM
    fig.savefig(stem.with_suffix(".svg"), transparent=True)
    fig.savefig(stem.with_suffix(".png"), dpi=PNG_DPI, transparent=True)
    fig.savefig(
        stem.with_suffix(".tiff"),
        dpi=PNG_DPI,
        transparent=True,
        pil_kwargs={"compression": "tiff_lzw"},
    )
    try:
        fig.savefig(stem.with_suffix(".pdf"), transparent=True)
    except PermissionError:
        print(f"Warning: PDF is open and could not be replaced: {stem.with_suffix('.pdf')}")


def main() -> None:
    configure_matplotlib()
    data = load_profit_increment_data(SOURCE_WORKBOOK)
    figure = create_figure(data)
    export_figure(figure, OUTPUT_DIR)
    plt.close(figure)

    print(f"Monthly observations: {len(data.monthly_percent)}")
    print(
        "Monthly range: "
        f"{data.monthly_percent.min():.2f}% to {data.monthly_percent.max():.2f}%"
    )
    print(
        "Annual rates: "
        + ", ".join(
            f"{year}={data.annual_percent[year]:.2f}%" for year in YEARS
        )
    )
    print(f"Created figures in: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
