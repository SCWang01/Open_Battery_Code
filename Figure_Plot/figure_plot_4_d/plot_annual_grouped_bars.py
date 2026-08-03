"""Create annual grouped-bar figures from the supplied 2023-2025 workbook.

The script exports one figure per year plus a vertically stacked three-year
figure.  Monthly values come from ``Monthly Analysis`` and annual values come
directly from ``Annual Summary``.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re

import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.axes import Axes
from matplotlib.patches import Patch
from matplotlib.ticker import FormatStrFormatter
import numpy as np
from openpyxl import load_workbook


SOURCE_WORKBOOK = Path(__file__).with_name("analysis_202301_202512.xlsx")
OUTPUT_DIR = Path(__file__).with_name("outputs") / "annual_grouped_bars"

STANDALONE_WIDTH_MM = 180
STANDALONE_HEIGHT_MM = 55
STACKED_WIDTH_MM = 180
STACKED_HEIGHT_MM = 120

YEARS = (2023, 2024, 2025)
X_LABELS = tuple(f"{month:02d}" for month in range(1, 13)) + ("annual",)

CARBON_COLOR = "#55B9DF"
COST_COLOR = "#79C85B"
PROFIT_COLOR = "#E99A70"
GRID_COLOR = "#B8B8B8"
SPINE_COLOR = "#555555"

LEFT_LABEL = "Reduction rate of carbon emissions\nand generation cost (%)"
RIGHT_LABEL = "Increase rate of battery profit (%)"

LEGEND_HANDLES = (
    Patch(facecolor=CARBON_COLOR, edgecolor="none", label="Carbon emission reduction rate"),
    Patch(facecolor=COST_COLOR, edgecolor="none", label="Generation cost reduction rate"),
    Patch(facecolor=PROFIT_COLOR, edgecolor="none", label="Battery profit increase rate"),
)


@dataclass(frozen=True)
class YearData:
    year: int
    carbon: np.ndarray
    cost: np.ndarray
    profit: np.ndarray


def configure_matplotlib() -> None:
    """Apply editable, publication-oriented typography and line settings."""
    mpl.rcParams.update(
        {
            # Times New Roman is the confirmed primary face; Arial is a safe
            # fallback for systems on which the requested font is unavailable.
            "font.family": ["Times New Roman", "Arial", "serif"],
            "font.size": 8,
            "axes.linewidth": 0.65,
            "axes.unicode_minus": True,
            "legend.frameon": True,
            "legend.fancybox": False,
            "svg.fonttype": "none",
            "pdf.fonttype": 42,
            "savefig.facecolor": "white",
            "savefig.edgecolor": "white",
        }
    )


def load_year_data(workbook_path: Path) -> dict[int, YearData]:
    """Read all monthly observations and Sheet2 annual summary values."""
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    monthly_sheet = workbook["Monthly Analysis"]
    annual_sheet = workbook["Annual Summary"]

    monthly_headers = {
        str(cell.value).strip().lower(): index
        for index, cell in enumerate(next(monthly_sheet.iter_rows()), start=0)
    }
    required_monthly = {"month", "profit increment", "cost reduction", "rate carbon"}
    missing = required_monthly.difference(monthly_headers)
    if missing:
        raise ValueError(f"Missing Monthly Analysis columns: {sorted(missing)}")

    monthly_by_year: dict[int, list[tuple[int, float, float, float]]] = {
        year: [] for year in YEARS
    }
    for row in monthly_sheet.iter_rows(min_row=2, values_only=True):
        month_code = str(row[monthly_headers["month"]]).strip()
        if len(month_code) != 6 or not month_code.isdigit():
            continue
        year = int(month_code[:4])
        if year not in monthly_by_year:
            continue
        month = int(month_code[4:])
        monthly_by_year[year].append(
            (
                month,
                float(row[monthly_headers["rate carbon"]]) * 100,
                float(row[monthly_headers["cost reduction"]]) * 100,
                float(row[monthly_headers["profit increment"]]) * 100,
            )
        )

    annual_headers = {
        str(cell.value).strip().lower(): index
        for index, cell in enumerate(next(annual_sheet.iter_rows()), start=0)
    }
    required_annual = {
        "annual",
        "profit increment rate",
        "cost reduction rate",
        "carbon reduction rate",
    }
    missing = required_annual.difference(annual_headers)
    if missing:
        raise ValueError(f"Missing Annual Summary columns: {sorted(missing)}")

    annual_by_year: dict[int, tuple[float, float, float]] = {}
    for row in annual_sheet.iter_rows(min_row=2, values_only=True):
        period = str(row[annual_headers["annual"]]).strip()
        match = re.fullmatch(r"(\d{4})-(\d{4})", period)
        if match is None:
            continue
        year = int(match.group(1))
        if year not in YEARS:
            continue
        annual_by_year[year] = (
            float(row[annual_headers["carbon reduction rate"]]) * 100,
            float(row[annual_headers["cost reduction rate"]]) * 100,
            float(row[annual_headers["profit increment rate"]]) * 100,
        )

    result: dict[int, YearData] = {}
    for year in YEARS:
        monthly = sorted(monthly_by_year[year], key=lambda item: item[0])
        observed_months = [item[0] for item in monthly]
        if observed_months != list(range(1, 13)):
            raise ValueError(f"{year} monthly observations are incomplete: {observed_months}")
        if year not in annual_by_year:
            raise ValueError(f"Annual Summary has no row corresponding to {year}")

        annual_carbon, annual_cost, annual_profit = annual_by_year[year]
        result[year] = YearData(
            year=year,
            carbon=np.array([item[1] for item in monthly] + [annual_carbon]),
            cost=np.array([item[2] for item in monthly] + [annual_cost]),
            profit=np.array([item[3] for item in monthly] + [annual_profit]),
        )

    workbook.close()
    return result


def style_axes(ax: Axes, right_ax: Axes, show_x_labels: bool) -> None:
    """Apply the shared dual-axis scale and reference-inspired styling."""
    ax.set_xlim(-0.65, len(X_LABELS) - 0.35)
    ax.set_ylim(-5, 20)
    ax.set_yticks(np.arange(-5, 21, 5))
    ax.yaxis.set_major_formatter(FormatStrFormatter("%d%%"))
    ax.tick_params(axis="y", labelsize=7.5, length=2.5, width=0.6, colors=SPINE_COLOR)

    # Match the visual position of 0% on both axes.  With the left axis at
    # -5%..20%, the zero baseline is 20% above the bottom; -10%..40% gives
    # the right axis the same zero position while retaining all profit values.
    right_ax.set_ylim(-10, 40)
    right_ax.set_yticks(np.arange(-10, 41, 10))
    right_ax.yaxis.set_major_formatter(FormatStrFormatter("%d%%"))
    right_ax.tick_params(
        axis="y", labelsize=7.5, length=2.5, width=0.6, colors=PROFIT_COLOR
    )

    x = np.arange(len(X_LABELS))
    ax.set_xticks(x)
    if show_x_labels:
        ax.set_xticklabels(X_LABELS, fontsize=7.5)
        ax.tick_params(axis="x", length=0, pad=3)
    else:
        ax.set_xticklabels([])
        ax.tick_params(axis="x", length=0)

    ax.yaxis.grid(True, linestyle="--", linewidth=0.55, color=GRID_COLOR, alpha=0.85)
    ax.set_axisbelow(True)
    ax.axhline(0, color="#8B8B8B", linewidth=0.8, zorder=1)

    for side in ("left", "bottom"):
        ax.spines[side].set_color(SPINE_COLOR)
        ax.spines[side].set_linewidth(0.65)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    right_ax.spines["top"].set_visible(False)
    right_ax.spines["left"].set_visible(False)
    right_ax.spines["bottom"].set_visible(False)
    right_ax.spines["right"].set_visible(True)
    right_ax.spines["right"].set_color(PROFIT_COLOR)
    right_ax.spines["right"].set_linewidth(0.65)
    right_ax.patch.set_alpha(0)


def draw_grouped_bars(ax: Axes, data: YearData, show_x_labels: bool) -> Axes:
    """Draw one year's 13 grouped observations and return the right axis."""
    right_ax = ax.twinx()
    x = np.arange(len(X_LABELS), dtype=float)
    width = 0.22

    ax.bar(
        x - width,
        data.carbon,
        width=width,
        color=CARBON_COLOR,
        edgecolor="none",
        alpha=0.86,
        zorder=3,
    )
    ax.bar(
        x,
        data.cost,
        width=width,
        color=COST_COLOR,
        edgecolor="none",
        alpha=0.86,
        zorder=3,
    )
    right_ax.bar(
        x + width,
        data.profit,
        width=width,
        color=PROFIT_COLOR,
        edgecolor="none",
        alpha=0.86,
        zorder=2,
    )
    style_axes(ax, right_ax, show_x_labels=show_x_labels)
    return right_ax


def style_legend(legend: mpl.legend.Legend) -> None:
    frame = legend.get_frame()
    frame.set_facecolor("white")
    frame.set_edgecolor("#999999")
    frame.set_linewidth(0.45)
    frame.set_alpha(1)
    for text in legend.get_texts():
        text.set_color(SPINE_COLOR)


def save_all_formats(fig: mpl.figure.Figure, output_stem: Path) -> None:
    """Export editable vector files and a 600-dpi raster preview."""
    output_stem.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_stem.with_suffix(".svg"), bbox_inches=None)
    fig.savefig(output_stem.with_suffix(".pdf"), bbox_inches=None)
    fig.savefig(output_stem.with_suffix(".png"), dpi=600, bbox_inches=None)


def plot_standalone(data: YearData, output_dir: Path) -> None:
    """Create a compact, self-contained figure for one year."""
    fig, ax = plt.subplots(
        figsize=(STANDALONE_WIDTH_MM / 25.4, STANDALONE_HEIGHT_MM / 25.4)
    )
    draw_grouped_bars(ax, data, show_x_labels=True)

    fig.subplots_adjust(left=0.115, right=0.885, bottom=0.27, top=0.74)
    fig.text(0.5, 0.955, str(data.year), ha="center", va="center", fontsize=8.5)
    fig.text(0.022, 0.505, LEFT_LABEL, ha="center", va="center", rotation=90, fontsize=8)
    fig.text(
        0.978,
        0.505,
        RIGHT_LABEL,
        ha="center",
        va="center",
        rotation=90,
        fontsize=8,
        color=PROFIT_COLOR,
    )
    fig.text(0.5, 0.055, "Month", ha="center", va="center", fontsize=8)

    legend = fig.legend(
        handles=LEGEND_HANDLES,
        loc="center",
        bbox_to_anchor=(0.5, 0.835),
        ncol=3,
        fontsize=7.2,
        handlelength=0.85,
        handletextpad=0.35,
        columnspacing=0.85,
        borderpad=0.3,
    )
    style_legend(legend)

    save_all_formats(fig, output_dir / f"grouped_bars_{data.year}")
    plt.close(fig)


def plot_stacked(all_data: dict[int, YearData], output_dir: Path) -> None:
    """Create the shared-axis 2023-2025 vertical figure."""
    fig, axes = plt.subplots(
        nrows=3,
        ncols=1,
        sharex=True,
        figsize=(STACKED_WIDTH_MM / 25.4, STACKED_HEIGHT_MM / 25.4),
    )
    panel_labels = ("(d.1) 2023", "(d.2) 2024", "(d.3) 2025")

    for index, (ax, year, panel_label) in enumerate(zip(axes, YEARS, panel_labels)):
        draw_grouped_bars(ax, all_data[year], show_x_labels=index == len(axes) - 1)
        ax.text(
            0.008,
            0.90,
            panel_label,
            transform=ax.transAxes,
            ha="left",
            va="top",
            fontsize=8.5,
        )

    fig.subplots_adjust(left=0.115, right=0.885, bottom=0.115, top=0.90, hspace=0.12)
    fig.text(0.022, 0.505, LEFT_LABEL, ha="center", va="center", rotation=90, fontsize=8)
    fig.text(
        0.978,
        0.505,
        RIGHT_LABEL,
        ha="center",
        va="center",
        rotation=90,
        fontsize=8,
        color=PROFIT_COLOR,
    )
    fig.text(0.5, 0.035, "Month", ha="center", va="center", fontsize=8)

    legend = fig.legend(
        handles=LEGEND_HANDLES,
        loc="upper center",
        bbox_to_anchor=(0.5, 0.985),
        ncol=3,
        fontsize=7.2,
        handlelength=0.85,
        handletextpad=0.35,
        columnspacing=0.85,
        borderpad=0.3,
    )
    style_legend(legend)

    save_all_formats(fig, output_dir / "grouped_bars_2023_2025_stacked")
    plt.close(fig)


def main() -> None:
    configure_matplotlib()
    all_data = load_year_data(SOURCE_WORKBOOK)
    for year in YEARS:
        plot_standalone(all_data[year], OUTPUT_DIR)
    plot_stacked(all_data, OUTPUT_DIR)
    print(f"Created figures in: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
