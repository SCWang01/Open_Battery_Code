"""Plot pooled hourly P_ESS-state counts and average CAISO prices.

Run export_hourly_state_counts.py first. This script obtains every selected
period and month from the generated workbook, reads five-minute prices directly
from Data/price, and creates one figure per period. For every clock hour, daily
hourly price means are the independent observations used for the mean and the
two-sided 90% Student-t confidence interval.
"""

from __future__ import annotations

import calendar
import csv
import math
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from matplotlib.patches import Patch
from matplotlib.ticker import MaxNLocator
from openpyxl import load_workbook
from scipy.stats import t


# --------------------------------------------------------------------------- #
# Stable paths and figure contract
# --------------------------------------------------------------------------- #
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = Path(__file__).resolve().parents[2]
PRICE_DIR = PROJECT_ROOT / "Data" / "price"
OUTPUT_DIR = SCRIPT_DIR / "output"
FIGURES_DIR = SCRIPT_DIR / "figures"
STATE_WORKBOOK = OUTPUT_DIR / "P_ESS_state_hourly_counts_selected.xlsx"

SELECTION_SHEET = "Selection"
SELECTION_HEADER = ("period_label", "sheet_name", "output_suffix", "year", "month")
STATE_ORDER = ("MC", "CC", "CD", "NU", "DC", "DD", "MD")
AUDIT_ROW_LABELS = (
    "Valid State Count",
    "Excluded Zero-Limit",
    "Excluded NC",
    "Raw Input Count",
)
EXPECTED_POINTS_PER_DAY_HOUR = 12
CONFIDENCE_LEVEL = 0.90

COLORS = {
    "MC": "#C90000",
    "CC": "#FF6B6B",
    "CD": "#F6B4B4",
    "NU": "#A6A6A6",
    "DC": "#9CC5E5",
    "DD": "#4F95D5",
    "MD": "#1D4B73",
}
PRICE_COLOR = "#B13AA3"
CI_FILL_ALPHA = 0.14
CI_HATCH = "//////"
BAR_WIDTH = 0.52

FIGURE_WIDTH_MM = 180.0
FIGURE_HEIGHT_MM = 64.0
PNG_DPI = 720


@dataclass(frozen=True)
class PeriodSpec:
    """One period reconstructed from the exporter's Selection sheet."""

    label: str
    sheet_name: str
    output_suffix: str
    months: tuple[tuple[int, int], ...]


@dataclass(frozen=True)
class StateCountData:
    """Seven plotted state series and four hourly audit series."""

    counts: dict[str, list[int]]
    valid: list[int]
    excluded_zero_limit: list[int]
    excluded_nc: list[int]
    raw: list[int]


@dataclass(frozen=True)
class HourlyPriceSummary:
    """Mean-price estimate and uncertainty for one clock hour."""

    hour: int
    mean: float
    ci_low: float
    ci_high: float
    standard_deviation: float
    sample_size: int


@dataclass(frozen=True)
class PriceAudit:
    """Included target-month rows and filtered extra rows."""

    included_rows: int
    excluded_rows: int


# --------------------------------------------------------------------------- #
# Workbook input validation
# --------------------------------------------------------------------------- #
def integer_value(value: object, *, context: str) -> int:
    """Validate a non-negative integer stored in an external workbook."""
    if (
        isinstance(value, bool)
        or not isinstance(value, (int, float))
        or not math.isfinite(float(value))
        or float(value) < 0
        or not float(value).is_integer()
    ):
        raise ValueError(f"{context} must be a non-negative integer; found {value!r}.")
    return int(value)


def read_hourly_row(row: tuple[object, ...], *, label: str) -> list[int]:
    """Read exactly 24 audited hourly integer values."""
    values = row[1:25]
    if len(values) != 24:
        raise ValueError(f"Row {label!r} must contain exactly 24 hourly values.")
    return [
        integer_value(value, context=f"Row {label!r}, hour {hour:02d}")
        for hour, value in enumerate(values)
    ]


def read_state_count_sheet(worksheet) -> StateCountData:
    """Read and reconcile one exported period's count table."""
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows)
    expected_header = ("P_ESS_state",) + tuple(
        f"{hour:02d}:00" for hour in range(24)
    )
    if tuple(header[:25]) != expected_header:
        raise ValueError(f"Worksheet {worksheet.title!r} has an invalid hour header.")

    values_by_label: dict[str, list[int]] = {}
    for row in rows:
        label = row[0]
        if label is None:
            continue
        if not isinstance(label, str):
            raise ValueError(
                f"Worksheet {worksheet.title!r}: invalid row label {label!r}."
            )
        if label in values_by_label:
            raise ValueError(
                f"Worksheet {worksheet.title!r}: duplicate row {label!r}."
            )
        values_by_label[label] = read_hourly_row(row, label=label)

    required = STATE_ORDER + AUDIT_ROW_LABELS
    missing = [label for label in required if label not in values_by_label]
    if missing:
        raise ValueError(
            f"Worksheet {worksheet.title!r} is missing rows: {', '.join(missing)}."
        )
    unexpected = [label for label in values_by_label if label not in required]
    if unexpected:
        raise ValueError(
            f"Worksheet {worksheet.title!r} has unexpected rows: "
            f"{', '.join(unexpected)}."
        )

    counts = {state: values_by_label[state] for state in STATE_ORDER}
    valid = values_by_label["Valid State Count"]
    excluded_zero_limit = values_by_label["Excluded Zero-Limit"]
    excluded_nc = values_by_label["Excluded NC"]
    raw = values_by_label["Raw Input Count"]
    for hour in range(24):
        state_sum = sum(counts[state][hour] for state in STATE_ORDER)
        if state_sum != valid[hour]:
            raise ValueError(
                f"Worksheet {worksheet.title!r}, hour {hour:02d}: state sum "
                f"{state_sum} does not equal valid count {valid[hour]}."
            )
        reconciled = valid[hour] + excluded_zero_limit[hour] + excluded_nc[hour]
        if reconciled != raw[hour]:
            raise ValueError(
                f"Worksheet {worksheet.title!r}, hour {hour:02d}: audited count "
                f"{reconciled} does not equal raw input count {raw[hour]}."
            )
    if max(valid) == 0:
        raise ValueError(f"Worksheet {worksheet.title!r} has no valid states.")

    return StateCountData(
        counts=counts,
        valid=valid,
        excluded_zero_limit=excluded_zero_limit,
        excluded_nc=excluded_nc,
        raw=raw,
    )


def read_periods_and_counts() -> list[tuple[PeriodSpec, StateCountData]]:
    """Read all period definitions and their state counts from one workbook."""
    if not STATE_WORKBOOK.is_file():
        raise FileNotFoundError(
            f"State workbook not found: {STATE_WORKBOOK}. "
            "Run export_hourly_state_counts.py first."
        )

    workbook = load_workbook(STATE_WORKBOOK, read_only=True, data_only=True)
    try:
        if SELECTION_SHEET not in workbook.sheetnames:
            raise ValueError(f"Worksheet {SELECTION_SHEET!r} is missing.")
        selection_rows = workbook[SELECTION_SHEET].iter_rows(values_only=True)
        header = next(selection_rows)
        if tuple(header[:5]) != SELECTION_HEADER:
            raise ValueError("Selection worksheet has an invalid header.")

        grouped: dict[str, dict[str, object]] = {}
        for excel_row, row in enumerate(selection_rows, start=2):
            if all(value is None for value in row):
                continue
            label, sheet_name, output_suffix, year_value, month_value = row[:5]
            if not all(
                isinstance(value, str) and value.strip()
                for value in (label, sheet_name, output_suffix)
            ):
                raise ValueError(
                    f"Selection row {excel_row}: label, sheet and suffix must be text."
                )
            year = integer_value(year_value, context=f"Selection row {excel_row} year")
            month = integer_value(
                month_value, context=f"Selection row {excel_row} month"
            )
            if year < 1 or month not in range(1, 13):
                raise ValueError(
                    f"Selection row {excel_row}: invalid month {year}-{month:02d}."
                )

            entry = grouped.setdefault(
                label,
                {
                    "sheet_name": sheet_name,
                    "output_suffix": output_suffix,
                    "months": [],
                },
            )
            if (
                entry["sheet_name"] != sheet_name
                or entry["output_suffix"] != output_suffix
            ):
                raise ValueError(
                    f"Selection period {label!r} has inconsistent sheet or suffix."
                )
            month_key = (year, month)
            months = entry["months"]
            if month_key in months:
                raise ValueError(
                    f"Selection period {label!r} repeats {year}-{month:02d}."
                )
            months.append(month_key)

        if not grouped:
            raise ValueError("Selection worksheet contains no periods.")

        result: list[tuple[PeriodSpec, StateCountData]] = []
        for label, entry in grouped.items():
            sheet_name = str(entry["sheet_name"])
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Selection period {label!r} references missing sheet "
                    f"{sheet_name!r}."
                )
            period = PeriodSpec(
                label=label,
                sheet_name=sheet_name,
                output_suffix=str(entry["output_suffix"]),
                months=tuple(entry["months"]),
            )
            result.append((period, read_state_count_sheet(workbook[sheet_name])))
        return result
    finally:
        workbook.close()


# --------------------------------------------------------------------------- #
# Price validation and statistical derivation
# --------------------------------------------------------------------------- #
def price_path(year: int, month: int) -> Path:
    """Return the monthly CAISO five-minute price CSV."""
    return PRICE_DIR / f"{year}{month:02d} CAISO Average Price.csv"


def parse_timestamp(value: object, *, source: Path, row: int) -> datetime:
    """Parse the timestamp formats used by the monthly CAISO CSV files."""
    if not isinstance(value, str):
        raise ValueError(f"{source.name} row {row}: date must be text.")
    for date_format in ("%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %H:%M:%S"):
        try:
            return datetime.strptime(value.strip(), date_format)
        except ValueError:
            pass
    raise ValueError(f"{source.name} row {row}: unsupported timestamp {value!r}.")


def read_daily_hourly_prices(
    period: PeriodSpec,
) -> tuple[dict[int, list[float]], PriceAudit]:
    """Read selected calendar rows and return daily-hourly price observations."""
    grouped: dict[tuple[int, int, int, int], list[float]] = defaultdict(list)
    included_rows = 0
    excluded_rows = 0

    for year, month in period.months:
        source = price_path(year, month)
        if not source.is_file():
            raise FileNotFoundError(f"Price CSV not found: {source}")
        seen_timestamps: set[datetime] = set()
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            if reader.fieldnames != ["date", "price"]:
                raise ValueError(
                    f"{source.name}: expected exactly the columns 'date' and 'price'."
                )
            for csv_row, row in enumerate(reader, start=2):
                timestamp = parse_timestamp(row["date"], source=source, row=csv_row)
                if (timestamp.year, timestamp.month) != (year, month):
                    excluded_rows += 1
                    continue
                try:
                    price = float(row["price"])
                except (TypeError, ValueError) as exc:
                    raise ValueError(
                        f"{source.name} row {csv_row}: invalid price {row['price']!r}."
                    ) from exc
                if not math.isfinite(price):
                    raise ValueError(
                        f"{source.name} row {csv_row}: price must be finite."
                    )
                if timestamp in seen_timestamps:
                    raise ValueError(
                        f"{source.name} row {csv_row}: duplicate timestamp "
                        f"{timestamp!s}."
                    )
                seen_timestamps.add(timestamp)
                grouped[(year, month, timestamp.day, timestamp.hour)].append(price)
                included_rows += 1

        expected_month_rows = (
            calendar.monthrange(year, month)[1]
            * 24
            * EXPECTED_POINTS_PER_DAY_HOUR
        )
        if len(seen_timestamps) != expected_month_rows:
            raise ValueError(
                f"{source.name}: found {len(seen_timestamps)} unique target-month "
                f"timestamps; expected {expected_month_rows}."
            )

    daily_hourly: dict[int, list[float]] = {hour: [] for hour in range(24)}
    for year, month in period.months:
        days = calendar.monthrange(year, month)[1]
        for day in range(1, days + 1):
            for hour in range(24):
                values = grouped[(year, month, day, hour)]
                if len(values) != EXPECTED_POINTS_PER_DAY_HOUR:
                    raise ValueError(
                        f"{year}-{month:02d}-{day:02d} {hour:02d}:00 has "
                        f"{len(values)} price points; expected "
                        f"{EXPECTED_POINTS_PER_DAY_HOUR}."
                    )
                daily_hourly[hour].append(sum(values) / len(values))

    return daily_hourly, PriceAudit(included_rows, excluded_rows)


def summarize_prices(
    daily_hourly: dict[int, list[float]],
) -> list[HourlyPriceSummary]:
    """Calculate two-sided 90% Student-t intervals from daily-hourly means."""
    summaries: list[HourlyPriceSummary] = []
    alpha = 1.0 - CONFIDENCE_LEVEL
    expected_n = len(daily_hourly[0])
    if expected_n < 2:
        raise ValueError("At least two selected calendar days are required for a CI.")

    for hour in range(24):
        values = daily_hourly[hour]
        if len(values) != expected_n:
            raise ValueError(
                f"Hour {hour:02d}: found {len(values)} daily means; "
                f"expected {expected_n}."
            )
        mean = sum(values) / expected_n
        variance = sum((value - mean) ** 2 for value in values) / (expected_n - 1)
        standard_deviation = math.sqrt(variance)
        critical_value = float(
            t.ppf(1.0 - alpha / 2.0, df=expected_n - 1)
        )
        margin = critical_value * standard_deviation / math.sqrt(expected_n)
        summaries.append(
            HourlyPriceSummary(
                hour=hour,
                mean=mean,
                ci_low=mean - margin,
                ci_high=mean + margin,
                standard_deviation=standard_deviation,
                sample_size=expected_n,
            )
        )
    return summaries


# --------------------------------------------------------------------------- #
# Source-data and figure assembly
# --------------------------------------------------------------------------- #
def configure_matplotlib() -> None:
    """Configure publication typography and editable vector text."""
    mpl.rcParams.update(
        {
            "font.family": ["Times New Roman", "Arial", "DejaVu Serif"],
            "font.serif": ["Times New Roman", "Times", "DejaVu Serif"],
            "font.size": 9,
            "axes.labelsize": 10,
            "xtick.labelsize": 8.5,
            "ytick.labelsize": 8.5,
            "legend.fontsize": 8.5,
            "axes.linewidth": 0.7,
            "axes.unicode_minus": False,
            "svg.fonttype": "none",
            "pdf.fonttype": 42,
            "hatch.linewidth": 0.55,
            "savefig.transparent": True,
        }
    )


def nice_count_scale(maximum: int) -> tuple[int, list[int]]:
    """Return a data-driven count-axis ceiling with about 10% headroom."""
    if maximum <= 0:
        raise ValueError(f"Count-axis maximum must be positive; found {maximum}.")
    raw_step = maximum / 7.0
    magnitude = 1.0
    if raw_step >= 1.0:
        while magnitude * 10.0 <= raw_step:
            magnitude *= 10.0
    else:
        while magnitude > raw_step:
            magnitude /= 10.0
    step = min(
        (factor * magnitude for factor in (1, 2, 2.5, 5, 10)),
        key=lambda candidate: abs(candidate - raw_step),
    )
    integer_step = max(1, int(math.ceil(step)))
    upper = int(math.ceil(maximum / integer_step) * integer_step)
    if upper < maximum * 1.08:
        upper += integer_step
    return upper, list(range(0, upper + 1, integer_step))


def price_axis_limits(summaries: list[HourlyPriceSummary]) -> tuple[float, float]:
    """Cover the complete CI extent with 10% data-relative headroom."""
    minimum = min(summary.ci_low for summary in summaries)
    maximum = max(summary.ci_high for summary in summaries)
    span = maximum - minimum
    if math.isclose(span, 0.0, rel_tol=0.0, abs_tol=1e-12):
        span = max(abs(minimum), 1.0)
    padding = 0.10 * span
    return minimum - padding, maximum + padding


def output_basename(period: PeriodSpec) -> str:
    """Match the reference plot's period-suffix naming convention."""
    return f"hourly_state_counts_and_price_{period.output_suffix}"


def write_source_data(
    period: PeriodSpec,
    state_data: StateCountData,
    summaries: list[HourlyPriceSummary],
) -> Path:
    """Write exact plotted values plus state-exclusion audit columns."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / f"{output_basename(period)}_source_data.csv"
    fieldnames = [
        "period_label",
        "selected_months",
        "hour",
        *[f"count_{state}" for state in STATE_ORDER],
        "state_rows_raw",
        "state_rows_valid",
        "state_rows_excluded_zero_limit",
        "state_rows_excluded_nc",
        "average_price_usd_per_mwh",
        "ci90_lower_usd_per_mwh",
        "ci90_upper_usd_per_mwh",
        "daily_mean_sd_usd_per_mwh",
        "n_days",
        "five_minute_points_per_day_hour",
    ]
    selected_months = ";".join(
        f"{year}-{month:02d}" for year, month in period.months
    )
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for summary in summaries:
            hour = summary.hour
            writer.writerow(
                {
                    "period_label": period.label,
                    "selected_months": selected_months,
                    "hour": hour,
                    **{
                        f"count_{state}": state_data.counts[state][hour]
                        for state in STATE_ORDER
                    },
                    "state_rows_raw": state_data.raw[hour],
                    "state_rows_valid": state_data.valid[hour],
                    "state_rows_excluded_zero_limit": (
                        state_data.excluded_zero_limit[hour]
                    ),
                    "state_rows_excluded_nc": state_data.excluded_nc[hour],
                    "average_price_usd_per_mwh": f"{summary.mean:.8f}",
                    "ci90_lower_usd_per_mwh": f"{summary.ci_low:.8f}",
                    "ci90_upper_usd_per_mwh": f"{summary.ci_high:.8f}",
                    "daily_mean_sd_usd_per_mwh": (
                        f"{summary.standard_deviation:.8f}"
                    ),
                    "n_days": summary.sample_size,
                    "five_minute_points_per_day_hour": (
                        EXPECTED_POINTS_PER_DAY_HOUR
                    ),
                }
            )
    return output_path


def build_figure(
    state_data: StateCountData,
    summaries: list[HourlyPriceSummary],
) -> plt.Figure:
    """Create the transparent stacked-bar and price-confidence-band figure."""
    configure_matplotlib()
    mm_per_inch = 25.4
    figure, count_axis = plt.subplots(
        figsize=(FIGURE_WIDTH_MM / mm_per_inch, FIGURE_HEIGHT_MM / mm_per_inch),
        facecolor="none",
    )
    figure.subplots_adjust(left=0.075, right=0.905, bottom=0.22, top=0.82)
    count_axis.set_facecolor("none")
    price_axis = count_axis.twinx()
    price_axis.set_facecolor("none")

    hours = list(range(24))
    bottoms = [0] * 24
    for state in STATE_ORDER:
        values = state_data.counts[state]
        count_axis.bar(
            hours,
            values,
            width=BAR_WIDTH,
            bottom=bottoms,
            color=COLORS[state],
            edgecolor="none",
            linewidth=0,
            zorder=2,
        )
        bottoms = [bottom + value for bottom, value in zip(bottoms, values)]

    means = [summary.mean for summary in summaries]
    lower = [summary.ci_low for summary in summaries]
    upper = [summary.ci_high for summary in summaries]
    price_axis.fill_between(
        hours,
        lower,
        upper,
        facecolor=mpl.colors.to_rgba(PRICE_COLOR, CI_FILL_ALPHA),
        edgecolor=PRICE_COLOR,
        hatch=CI_HATCH,
        linewidth=0.0,
        zorder=3,
    )
    price_axis.plot(
        hours,
        means,
        color=PRICE_COLOR,
        linewidth=1.15,
        marker="D",
        markersize=3.0,
        markerfacecolor="white",
        markeredgecolor=PRICE_COLOR,
        markeredgewidth=0.8,
        zorder=4,
    )

    count_upper, count_ticks = nice_count_scale(max(state_data.valid))
    count_axis.set_xlim(-0.7, 23.7)
    count_axis.set_ylim(0, count_upper)
    count_axis.set_xticks(hours)
    count_axis.set_yticks(count_ticks)
    count_axis.set_xlabel("Time (hours)")
    count_axis.set_ylabel("Count")
    count_axis.tick_params(axis="both", direction="out", length=2.8, width=0.65)

    price_axis.set_ylim(*price_axis_limits(summaries))
    price_axis.yaxis.set_major_locator(MaxNLocator(nbins=6))
    price_axis.set_ylabel("Price (USD/MWh)", color=PRICE_COLOR)
    price_axis.tick_params(
        axis="y",
        colors=PRICE_COLOR,
        direction="out",
        length=2.8,
        width=0.65,
    )

    frame_color = "#B7B7B7"
    for axis in (count_axis, price_axis):
        for spine in axis.spines.values():
            spine.set_color(frame_color)
            spine.set_linewidth(0.65)
    count_axis.spines["right"].set_visible(False)
    price_axis.spines["left"].set_visible(False)

    figure.legend(
        handles=[
            Patch(
                facecolor=mpl.colors.to_rgba(PRICE_COLOR, CI_FILL_ALPHA),
                edgecolor=PRICE_COLOR,
                hatch=CI_HATCH,
                linewidth=0,
                label="90% Confidence interval",
            ),
            Line2D(
                [0],
                [0],
                color=PRICE_COLOR,
                linewidth=1.15,
                marker="D",
                markersize=3.0,
                markerfacecolor="white",
                markeredgecolor=PRICE_COLOR,
                markeredgewidth=0.8,
                label="Average price",
            ),
        ],
        loc="upper center",
        bbox_to_anchor=(0.5, 0.985),
        ncol=2,
        frameon=False,
        handlelength=1.6,
        handletextpad=0.35,
        columnspacing=5.5,
        borderaxespad=0,
    )
    return figure


def save_figure(figure: plt.Figure, period: PeriodSpec) -> list[Path]:
    """Export one transparent figure in PNG, SVG and PDF formats."""
    FIGURES_DIR.mkdir(parents=True, exist_ok=True)
    basename = output_basename(period)
    base = FIGURES_DIR / basename
    description = (
        f"{period.label} hourly state counts and average prices "
        f"for {len(period.months)} selected months"
    )
    paths = [
        base.with_suffix(".png"),
        base.with_suffix(".svg"),
        base.with_suffix(".pdf"),
    ]
    figure.savefig(
        paths[0],
        dpi=PNG_DPI,
        transparent=True,
        metadata={"Title": description},
    )
    figure.savefig(
        paths[1],
        transparent=True,
        metadata={"Creator": "Matplotlib", "Description": description},
    )
    figure.savefig(
        paths[2],
        transparent=True,
        metadata={"Creator": "Matplotlib", "Title": description},
    )
    return paths


def main() -> None:
    for period, state_data in read_periods_and_counts():
        daily_hourly, price_audit = read_daily_hourly_prices(period)
        summaries = summarize_prices(daily_hourly)
        source_data_path = write_source_data(period, state_data, summaries)
        figure = build_figure(state_data, summaries)
        figure_paths = save_figure(figure, period)
        plt.close(figure)

        print(f"Period: {period.label}")
        print(
            "  Months: "
            + ", ".join(f"{year}-{month:02d}" for year, month in period.months)
        )
        print(f"  State rows: raw={sum(state_data.raw)}")
        print(f"  State rows: valid={sum(state_data.valid)}")
        print(
            "  State rows: excluded zero-limit="
            f"{sum(state_data.excluded_zero_limit)}"
        )
        print(f"  State rows: excluded NC={sum(state_data.excluded_nc)}")
        print(f"  Price rows included={price_audit.included_rows}")
        print(f"  Extra price rows filtered={price_audit.excluded_rows}")
        print(
            f"  Confidence level={CONFIDENCE_LEVEL:.0%}; "
            f"n={summaries[0].sample_size} days; "
            f"df={summaries[0].sample_size - 1}"
        )
        print(f"  Source data: {source_data_path}")
        for path in figure_paths:
            print(f"  Figure: {path}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        raise SystemExit(f"Error: {exc}") from None
