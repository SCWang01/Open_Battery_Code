"""Export hourly P_ESS-state counts for selected calendar months.

Edit SELECTION_MATRIX to pool an arbitrary set of months from SELECTION_YEARS.
Set PLOT_SEASONS to True to export Winter, Spring, Summer and Autumn as four
independent periods. Classified workbooks are read directly from
Results/Bidding; no copied input folder is used.

Rows intentionally left unclassified because either power limit is zero are
excluded from state counts and audited by hour. NC is also excluded and
reported separately. Unknown non-empty labels remain hard errors.
"""

from __future__ import annotations

import calendar
import math
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# --------------------------------------------------------------------------- #
# Editable selection and stable data contract
# --------------------------------------------------------------------------- #
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = Path(__file__).resolve().parents[2]
BIDDING_DIR = PROJECT_ROOT / "Results" / "Bidding"
OUTPUT_DIR = SCRIPT_DIR / "output"
OUTPUT = OUTPUT_DIR / "P_ESS_state_hourly_counts_selected.xlsx"

SELECTION_YEARS = (2023, 2024, 2025)
SELECTION_MATRIX: list[list[int]] = [
    [1, 1, 1],  # January
    [1, 1, 1],  # February
    [1, 1, 1],  # March
    [1, 1, 1],  # April
    [1, 1, 1],  # May
    [1, 1, 1],  # June
    [1, 1, 1],  # July
    [1, 1, 1],  # August
    [1, 1, 1],  # September
    [1, 1, 1],  # October
    [1, 1, 1],  # November
    [1, 1, 1],  # December
]

# False: export the custom matrix above. True: export all four seasons for
# every year in SELECTION_YEARS, matching plot_state_distribution_pies.py.
PLOT_SEASONS = False

SEASON_MONTHS = {
    "Winter": frozenset((12, 1, 2)),
    "Spring": frozenset((3, 4, 5)),
    "Summer": frozenset((6, 7, 8)),
    "Autumn": frozenset((9, 10, 11)),
}

PLOTTED_STATE_ORDER = ("MC", "CC", "CD", "NU", "DC", "DD", "MD")
KNOWN_STATE_ORDER = PLOTTED_STATE_ORDER + ("NC",)
SELECTION_SHEET = "Selection"
COUNT_SHEET_PREFIX = "Hourly Counts - "
SELECTION_HEADER = ("period_label", "sheet_name", "output_suffix", "year", "month")
AUDIT_ROW_LABELS = (
    "Valid State Count",
    "Excluded Zero-Limit",
    "Excluded NC",
    "Raw Input Count",
)
ZERO_ABS_TOL = 1e-9


@dataclass(frozen=True)
class PeriodSpec:
    """One output period and the exact calendar months pooled into it."""

    label: str
    months: tuple[tuple[int, int], ...]
    output_suffix: str

    @property
    def sheet_name(self) -> str:
        return f"{COUNT_SHEET_PREFIX}{self.label}"


@dataclass
class HourlyStateTally:
    """Hourly state counts plus explicit exclusion accounting."""

    counts: dict[int, Counter[str]] = field(
        default_factory=lambda: {hour: Counter() for hour in range(24)}
    )
    valid: Counter[int] = field(default_factory=Counter)
    excluded_zero_limit: Counter[int] = field(default_factory=Counter)
    excluded_nc: Counter[int] = field(default_factory=Counter)
    raw: Counter[int] = field(default_factory=Counter)

    def validate(self) -> None:
        """Ensure every hourly audit identity reconciles exactly."""
        for hour in range(24):
            counted = sum(self.counts[hour][state] for state in PLOTTED_STATE_ORDER)
            if counted != self.valid[hour]:
                raise ValueError(
                    f"Hour {hour:02d}: state sum {counted} does not equal "
                    f"valid count {self.valid[hour]}."
                )
            reconciled = (
                self.valid[hour]
                + self.excluded_zero_limit[hour]
                + self.excluded_nc[hour]
            )
            if reconciled != self.raw[hour]:
                raise ValueError(
                    f"Hour {hour:02d}: audited rows {reconciled} do not equal "
                    f"raw rows {self.raw[hour]}."
                )


# --------------------------------------------------------------------------- #
# Selection validation and derivation
# --------------------------------------------------------------------------- #
def month_name(year: int, month: int) -> str:
    """Return the workbook/sheet month token, for example May2025."""
    return f"{calendar.month_name[month]}{year}"


def workbook_path(year: int, month: int) -> Path:
    """Return the classified-width workbook for one calendar month."""
    return BIDDING_DIR / (
        f"dsfunction_{month_name(year, month)}_exact_V5_k20_classified_width.xlsx"
    )


def select_months(
    matrix: list[list[int]] = SELECTION_MATRIX,
    years: tuple[int, ...] = SELECTION_YEARS,
) -> tuple[tuple[int, int], ...]:
    """Validate and expand the 12-by-year month-selection matrix."""
    if len(matrix) != 12:
        raise ValueError(f"SELECTION_MATRIX must have 12 rows; found {len(matrix)}.")
    for month_index, row in enumerate(matrix, start=1):
        if len(row) != len(years):
            raise ValueError(
                f"SELECTION_MATRIX row {month_index} "
                f"({calendar.month_name[month_index]}) must have {len(years)} entries."
            )
        for year_index, value in enumerate(row):
            if value not in (0, 1):
                raise ValueError(
                    f"SELECTION_MATRIX[{month_index - 1}][{year_index}] for "
                    f"{years[year_index]}-{month_index:02d} must be 0 or 1; "
                    f"found {value!r}."
                )

    selected = tuple(
        (year, month)
        for month in range(1, 13)
        for year_index, year in enumerate(years)
        if matrix[month - 1][year_index] == 1
    )
    if not selected:
        raise ValueError("SELECTION_MATRIX selects no months.")
    return selected


def season_selection_matrix(
    season: str, years: tuple[int, ...] = SELECTION_YEARS
) -> list[list[int]]:
    """Build the reference script's same-year meteorological-season matrix."""
    try:
        selected_months = SEASON_MONTHS[season]
    except KeyError as exc:
        raise ValueError(f"Unknown season: {season!r}.") from exc
    return [
        [1 if month in selected_months else 0 for _year in years]
        for month in range(1, 13)
    ]


def per_year_counts(
    months: tuple[tuple[int, int], ...],
    years: tuple[int, ...] = SELECTION_YEARS,
) -> dict[int, int]:
    """Count selected months for every configured year, including zeros."""
    counts = {year: 0 for year in years}
    for year, _month in months:
        if year not in counts:
            raise ValueError(f"Selected year {year} is not in SELECTION_YEARS.")
        counts[year] += 1
    return counts


def season_label(
    months: tuple[tuple[int, int], ...],
    years: tuple[int, ...] = SELECTION_YEARS,
) -> str | None:
    """Return the shared season name when every configured year matches it."""
    selected_by_year = {
        year: {month for selected_year, month in months if selected_year == year}
        for year in years
    }
    for label, season_months in SEASON_MONTHS.items():
        if all(selected_by_year[year] == season_months for year in years):
            return label
    return None


def period_basename_suffix(
    months: tuple[tuple[int, int], ...],
    years: tuple[int, ...] = SELECTION_YEARS,
) -> str:
    """Match the reference figure's count-based filename convention."""
    counts = per_year_counts(months, years)
    suffix = "_".join(f"{year}_{counts[year]}" for year in years)
    label = season_label(months, years)
    return f"{suffix}_{label}" if label is not None else suffix


def make_period(label: str, months: tuple[tuple[int, int], ...]) -> PeriodSpec:
    """Create one normalized period without repeating suffix derivation."""
    return PeriodSpec(
        label=label,
        months=months,
        output_suffix=period_basename_suffix(months),
    )


def selected_periods() -> tuple[PeriodSpec, ...]:
    """Derive either one custom period or four season periods."""
    if PLOT_SEASONS:
        periods: list[PeriodSpec] = []
        for season in SEASON_MONTHS:
            months = select_months(season_selection_matrix(season))
            periods.append(make_period(season, months))
        return tuple(periods)

    months = select_months()
    return (make_period("Custom", months),)


# --------------------------------------------------------------------------- #
# Source validation and aggregation
# --------------------------------------------------------------------------- #
def find_unique_header(headers: tuple[object, ...], name: str, source: Path) -> int:
    """Return a zero-based index for one required, unique column."""
    matches = [index for index, value in enumerate(headers) if value == name]
    if len(matches) != 1:
        raise ValueError(
            f"{source.name}: header {name!r} occurs {len(matches)} times; "
            "exactly one is required."
        )
    return matches[0]


def finite_number(value: object, *, field_name: str, source: Path, row: int) -> float:
    """Validate one finite numeric workbook value at the input boundary."""
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(
            f"{source.name} row {row}: {field_name} must be numeric; found {value!r}."
        )
    normalized = float(value)
    if not math.isfinite(normalized):
        raise ValueError(
            f"{source.name} row {row}: {field_name} must be finite; found {value!r}."
        )
    return normalized


def read_one_month(year: int, month: int, tally: HourlyStateTally) -> None:
    """Validate and pool one classified workbook into an hourly tally."""
    source = workbook_path(year, month)
    if not source.is_file():
        raise FileNotFoundError(f"Source workbook not found: {source}")

    sheet_name = month_name(year, month)
    workbook = load_workbook(source, read_only=True, data_only=True)
    seen_timestamps: set[datetime] = set()
    nonempty_rows = 0
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"{source.name}: worksheet {sheet_name!r} not found; available: "
                f"{', '.join(workbook.sheetnames)}."
            )
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows)
        time_col = find_unique_header(headers, "time", source)
        state_col = find_unique_header(headers, "P_ESS_state", source)
        pcmax_col = find_unique_header(headers, "Pcmax", source)
        pdmax_col = find_unique_header(headers, "Pdmax", source)
        width_col = find_unique_header(headers, "dsfunction_state_width", source)

        for excel_row, row in enumerate(rows, start=2):
            if all(value is None for value in row):
                continue
            nonempty_rows += 1
            timestamp = row[time_col]
            if not isinstance(timestamp, datetime):
                raise ValueError(
                    f"{source.name} row {excel_row}: invalid time {timestamp!r}."
                )
            if (timestamp.year, timestamp.month) != (year, month):
                raise ValueError(
                    f"{source.name} row {excel_row}: time {timestamp!s} lies outside "
                    f"{year}-{month:02d}."
                )
            if timestamp in seen_timestamps:
                raise ValueError(
                    f"{source.name} row {excel_row}: duplicate time {timestamp!s}."
                )
            seen_timestamps.add(timestamp)

            hour = timestamp.hour
            state = row[state_col]
            tally.raw[hour] += 1

            if state is None or (isinstance(state, str) and not state.strip()):
                pcmax = finite_number(
                    row[pcmax_col], field_name="Pcmax", source=source, row=excel_row
                )
                pdmax = finite_number(
                    row[pdmax_col], field_name="Pdmax", source=source, row=excel_row
                )
                is_zero_limit = math.isclose(
                    pcmax, 0.0, rel_tol=0.0, abs_tol=ZERO_ABS_TOL
                ) or math.isclose(pdmax, 0.0, rel_tol=0.0, abs_tol=ZERO_ABS_TOL)
                if not is_zero_limit or row[width_col] is not None:
                    raise ValueError(
                        f"{source.name} row {excel_row}: blank P_ESS_state is only "
                        "valid for a zero power limit with blank dsfunction_state_width."
                    )
                tally.excluded_zero_limit[hour] += 1
                continue

            if state == "NC":
                tally.excluded_nc[hour] += 1
                continue
            if state not in PLOTTED_STATE_ORDER:
                raise ValueError(
                    f"{source.name} row {excel_row}: unknown P_ESS_state {state!r}; "
                    f"expected one of {', '.join(KNOWN_STATE_ORDER)} or a verified "
                    "zero-limit blank."
                )

            tally.counts[hour][state] += 1
            tally.valid[hour] += 1
    finally:
        workbook.close()

    expected_rows = calendar.monthrange(year, month)[1] * 24
    if nonempty_rows != expected_rows or len(seen_timestamps) != expected_rows:
        raise ValueError(
            f"{source.name}: expected {expected_rows} unique hourly rows for "
            f"{year}-{month:02d}; found {nonempty_rows} rows and "
            f"{len(seen_timestamps)} unique timestamps."
        )


def aggregate_period(period: PeriodSpec) -> HourlyStateTally:
    """Pool all raw hourly state observations for one selected period."""
    tally = HourlyStateTally()
    for year, month in period.months:
        read_one_month(year, month, tally)
    tally.validate()
    return tally


# --------------------------------------------------------------------------- #
# Workbook assembly
# --------------------------------------------------------------------------- #
def apply_sheet_style(worksheet, data_end_row: int) -> None:
    """Apply the compact, auditable table style used by the prior exporter."""
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "B2"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    state_fill = PatternFill("solid", fgColor="D9EAF7")
    audit_fill = PatternFill("solid", fgColor="E2F0D9")
    body_font = Font(name="Calibri", size=11, color="1F1F1F")
    thin_blue = Side(style="thin", color="9EBDD3")
    medium_blue = Side(style="medium", color="4472C4")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=medium_blue)

    for row in worksheet.iter_rows(
        min_row=2, max_row=data_end_row, min_col=1, max_col=25
    ):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.column > 1:
                cell.number_format = "0"
        row[0].font = Font(name="Calibri", size=11, bold=True, color="1F1F1F")
        row[0].border = Border(right=thin_blue)
        row[0].fill = (
            state_fill if row[0].value in PLOTTED_STATE_ORDER else audit_fill
        )

    first_audit_row = len(PLOTTED_STATE_ORDER) + 2
    for cell in worksheet[first_audit_row]:
        cell.border = Border(top=medium_blue)

    worksheet.conditional_formatting.add(
        f"B2:Y{len(PLOTTED_STATE_ORDER) + 1}",
        ColorScaleRule(
            start_type="min",
            start_color="FFF2CC",
            mid_type="percentile",
            mid_value=50,
            mid_color="9DC3E6",
            end_type="max",
            end_color="4472C4",
        ),
    )
    worksheet.column_dimensions["A"].width = 24
    for column in range(2, 26):
        worksheet.column_dimensions[get_column_letter(column)].width = 8
    worksheet.row_dimensions[1].height = 24
    for row_index in range(2, data_end_row + 1):
        worksheet.row_dimensions[row_index].height = 21
    worksheet.auto_filter.ref = f"A1:Y{data_end_row}"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True


def add_selection_sheet(workbook: Workbook, periods: tuple[PeriodSpec, ...]) -> None:
    """Write the single source of truth consumed by the plotting script."""
    worksheet = workbook.active
    worksheet.title = SELECTION_SHEET
    worksheet.sheet_view.showGridLines = False
    worksheet.append(SELECTION_HEADER)
    for period in periods:
        for year, month in period.months:
            worksheet.append(
                (period.label, period.sheet_name, period.output_suffix, year, month)
            )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for column, width in zip("ABCDE", (18, 26, 38, 12, 12)):
        worksheet.column_dimensions[column].width = width
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:E{worksheet.max_row}"


def add_count_sheet(
    workbook: Workbook, period: PeriodSpec, tally: HourlyStateTally
) -> None:
    """Write one period's seven plotted states and four audit rows."""
    worksheet = workbook.create_sheet(period.sheet_name)
    worksheet.append(("P_ESS_state", *(f"{hour:02d}:00" for hour in range(24))))
    for state in PLOTTED_STATE_ORDER:
        worksheet.append(
            (state, *(tally.counts[hour][state] for hour in range(24)))
        )

    audit_values = {
        "Valid State Count": tally.valid,
        "Excluded Zero-Limit": tally.excluded_zero_limit,
        "Excluded NC": tally.excluded_nc,
        "Raw Input Count": tally.raw,
    }
    for label in AUDIT_ROW_LABELS:
        worksheet.append(
            (label, *(audit_values[label][hour] for hour in range(24)))
        )

    apply_sheet_style(worksheet, worksheet.max_row)


def build_workbook(
    period_tallies: tuple[tuple[PeriodSpec, HourlyStateTally], ...]
) -> Workbook:
    """Assemble the fixed intermediate workbook for custom or season mode."""
    workbook = Workbook()
    add_selection_sheet(
        workbook, tuple(period for period, _tally in period_tallies)
    )
    for period, tally in period_tallies:
        add_count_sheet(workbook, period, tally)
    return workbook


def main() -> None:
    periods = selected_periods()
    period_tallies = tuple((period, aggregate_period(period)) for period in periods)
    workbook = build_workbook(period_tallies)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)

    print(f"Mode: {'seasons' if PLOT_SEASONS else 'custom selection'}")
    for period, tally in period_tallies:
        selected = ", ".join(f"{year}-{month:02d}" for year, month in period.months)
        print(f"Period: {period.label} -> [{selected}]")
        print(f"  Valid states: {sum(tally.valid.values())}")
        print(
            "  Excluded zero-limit states: "
            f"{sum(tally.excluded_zero_limit.values())}"
        )
        print(f"  Excluded NC states: {sum(tally.excluded_nc.values())}")
    print(f"Exported: {OUTPUT}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        raise SystemExit(f"Error: {exc}") from None
