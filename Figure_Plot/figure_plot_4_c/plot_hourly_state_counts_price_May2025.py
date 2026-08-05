"""Plot hourly P_ESS state counts with May 2025 CAISO prices.

For each clock hour, the twelve 5-minute prices within each day are first
averaged.  The line shows the mean of the resulting 31 daily hourly means, and
the hatched band shows its two-sided 90% Student-t confidence interval.
"""

from __future__ import annotations

import csv
import math
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from matplotlib.patches import Patch
from openpyxl import load_workbook
from scipy.stats import t


SCRIPT_DIR = Path(__file__).resolve().parent
INPUT_DIR = SCRIPT_DIR / "input"
OUTPUT_DIR = SCRIPT_DIR / "output"
FIGURES_DIR = SCRIPT_DIR / "figures"
STATE_WORKBOOK = OUTPUT_DIR / "P_ESS_state_hourly_counts_May2025.xlsx"
PRICE_CSV = INPUT_DIR / "202505 CAISO Average Price.csv"
STATE_SHEET = "Hourly Counts"
OUTPUT_BASENAME = "hourly_state_counts_and_price_May2025"

TARGET_YEAR = 2025
TARGET_MONTH = 5
EXPECTED_DAYS = 31
EXPECTED_POINTS_PER_DAY_HOUR = 12
CONFIDENCE_LEVEL = 0.90

STATE_ORDER = ("MC", "CC", "CD", "NU", "DC", "DD", "MD")
ALL_STATES = STATE_ORDER + ("NC",)
COLORS = {
    "MC": "#C90000",
    "CC": "#FF6B6B",
    "CD": "#F6B4B4",
    "NU": "#A6A6A6",
    "DC": "#9CC5E5",
    "DD": "#4F95D5",
    "MD": "#1D4B73",
}
PRICE_COLOR = "#B13AA3"
CI_FILL_ALPHA = 0.14
CI_HATCH = "//////"
BAR_WIDTH = 0.52

FIGURE_WIDTH_MM = 180.0
FIGURE_HEIGHT_MM = 64.0
PNG_DPI = 720


@dataclass(frozen=True)
class HourlyPriceSummary:
    hour: int
    mean: float
    ci_low: float
    ci_high: float
    standard_deviation: float
    sample_size: int


def configure_matplotlib() -> None:
    """Configure publication typography and editable vector text."""
    mpl.rcParams.update(
        {
            "font.family": ["Times New Roman", "Arial", "DejaVu Serif"],
            "font.serif": ["Times New Roman", "Times", "DejaVu Serif"],
            "font.size": 9,
            "axes.labelsize": 10,
            "xtick.labelsize": 8.5,
            "ytick.labelsize": 8.5,
            "legend.fontsize": 8.5,
            "axes.linewidth": 0.7,
            "axes.unicode_minus": False,
            "svg.fonttype": "none",
            "pdf.fonttype": 42,
            "hatch.linewidth": 0.55,
            "savefig.transparent": True,
        }
    )


def read_state_counts() -> dict[str, list[int]]:
    """Read the 24 hourly counts for every state and validate all totals."""
    if not STATE_WORKBOOK.is_file():
        raise FileNotFoundError(f"State workbook not found: {STATE_WORKBOOK.name}")

    workbook = load_workbook(STATE_WORKBOOK, read_only=True, data_only=True)
    try:
        if STATE_SHEET not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet {STATE_SHEET!r} not found; available sheets: "
                f"{', '.join(workbook.sheetnames)}"
            )
        worksheet = workbook[STATE_SHEET]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows)
        expected_header = ("P_ESS_state",) + tuple(
            f"{hour:02d}:00" for hour in range(24)
        )
        if tuple(header[:25]) != expected_header:
            raise ValueError("Unexpected state-workbook header or hour order.")

        counts: dict[str, list[int]] = {}
        for row in rows:
            state = row[0]
            if state not in ALL_STATES:
                continue
            if state in counts:
                raise ValueError(f"Duplicate state row: {state}")
            values = list(row[1:25])
            if len(values) != 24 or any(
                isinstance(value, bool)
                or not isinstance(value, (int, float))
                or not math.isfinite(float(value))
                or float(value) < 0
                or not float(value).is_integer()
                for value in values
            ):
                raise ValueError(f"State {state} must contain 24 non-negative counts.")
            counts[state] = [int(value) for value in values]
    finally:
        workbook.close()

    missing = [state for state in ALL_STATES if state not in counts]
    if missing:
        raise ValueError(f"Missing state rows: {', '.join(missing)}")
    if any(counts["NC"]):
        raise ValueError("NC has non-zero counts and cannot be silently omitted.")

    for hour in range(24):
        total = sum(counts[state][hour] for state in ALL_STATES)
        if total != EXPECTED_DAYS:
            raise ValueError(
                f"Hour {hour:02d}:00 has {total} state observations; "
                f"expected {EXPECTED_DAYS}."
            )
    return counts


def parse_timestamp(value: str) -> datetime:
    """Parse the timestamp formats accepted by the source CSV."""
    for date_format in ("%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %H:%M:%S"):
        try:
            return datetime.strptime(value.strip(), date_format)
        except ValueError:
            pass
    raise ValueError(f"Unsupported timestamp: {value!r}")


def read_daily_hourly_prices() -> tuple[dict[int, list[float]], int, int]:
    """Return 31 daily means per hour after excluding all non-May rows."""
    if not PRICE_CSV.is_file():
        raise FileNotFoundError(f"Price CSV not found: {PRICE_CSV.name}")

    grouped: dict[tuple[int, int], list[float]] = defaultdict(list)
    included_rows = 0
    excluded_rows = 0
    with PRICE_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames != ["date", "price"]:
            raise ValueError("Price CSV must have exactly the columns 'date' and 'price'.")
        for csv_row, row in enumerate(reader, start=2):
            timestamp = parse_timestamp(row["date"])
            try:
                price = float(row["price"])
            except (TypeError, ValueError) as exc:
                raise ValueError(f"CSV row {csv_row}: invalid price.") from exc
            if not math.isfinite(price):
                raise ValueError(f"CSV row {csv_row}: price is not finite.")

            if (timestamp.year, timestamp.month) != (TARGET_YEAR, TARGET_MONTH):
                excluded_rows += 1
                continue
            grouped[(timestamp.day, timestamp.hour)].append(price)
            included_rows += 1

    expected_rows = EXPECTED_DAYS * 24 * EXPECTED_POINTS_PER_DAY_HOUR
    if included_rows != expected_rows:
        raise ValueError(
            f"Found {included_rows} May price rows; expected {expected_rows}."
        )

    daily_hourly: dict[int, list[float]] = {hour: [] for hour in range(24)}
    for day in range(1, EXPECTED_DAYS + 1):
        for hour in range(24):
            values = grouped[(day, hour)]
            if len(values) != EXPECTED_POINTS_PER_DAY_HOUR:
                raise ValueError(
                    f"{TARGET_YEAR}-{TARGET_MONTH:02d}-{day:02d} {hour:02d}:00 "
                    f"has {len(values)} price points; expected "
                    f"{EXPECTED_POINTS_PER_DAY_HOUR}."
                )
            daily_hourly[hour].append(sum(values) / len(values))
    return daily_hourly, included_rows, excluded_rows


def summarize_prices(
    daily_hourly: dict[int, list[float]],
) -> list[HourlyPriceSummary]:
    """Calculate two-sided 90% t confidence intervals for the hourly means."""
    summaries: list[HourlyPriceSummary] = []
    alpha = 1.0 - CONFIDENCE_LEVEL
    for hour in range(24):
        values = daily_hourly[hour]
        n = len(values)
        if n != EXPECTED_DAYS:
            raise ValueError(f"Hour {hour:02d}:00 has {n} daily means.")
        mean = sum(values) / n
        variance = sum((value - mean) ** 2 for value in values) / (n - 1)
        standard_deviation = math.sqrt(variance)
        critical_value = float(t.ppf(1.0 - alpha / 2.0, df=n - 1))
        margin = critical_value * standard_deviation / math.sqrt(n)
        summaries.append(
            HourlyPriceSummary(
                hour=hour,
                mean=mean,
                ci_low=mean - margin,
                ci_high=mean + margin,
                standard_deviation=standard_deviation,
                sample_size=n,
            )
        )
    return summaries


def write_source_data(
    state_counts: dict[str, list[int]], summaries: list[HourlyPriceSummary]
) -> Path:
    """Write the exact values displayed in the figure."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / f"{OUTPUT_BASENAME}_source_data.csv"
    fieldnames = [
        "hour",
        *[f"count_{state}" for state in STATE_ORDER],
        "average_price_usd_per_mwh",
        "ci90_lower_usd_per_mwh",
        "ci90_upper_usd_per_mwh",
        "daily_mean_sd_usd_per_mwh",
        "n_days",
        "five_minute_points_per_day_hour",
    ]
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for summary in summaries:
            row: dict[str, object] = {
                "hour": summary.hour,
                **{
                    f"count_{state}": state_counts[state][summary.hour]
                    for state in STATE_ORDER
                },
                "average_price_usd_per_mwh": f"{summary.mean:.8f}",
                "ci90_lower_usd_per_mwh": f"{summary.ci_low:.8f}",
                "ci90_upper_usd_per_mwh": f"{summary.ci_high:.8f}",
                "daily_mean_sd_usd_per_mwh": (
                    f"{summary.standard_deviation:.8f}"
                ),
                "n_days": summary.sample_size,
                "five_minute_points_per_day_hour": EXPECTED_POINTS_PER_DAY_HOUR,
            }
            writer.writerow(row)
    return output_path


def build_figure(
    state_counts: dict[str, list[int]], summaries: list[HourlyPriceSummary]
) -> plt.Figure:
    """Create the transparent stacked-bar and price confidence-band figure."""
    configure_matplotlib()
    mm_per_inch = 25.4
    figure, count_axis = plt.subplots(
        figsize=(FIGURE_WIDTH_MM / mm_per_inch, FIGURE_HEIGHT_MM / mm_per_inch),
        facecolor="none",
    )
    figure.subplots_adjust(left=0.075, right=0.905, bottom=0.22, top=0.82)
    count_axis.set_facecolor("none")
    price_axis = count_axis.twinx()
    price_axis.set_facecolor("none")

    hours = list(range(24))
    bottoms = [0] * 24
    for state in STATE_ORDER:
        values = state_counts[state]
        count_axis.bar(
            hours,
            values,
            width=BAR_WIDTH,
            bottom=bottoms,
            color=COLORS[state],
            edgecolor="none",
            linewidth=0,
            zorder=2,
        )
        bottoms = [bottom + value for bottom, value in zip(bottoms, values)]

    means = [summary.mean for summary in summaries]
    lower = [summary.ci_low for summary in summaries]
    upper = [summary.ci_high for summary in summaries]
    price_axis.fill_between(
        hours,
        lower,
        upper,
        facecolor=mpl.colors.to_rgba(PRICE_COLOR, CI_FILL_ALPHA),
        edgecolor=PRICE_COLOR,
        hatch=CI_HATCH,
        linewidth=0.0,
        zorder=3,
    )
    price_axis.plot(
        hours,
        means,
        color=PRICE_COLOR,
        linewidth=1.15,
        marker="D",
        markersize=3.0,
        markerfacecolor="white",
        markeredgecolor=PRICE_COLOR,
        markeredgewidth=0.8,
        zorder=4,
    )

    count_axis.set_xlim(-0.7, 23.7)
    count_axis.set_ylim(0, 35)
    count_axis.set_xticks(hours)
    count_axis.set_yticks(range(0, 36, 5))
    count_axis.set_xlabel("Hours")
    count_axis.set_ylabel("Count")
    count_axis.tick_params(axis="both", direction="out", length=2.8, width=0.65)

    price_axis.set_ylim(0, 60)
    price_axis.set_yticks(range(0, 61, 10))
    price_axis.set_ylabel("Price (USD/MWh)", color=PRICE_COLOR)
    price_axis.tick_params(
        axis="y",
        colors=PRICE_COLOR,
        direction="out",
        length=2.8,
        width=0.65,
    )

    frame_color = "#B7B7B7"
    for axis in (count_axis, price_axis):
        for spine in axis.spines.values():
            spine.set_color(frame_color)
            spine.set_linewidth(0.65)
    count_axis.spines["right"].set_visible(False)
    price_axis.spines["left"].set_visible(False)

    legend_handles = [
        Patch(
            facecolor=mpl.colors.to_rgba(PRICE_COLOR, CI_FILL_ALPHA),
            edgecolor=PRICE_COLOR,
            hatch=CI_HATCH,
            linewidth=0,
            label="90% Confidence Interval",
        ),
        Line2D(
            [0],
            [0],
            color=PRICE_COLOR,
            linewidth=1.15,
            marker="D",
            markersize=3.0,
            markerfacecolor="white",
            markeredgecolor=PRICE_COLOR,
            markeredgewidth=0.8,
            label="Average price",
        ),
    ]
    figure.legend(
        handles=legend_handles,
        loc="upper center",
        bbox_to_anchor=(0.5, 0.985),
        ncol=2,
        frameon=False,
        handlelength=1.6,
        handletextpad=0.35,
        columnspacing=5.5,
        borderaxespad=0,
    )
    return figure


def save_figure(figure: plt.Figure) -> list[Path]:
    """Export one transparent figure in PNG, SVG, and PDF formats."""
    FIGURES_DIR.mkdir(parents=True, exist_ok=True)
    base = FIGURES_DIR / OUTPUT_BASENAME
    paths = [
        base.with_suffix(".png"),
        base.with_suffix(".svg"),
        base.with_suffix(".pdf"),
    ]
    figure.savefig(
        paths[0],
        dpi=PNG_DPI,
        transparent=True,
        metadata={"Title": "May 2025 hourly state counts and average prices"},
    )
    figure.savefig(
        paths[1],
        transparent=True,
        metadata={
            "Creator": "Matplotlib",
            "Description": "May 2025 hourly state counts and average prices",
        },
    )
    figure.savefig(
        paths[2],
        transparent=True,
        metadata={
            "Creator": "Matplotlib",
            "Title": "May 2025 hourly state counts and average prices",
        },
    )
    return paths


def main() -> None:
    state_counts = read_state_counts()
    daily_hourly, included_rows, excluded_rows = read_daily_hourly_prices()
    summaries = summarize_prices(daily_hourly)
    source_data_path = write_source_data(state_counts, summaries)
    figure = build_figure(state_counts, summaries)
    figure_paths = save_figure(figure)
    plt.close(figure)

    print(f"May price rows included: {included_rows}")
    print(f"Non-May price rows excluded: {excluded_rows}")
    print(f"Daily hourly means per hour: {EXPECTED_DAYS}")
    print(f"Confidence level: {CONFIDENCE_LEVEL:.0%} (Student t, df=30)")
    print(
        "Price mean range: "
        f"{min(summary.mean for summary in summaries):.4f} to "
        f"{max(summary.mean for summary in summaries):.4f} USD/MWh"
    )
    print(
        "CI range: "
        f"{min(summary.ci_low for summary in summaries):.4f} to "
        f"{max(summary.ci_high for summary in summaries):.4f} USD/MWh"
    )
    print(f"Source data: {source_data_path.name}")
    for path in figure_paths:
        print(f"Figure: {path.name}")


if __name__ == "__main__":
    main()
