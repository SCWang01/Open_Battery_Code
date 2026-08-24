# -*- coding: utf-8 -*-
"""
Created on Apr 22nd, 2025

@author:gcg

Variant V5.  V3 collapses the entire CAISO battery
fleet into a single equivalent battery and assumes 100% of it bids optimally
through one aggregated demand function, which is unrealistic.  V5 splits the
aggregate fleet with a ratio k in (0, 1]:

  * a controllable k-unit (capacity Cap*k, power Pdmax*k / Pcmax*k) is optimised
    by the bidding method (M1), and
  * the remaining (1-k) portion keeps its original actual output,
    (1-k) * hourly_battery.

The combined "method" output is therefore
    P_method = P_optimised(k-unit) + (1-k) * hourly_battery,
compared against the full actual fleet (100% hourly_battery) baseline.  At k=1
this reproduces V3 exactly.  Data source and cost path are identical to V3:
the hourly CAISO natural-gas data in ng_data/gasYYYYMM.xlsx with the 'exact'
piecewise merit-order cost model.

The bidding optimization enforces mutually exclusive charging and discharging
through binary charge/discharge states.  Solver-scale residuals are normalized
during post-solve power classification.

"""

from cost_calculation import gas_cost, gas_marginal_price, ng_carbon_emission
from Random_Generator import (
    END_YEAR_MONTH as RANDOM_END_YEAR_MONTH,
    N_T as RANDOM_N_T,
    START_YEAR_MONTH as RANDOM_START_YEAR_MONTH,
    load_monthly_innovations,
)
import gurobipy as gp
from gurobipy import GRB
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import zipfile
import csv
import json
import calendar
import tempfile
from pathlib import Path
from io import StringIO
from scipy.optimize import minimize
from tqdm import tqdm
from openpyxl import load_workbook

#%% Global settings

N_t = RANDOM_N_T  # The number of time intervals in one day
# Aliases of the Random_Generator constants: change them there, not here, so
# the scenario data and the study period stay consistent.
START_YEAR_MONTH = RANDOM_START_YEAR_MONTH
END_YEAR_MONTH = RANDOM_END_YEAR_MONTH
year_month_list = [
    (year, month)
    for year in range(START_YEAR_MONTH[0], END_YEAR_MONTH[0] + 1)
    for month in range(1, 13)
    if START_YEAR_MONTH <= (year, month) <= END_YEAR_MONTH
]
monthlist = np.array([calendar.month_name[month] for _, month in year_month_list])
monthnumlist = np.array([f'{month:02d}' for _, month in year_month_list])
eta = 0.95 # Set the efficiency of the equivalent battery
Smax = 1 # Set the maximum SOC of the equivalent battery
Smin = 0 # Set the minimum SOC of the equivalent battery
N_price = 300 # the number of traversed prices
meanstd = 2 # prediction error variable
k = 0.2 # fraction of the fleet treated as the controllable (optimised) unit; must be in (0, 1]
PROGRAM_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = PROGRAM_DIR.parent
DATA_DIR = PROJECT_ROOT / 'data'
RESULTS_DIR = PROJECT_ROOT / 'Results'
COST_MODE = 'exact' # gas cost model: 'exact' (piecewise merit order) or 'quadratic' (Fuel_Coe.xlsx fit)


def ensure_results_dir():
    """Create and return the project-level Results directory."""
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    return RESULTS_DIR


def load_monthly_price_error_data(year_value, month_value, n_day):
    """Load the fixed random scenario assigned to one calendar month."""
    return load_monthly_innovations(year_value, month_value, n_day=n_day)


def Marginal_Check(input_path, control_capacity, eta, initial_soc):
    """Append each dsfunction stair's one-hour post-action SOC as column 8.

    The workbook's ``SOC`` column records the Control battery state after the
    cleared ``P_ESS`` for that hour.  Therefore each candidate stair is applied
    to the state before the hour: ``initial_soc`` for the first row and the
    preceding row's recorded SOC thereafter.

    Existing eighth values are replaced so the operation is idempotent.  The
    workbook is validated completely before a temporary file atomically
    replaces the input file.
    """
    input_path = Path(input_path)
    if not input_path.is_file():
        raise FileNotFoundError(f'dsfunction workbook not found: {input_path}')

    control_capacity = float(control_capacity)
    eta = float(eta)
    initial_soc = float(initial_soc)
    if not np.isfinite(control_capacity) or control_capacity <= 0:
        raise ValueError('control_capacity must be a finite positive number')
    if not np.isfinite(eta) or eta <= 0:
        raise ValueError('eta must be a finite positive number')
    if not np.isfinite(initial_soc):
        raise ValueError('initial_soc must be finite')

    workbook = load_workbook(input_path, data_only=False)
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    required_headers = ('time', 'dsfunction', 'P_ESS', 'SOC')
    header_columns = {}
    for header in required_headers:
        matches = [
            index for index, value in enumerate(headers, start=1)
            if value == header
        ]
        if len(matches) != 1:
            raise ValueError(
                f'Header {header!r} appears {len(matches)} times in '
                f'{input_path}; exactly one is required'
            )
        header_columns[header] = matches[0]

    if worksheet.max_row < 2:
        raise ValueError(f'No hourly records found in {input_path}')

    excel_rows = range(2, worksheet.max_row + 1)
    times = pd.to_datetime(
        [worksheet.cell(row, header_columns['time']).value for row in excel_rows],
        errors='coerce',
    )
    if times.isna().any():
        bad_index = int(np.flatnonzero(times.isna())[0])
        raise ValueError(f'Invalid time at Excel row {bad_index + 2}')
    if len(times) > 1:
        hourly_steps = np.diff(times.to_numpy(dtype='datetime64[ns]'))
        bad_steps = np.flatnonzero(hourly_steps != np.timedelta64(1, 'h'))
        if bad_steps.size:
            row = int(bad_steps[0]) + 3
            raise ValueError(f'Time is not hourly and continuous at Excel row {row}')
    if (
        times[0].day != 1
        or times[0].hour != 0
        or times[-1].year != times[0].year
        or times[-1].month != times[0].month
        or times[-1].day != calendar.monthrange(times[0].year, times[0].month)[1]
        or times[-1].hour != 23
    ):
        raise ValueError('Time column must cover one complete calendar month')

    time_step = 24 / N_t
    serialized_dsfunctions = []
    matrix_widths = set()
    previous_actual_soc = initial_soc
    for excel_row in excel_rows:
        p_ess_value = worksheet.cell(excel_row, header_columns['P_ESS']).value
        actual_soc_value = worksheet.cell(excel_row, header_columns['SOC']).value
        try:
            p_ess = float(p_ess_value)
            actual_soc = float(actual_soc_value)
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f'P_ESS and SOC must be numeric at Excel row {excel_row}'
            ) from exc
        if not np.isfinite(p_ess) or not np.isfinite(actual_soc):
            raise ValueError(
                f'P_ESS and SOC must be finite at Excel row {excel_row}'
            )

        if p_ess >= 0:
            expected_actual_soc = (
                previous_actual_soc
                - p_ess / eta / control_capacity * time_step
            )
        else:
            expected_actual_soc = (
                previous_actual_soc
                - p_ess * eta / control_capacity * time_step
            )
        if not np.isclose(actual_soc, expected_actual_soc, rtol=1e-9, atol=1e-9):
            raise ValueError(
                f'Control SOC is inconsistent with P_ESS at Excel row '
                f'{excel_row}: expected {expected_actual_soc}, found {actual_soc}'
            )

        dsfunction_value = worksheet.cell(
            excel_row, header_columns['dsfunction']
        ).value
        if not isinstance(dsfunction_value, str):
            raise ValueError(f'dsfunction must be JSON text at Excel row {excel_row}')
        try:
            dsfunction = json.loads(dsfunction_value)
        except json.JSONDecodeError as exc:
            raise ValueError(
                f'Invalid dsfunction JSON at Excel row {excel_row}'
            ) from exc
        if not isinstance(dsfunction, list) or not dsfunction:
            raise ValueError(
                f'dsfunction must be a non-empty matrix at Excel row {excel_row}'
            )

        enriched_dsfunction = []
        for stair_index, stair in enumerate(dsfunction, start=1):
            if not isinstance(stair, list) or len(stair) not in (7, 8):
                raise ValueError(
                    f'dsfunction stair {stair_index} at Excel row {excel_row} '
                    'must contain 7 or 8 values'
                )
            matrix_widths.add(len(stair))
            try:
                power = float(stair[2])
            except (TypeError, ValueError) as exc:
                raise ValueError(
                    f'dsfunction power at Excel row {excel_row}, stair '
                    f'{stair_index} must be numeric'
                ) from exc
            if not np.isfinite(power):
                raise ValueError(
                    f'dsfunction power at Excel row {excel_row}, stair '
                    f'{stair_index} must be finite'
                )

            if power >= 0:
                candidate_soc = (
                    previous_actual_soc
                    - power / eta / control_capacity * time_step
                )
            else:
                candidate_soc = (
                    previous_actual_soc
                    - power * eta / control_capacity * time_step
                )
            enriched_dsfunction.append([*stair[:7], candidate_soc])

        serialized_dsfunctions.append(
            json.dumps(enriched_dsfunction, separators=(',', ':'))
        )
        previous_actual_soc = actual_soc

    if len(matrix_widths) != 1:
        raise ValueError(
            'All dsfunction stairs must consistently contain either 7 or 8 values'
        )

    for excel_row, serialized in zip(excel_rows, serialized_dsfunctions):
        worksheet.cell(excel_row, header_columns['dsfunction']).value = serialized

    temporary_path = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f'.{input_path.stem}.',
            suffix=input_path.suffix,
            dir=input_path.parent,
            delete=False,
        ) as temporary_file:
            temporary_path = Path(temporary_file.name)
        workbook.save(temporary_path)
        temporary_path.replace(input_path)
    finally:
        if temporary_path is not None and temporary_path.exists():
            temporary_path.unlink()

    return input_path
#%% Read CAISO data

def read_monthly_natural_gas(year, monthnum, Nday):
    """Read the existing hourly monthly natural-gas workbook."""
    input_path = DATA_DIR / 'ng_data' / f'gas{year}{monthnum}.xlsx'
    monthly_gas = pd.read_excel(input_path, usecols=['date', 'hour', 'gas.gas'])
    monthly_gas['date'] = pd.to_numeric(monthly_gas['date'], errors='coerce')
    monthly_gas['hour'] = pd.to_numeric(monthly_gas['hour'], errors='coerce')
    monthly_gas['gas.gas'] = pd.to_numeric(monthly_gas['gas.gas'], errors='coerce')
    monthly_gas = monthly_gas.dropna(subset=['date', 'hour', 'gas.gas'])
    monthly_gas[['date', 'hour']] = monthly_gas[['date', 'hour']].astype(int)

    gas_by_time = monthly_gas.set_index(['date', 'hour'])['gas.gas']
    if gas_by_time.index.has_duplicates:
        raise ValueError(f'Duplicate date/hour rows found in {input_path}')

    expected_dates = [int(f'{year}{monthnum}{day:02d}') for day in range(1, Nday + 1)]
    expected_index = pd.MultiIndex.from_product(
        [expected_dates, range(N_t)], names=['date', 'hour']
    )
    gas_by_time = gas_by_time.reindex(expected_index)
    skipped_dates = [
        str(date) for date in expected_dates
        if gas_by_time.loc[date].isna().any()
    ]
    return gas_by_time.fillna(0).to_numpy(dtype=float), skipped_dates

def readdata(Ntall,Nt,Nday):
    df_price = pd.read_csv(DATA_DIR / 'price' / f'{year}{monthnum} CAISO Average Price.csv')  # real time average LMP
    price = df_price['price'].values
    hourly_gas, skipped_gas_dates = read_monthly_natural_gas(year, monthnum, Nday)
    battery = np.zeros((Ntall*Nday))  # real time battery discharge power (>0 discharge;<0 charge)
    for d in range(Nday): # cut the data by days
        date = monthnum + f'{d + 1:02d}'
        if f'{year}{date}' in skipped_gas_dates:
            continue

        try:
            df_ess = pd.read_csv(DATA_DIR / 'battery_data' / f'CAISO-batteries-{year}{date}.csv')  # read the real time battery power
            df_ess = df_ess.set_index(df_ess.columns[0])
            daily_battery = pd.to_numeric(
                df_ess.loc['Total batteries', :], errors='raise'
            ).to_numpy(dtype=float)
            if len(daily_battery) != Ntall:
                raise ValueError(f'Expected {Ntall} battery values for {year}{date}')
            if not np.isfinite(daily_battery).all():
                raise ValueError(f'Non-finite battery values for {year}{date}')
            battery[Ntall*d:Ntall*(d+1)] = daily_battery
        except (OSError, pd.errors.ParserError, KeyError, ValueError):
            # The original data set contains truncated and incomplete CAISO days.
            pass

    df_cu = pd.read_csv(DATA_DIR / 'curtailment' / f'curtailment_{year}{monthnum}.csv')  # read the curtailment of renewables
    curtailment = df_cu['total_curtailment_mwh'].to_numpy()[:Nt*Nday]
    if len(curtailment) < Nt*Nday:
        # Keep the calendar alignment when a source month is short (March
        # 2025 contains 743 rather than 744 hourly values).
        curtailment = np.pad(curtailment, (0, Nt*Nday-len(curtailment)))

    # transform data sampled per 5 minutes to data sampled per hour
    points_per_hour = 12
    n_hours_price = len(price) // points_per_hour
    n_hours_battery = len(battery) // points_per_hour
    price = price[:n_hours_price * points_per_hour]
    battery = battery[:n_hours_battery * points_per_hour]
    hourly_price = price.reshape(n_hours_price, points_per_hour).mean(axis=1)
    hourly_battery = battery.reshape(n_hours_battery, points_per_hour).mean(axis=1)

    # Use the corresponding monthly charge/discharge peak on every valid day,
    # but set both limits to zero on missing/invalid all-zero days.
    daily_battery = hourly_battery.reshape(Nday, Nt)
    monthly_Pdmax = np.ceil(max(0.0, hourly_battery.max())/100)*100
    monthly_Pcmax = np.ceil(max(0.0, -hourly_battery.min())/100)*100
    zero_output_days = np.all(daily_battery == 0, axis=1)
    Pdmax = np.full(Nday, monthly_Pdmax, dtype=float)
    Pcmax = np.full(Nday, monthly_Pcmax, dtype=float)
    Pdmax[zero_output_days] = 0
    Pcmax[zero_output_days] = 0

    # calculate the capacity and the SOC at 0:00 on the first day of the month
    E = np.zeros((Nt*Nday+1))
    for i in range(1,Nt*Nday+1):
        if  hourly_battery[i-1] < 0:
            E[i] = E[i-1] -  hourly_battery[i-1]*eta
        else:
            E[i] = E[i-1] -  hourly_battery[i-1]/eta
    Emax = max(E)
    Emin = min(E)
    Cap = 1000 *np.ceil((Emax-Emin)/1000)
    SOC = E/Cap
    smin_original = np.min(SOC)
    # Shift the reconstructed trajectory by exactly enough to make its
    # minimum zero.  Rounding this offset to a 5% grid can push the maximum
    # above 100% (the passive and actual-baseline paths are not optimized
    # subject to SOC bounds).  Since Cap already covers the full E range,
    # the exact shift gives 0 <= SINI + E/Cap <= 1 up to floating-point error.
    SINI = max(0.0, float(-smin_original))
    return hourly_price, hourly_gas, hourly_battery, Pdmax, Pcmax, Smax, Smin, Cap, eta, SINI, curtailment, skipped_gas_dates


#%% M1: the demand function bidding -- the proposed method
def build_dsfunction(pri0, Res0, numncd0):
    """Build price stairs enriched with representative NCD and SOC-limit values.

    Consecutive price samples stay in the same stair only while the rounded
    power, the sign of (ncd0 - ncd1), and the SOC-limit state all remain
    unchanged.  Each stair stores the most frequent NCD/SOC vector in its
    interval; ties use the vector that first appears at the lower price.
    """
    Res0_rounded = np.round(Res0, 4)
    sign_state = np.sign(numncd0[0, :] - numncd0[1, :]).astype(int)
    soc_limit_state = numncd0[3, :].astype(int)

    segment_bounds = []
    segment_start = 0
    for K in range(1, len(Res0_rounded) + 1):
        segment_finished = (
            K == len(Res0_rounded)
            or Res0_rounded[K] != Res0_rounded[segment_start]
            or sign_state[K] != sign_state[segment_start]
            or soc_limit_state[K] != soc_limit_state[segment_start]
        )
        if segment_finished:
            segment_bounds.append((segment_start, K))  # K is exclusive
            segment_start = K

    dsfunction = np.zeros((len(segment_bounds), 7))
    previous_price_high = pri0[0]
    for i, (start, end) in enumerate(segment_bounds):
        ncd_soc_vectors = numncd0[:, start:end].T
        unique_vectors, first_indices, counts = np.unique(
            ncd_soc_vectors, axis=0, return_index=True, return_counts=True
        )
        max_count = np.max(counts)
        mode_candidates = np.flatnonzero(counts == max_count)
        mode_index = mode_candidates[np.argmin(first_indices[mode_candidates])]

        dsfunction[i, 0] = previous_price_high
        dsfunction[i, 1] = pri0[end - 1]
        dsfunction[i, 2] = Res0_rounded[start]
        dsfunction[i, 3:7] = unique_vectors[mode_index]
        previous_price_high = pri0[end - 1]

    return dsfunction


def clean_power(value, power_max, abs_tol=1e-3):
    """Remove solver-scale residual power during post-solve processing.

    The optimization model keeps its original solution.  This helper only
    normalizes values used for output classification, using a fixed absolute
    threshold.
    """
    if power_max <= 0 or abs(value) <= abs_tol:
        return 0.0
    return value


def biddingNEW(N_t, N_price, eta,Cap,Smax,Smin,Pdmax,Pcmax,price,SOC_ini):
    # get prices
    pri = price[1:N_t]
    price_min = np.min(price)
    price_max = np.max(price)
    price_span = max(price_max - price_min, 1.0)
    # Expand both sides by one full-price span.  The resulting interval is three
    # times as wide and always contains every input price, regardless of sign.
    pri0 = np.linspace(
        price_min - price_span,
        price_max + price_span,
        N_price,
    )

    # initialize variables for the stair recognization
    iup= np.zeros((N_price)) # the index of interval when the battery first reach the maximum energy limit
    ilow= np.zeros((N_price)) # the index of interval when the battery first reach the minimum energy limit
    numncd0= np.zeros((4,N_price)) # three NCD counts plus the first-reached SOC-limit state (Smax=1, Smin=-1, neither=0)
    Tb= np.zeros((N_price)) # the index of interval when the battery first reach the maximum/minimum energy limit
    ResE= np.zeros((N_t,N_price)) # the energy of battery at each time interval
    Res= np.zeros((N_t,N_price)) # the power of battery at each time interval
    ResPd= np.zeros((N_t,N_price)) # the discharge power of battery at each time interval
    ResPc= np.zeros((N_t,N_price)) # the charge power of battery at each time interval
    Res0 = np.zeros((N_price)) # the power of battery at the current time interval
    # Retained for compatibility with the existing result schema.  Explicit
    # simultaneous charge/discharge detection is no longer performed.
    initial_value = 0
    following_value = 0

    # traverse the price for the current interval from the minimum to the maximum
    for K in range(N_price):
        # build the model
        model = gp.Model()
        # variable definition
        SOC = model.addVars( N_t+1, vtype=GRB.CONTINUOUS, lb=Smin, ub=Smax, name='SOC')
        Pc = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pcmax, name='Pc')
        Pd = model.addVars( N_t, vtype=GRB.CONTINUOUS, lb=0, ub=Pdmax, name='Pd')
        sc = model.addVars(N_t, vtype=GRB.BINARY, name='sc')
        sd = model.addVars(N_t, vtype=GRB.BINARY, name='sd')
        # Constraints
        # SOC initialization
        model.addConstr((SOC[0] == SOC_ini), 'SOC_initial')
        # SOC update
        model.addConstrs((Cap*SOC[t] == Cap*SOC[t-1] - (Pd[t-1]/eta - eta*Pc[t-1])*(24/N_t) for t in range(1, N_t+1)), 'SOC')
        # SOC bounds
        model.addConstrs(((SOC[t] >= Smin) for t in range(1, N_t+1)), 'SOC_lower_bound')
        model.addConstrs(((SOC[t] <= Smax) for t in range(1, N_t+1)), 'SOC_upper_bound')
        # Charging and discharging are mutually exclusive in the optimization.
        model.addConstrs((Pc[t] <= Pcmax * sc[t] for t in range(N_t)), 'charge_state')
        model.addConstrs((Pd[t] <= Pdmax * sd[t] for t in range(N_t)), 'discharge_state')
        model.addConstrs((sc[t] + sd[t] <= 1 for t in range(N_t)), 'mutually_exclusive_state')
        # objective function
        model.setObjective(((pri0[K]*((Pd[0]-Pc[0]))*(24/N_t) +(sum(pri[t-1]*((Pd[t]-Pc[t])*(24/N_t)) for t in range(1,N_t)))) ), GRB.MAXIMIZE)
        # solve the model
        model.setParam('OutputFlag', 0)
        model.optimize()
        # Normalize solver-scale residuals before recording the net power.
        pd0_clean = clean_power(Pd[0].X, Pdmax)
        pc0_clean = clean_power(Pc[0].X, Pcmax)
        Res0[K] = pd0_clean-pc0_clean

        # stair recognization
        for t in range(N_t):
            pd_clean = clean_power(Pd[t].X, Pdmax)
            pc_clean = clean_power(Pc[t].X, Pcmax)
            if pd_clean>pc_clean:
                ResPd[t,K] = pd_clean
                ResPc[t,K] = 0
                Res[t,K] = pd_clean
            else:
                ResPc[t,K] = -pc_clean
                ResPd[t,K] = 0
                Res[t,K] = -pc_clean
        for t in range(N_t):
            ResE[t,K] = SOC[t+1].X
        # the index of interval when the battery first reach the maximum energy limit
        upper_limit_hits = np.flatnonzero(np.isclose(ResE[:, K], Smax, rtol=0, atol=1e-6))
        if len(upper_limit_hits) == 0:
            iup[K] = 24
        else:
            iup[K] = upper_limit_hits[0] + 1
        # the index of interval when the battery first reach the minimum energy limit
        lower_limit_hits = np.flatnonzero(np.isclose(ResE[:, K], Smin, rtol=0, atol=1e-6))
        if len(lower_limit_hits) == 0:
            ilow[K] = 24
        else:
            ilow[K] = lower_limit_hits[0] + 1
        Tb[K]=min(iup[K],ilow[K]) # the index of interval when the battery first reach the maximum/minimum energy limit
        numncd0[0,K] = np.sum(ResPc[:int(Tb[K]),K]==-Pcmax) # the number of min-rate power before the battery first reach the maximum/mimium energy limit
        numncd0[1,K] = np.sum(ResPd[:int(Tb[K]),K]==Pdmax) # the number of max-rate power before the battery first reach the maximum/mimium energy limit
        numncd0[2,K] = np.sum((ResPc[:int(Tb[K]), K] + ResPd[:int(Tb[K]), K] >= -1e-3) &
                       (ResPc[:int(Tb[K]), K] + ResPd[:int(Tb[K]), K] <= 1e-3)) # the number of the null stair before the battery first reach the maximum/mimium energy limit
        if iup[K] < ilow[K]:
            numncd0[3,K] = 1
        elif ilow[K] < iup[K]:
            numncd0[3,K] = -1
        else:
            numncd0[3,K] = 0

    # Columns: price_low, price_high, powerstair, ncd0, ncd1, ncd2,
    # and first-reached SOC-limit state.
    dsfunction = build_dsfunction(pri0, Res0, numncd0)
    return [numncd0, dsfunction, initial_value, following_value]


#%% calculate the profit of the battery that bid by M1
def calculate_profit(
    N_t, Nday, N_price, eta, CAP, Smax, Smin, Pdmax, Pcmax, price,
    Sinitial, meanstd, price_error_z, progress_desc,
):
    expected_noise_shape = (N_t * Nday, N_t)
    if np.shape(price_error_z) != expected_noise_shape:
        raise ValueError(
            f'price_error_z has shape {np.shape(price_error_z)}, '
            f'expected {expected_noise_shape}'
        )
    if not np.isfinite(price_error_z).all():
        raise ValueError('price_error_z contains non-finite values')

    # Initialize the battery
    ESSOC_status = np.zeros((N_t*Nday))
    profit = np.zeros((N_t*Nday))
    P_cleared = np.zeros((N_t*Nday))
    ncd = np.zeros((N_t*Nday,4,N_price))
    dsfunctions = []
    initial_value = 0
    following_value = 0

    # for each time interval bid a demand-supply function
    for t in tqdm(range(N_t*Nday), desc=progress_desc, unit='hour', leave=False):
        # generate prices with prediction error
        priceN_t = price[t:t+N_t] # get price of next N_t intervals
        priceN_t_with_error = np.zeros((N_t)) # initialize the priceN_t_with_error
        for tt in range(N_t):
            if tt == 0:
                # The current price is observed exactly; prediction errors
                # apply only to the future prices used by biddingNEW.
                priceN_t_with_error[tt] = priceN_t[tt]
                continue
            # The fixed-price forecasts used by biddingNEW start at tt=1.
            # The base error is meanstd% (2% by default) at horizon 1, growing
            # by 0.1 percentage point per horizon step.
            rate = meanstd / 100 + (tt - 1) * 0.001
            noise = price_error_z[t, tt] * rate
            priceN_t_with_error[tt] = np.array(priceN_t[tt]*(1 + noise)) # generate the prices with errors
            # adjust the extreme values
            if priceN_t_with_error[tt]<np.min(price):
                priceN_t_with_error[tt] =np.min(price)
            if priceN_t_with_error[tt]>np.max(price):
                priceN_t_with_error[tt]=np.max(price)

        # bid a demand-supply function using the current day's power bounds
        # (Pdmax/Pcmax are length-Nday arrays; fixed within a calendar day)
        d = t // N_t
        numncd0, dsfunction, hourly_initial_value, hourly_following_value = biddingNEW(
            N_t, N_price, eta, CAP, Smax, Smin,
            Pdmax[d], Pcmax[d], priceN_t_with_error, Sinitial
        )
        initial_value += hourly_initial_value
        following_value += hourly_following_value
        dsfunctions.append(dsfunction.copy())

        # clearing and profit calculation
        price_cleared = priceN_t[0]
        # The first and last stair extend beyond the sampled forecast grid so
        # every possible clearing price maps to a bid, including price spikes.
        if price_cleared < dsfunction[0, 0]:
            P_cleared[t] = dsfunction[0, 2]
        elif price_cleared > dsfunction[-1, 1]:
            P_cleared[t] = dsfunction[-1, 2]
        else:
            for i in range(len(dsfunction)):
                price_low = dsfunction[i, 0]
                price_high = dsfunction[i, 1]
                if price_low <= price_cleared <= price_high: # the battery is cleared by regarded as a price-taker
                    P_cleared[t] = dsfunction[i, 2]
        if P_cleared[t] >=0:
            Sinitial = Sinitial - P_cleared[t]/eta/CAP*(24/N_t)
        else:
            Sinitial = Sinitial - P_cleared[t]*eta/CAP*(24/N_t)
        ESSOC_status[t] = Sinitial
        profit[t] = price_cleared*P_cleared[t]*(24/N_t) # calculate the profit for each time interval
        ncd[t,:,:] = numncd0
    aveprice = np.mean(price[0:N_t*Nday])
    ESSend = ESSOC_status[-1]*CAP # the finally remained energy
    end_value = ESSend*aveprice # the profit of the finally remained energy
    return (
        profit, ESSOC_status, P_cleared, end_value, ncd, dsfunctions,
        initial_value, following_value,
    )

#%% calculate the profit of the battery that bid by M2 (the actual data)
def calculate_profit_actual(eta,CAP,price, N_t,Nday, netoutput,Sinitial):
    # Initialize the battery
    ESSOC_status = np.zeros((N_t*Nday))
    profit = np.zeros((N_t*Nday))
    # profit calculation
    for t in range(N_t*Nday):
        x = price[t]
        profit[t] = x*netoutput[t]*(24/N_t) # calculate the profit for each time interval
        if netoutput[t] >=0:
            Sinitial = Sinitial - netoutput[t]/eta/CAP*(24/N_t)
        else:
            Sinitial = Sinitial - netoutput[t]*eta/CAP*(24/N_t)
        ESSOC_status[t] = Sinitial
    aveprice = np.mean(price[0:N_t*Nday])
    ESSend = ESSOC_status[-1]*CAP # the finally remained energy
    end_value = ESSend*aveprice # the profit of the finally remained energy
    return profit, ESSOC_status,netoutput,end_value
#%% Generation cost and carbon emission

def carbon_emission_calculation(gas_generation, year_value, month_value):
    """Calculate natural-gas CO2 using the cost-ranked dispatch order.

    Each dispatched plant contributes ``mmbtu_per_mwh * MWh * 0.05306``
    metric tonnes of CO2.  The underlying merit-order stack is the same stable
    ``$_per_mwh`` ordering used by the exact gas-cost calculation.
    """
    return ng_carbon_emission(gas_generation, year_value, month_value)


# The 'quadratic' cost mode is implemented in cost_calculation.py
# (quad_cost / quad_marginal_price), which reads Program/Fuel_Coe.xlsx.

# calculate the generation cost and the carbon emission
def calculate_cost_and_carbon(gas_actual,curtailment, PESS_method,PESS_actual,N_t,Nday):

    gas_old = np.asarray(gas_actual, dtype=float)
    curtailment_old = np.asarray(curtailment, dtype=float)

    Pgas = np.zeros((N_t*Nday,3))
    Pgas[:,1] = gas_old
    Pgas[:,2] = np.maximum(gas_old+PESS_actual, 0.0) # without battery

    diff = PESS_method-PESS_actual # power discharged more than the actual case

    gas_new = gas_old.copy()
    curtailment_new = curtailment_old.copy()

    supply = diff > 0   # extra ESS supply reduces gas, then adds curtailment
    demand = diff < 0   # extra ESS demand absorbs curtailment, then adds gas

    gas_reduction = np.minimum(diff[supply], np.maximum(gas_old[supply], 0.0))
    gas_new[supply] = gas_old[supply] - gas_reduction
    curtailment_new[supply] = curtailment_old[supply] + (diff[supply] - gas_reduction)

    extra_demand = -diff[demand]
    curtailment_reduction = np.minimum(extra_demand, curtailment_old[demand])
    curtailment_new[demand] = curtailment_old[demand] - curtailment_reduction
    gas_new[demand] = gas_old[demand] + (extra_demand - curtailment_reduction)

    #gas_new = np.maximum(gas_new, 0.0)
    #curtailment_new = np.maximum(curtailment_new, 0.0)

    # renewable energy newly absorbed = curtailment reduction (negative = more curtailment)
    absorbed = curtailment_old - curtailment_new

    Pgas[:,0] = gas_new

    month_int = int(monthnum)
    Cost = np.column_stack([
        gas_cost(Pgas[:, 0], year, month_int, COST_MODE),
        gas_cost(Pgas[:, 1], year, month_int, COST_MODE),
        gas_cost(Pgas[:, 2], year, month_int, COST_MODE),
    ])
    # marginal ($/MWh) gas price at each gas output, consistent with COST_MODE
    MargPrice = np.column_stack([
        gas_marginal_price(Pgas[:, 0], year, month_int, COST_MODE),
        gas_marginal_price(Pgas[:, 1], year, month_int, COST_MODE),
        gas_marginal_price(Pgas[:, 2], year, month_int, COST_MODE),
    ])
    Carbon = np.column_stack([
        carbon_emission_calculation(Pgas[:, 0], year, month_int),
        carbon_emission_calculation(Pgas[:, 1], year, month_int),
        carbon_emission_calculation(Pgas[:, 2], year, month_int),
    ])
    return Pgas, Cost, absorbed, diff, MargPrice, Carbon

def calculate_main(
    meanstd, price, gas, battery, curtailment, Pdmax, Pcmax, Smax, Smin,
    Cap, eta, SINI, price_error_z,
):
    # V6: split the aggregate fleet into a controllable k-unit (optimised by the
    # bidding method) and a passive (1-k) remainder that keeps its actual output.
    # The controllable unit is a proportionally shrunk copy of the fleet, so its
    # capacity and power scale by k while its SOC trajectory (hence SINI) is
    # unchanged.  P_method = P_cleared_controlled + (1-k)*battery.

    # controllable k-unit: optimised bidding on the k-scaled battery
    (
        profit_ctrl, ESSOC_status_ctrl, P_cleared_ctrl, end_value_ctrl, ncd,
        dsfunctions, initial_value, following_value,
    ) = calculate_profit(
        N_t, N_day, N_price, eta, Cap*k, Smax, Smin, Pdmax*k, Pcmax*k,
        price, SINI, meanstd, price_error_z, f'{year}-{monthnum}'
    )
    # passive (1-k) remainder: keeps its original actual output, (1-k)*battery.
    # At k=1 the remainder and its capacity are both zero; skip it to avoid a
    # 0/0 in the SOC update and let k=1 reproduce V3 exactly.
    battery_passive = (1 - k) * battery
    if (1 - k) > 0:
        profit_passive, ESSOC_status_passive, _, end_value_passive = calculate_profit_actual(
            eta, Cap*(1 - k), price, N_t, N_day, battery_passive, SINI
        )
    else:
        profit_passive = np.zeros((N_t*N_day))
        end_value_passive = 0.0
    # full actual baseline (100% of the fleet), unchanged from V3
    profit_actual, ESSOC_status_actual, P_cleared_actual, end_value_actual = calculate_profit_actual(
        eta, Cap, price, N_t, N_day, battery, SINI
    )

    # combined method output and profit
    P_cleared = P_cleared_ctrl + battery_passive
    profit = profit_ctrl + profit_passive
    total_profit = sum(profit_ctrl) + end_value_ctrl + sum(profit_passive) + end_value_passive
    total_profit_actual = sum(profit_actual) + end_value_actual

    # cost/carbon: full method output vs. full actual baseline
    Pgas, Cost, absorbed, diff, MargPrice, Carbon = calculate_cost_and_carbon(
        gas, curtailment, P_cleared, P_cleared_actual, N_t, N_day
    )

    # summarize the results
    total_cost = sum(Cost[:,0])
    total_cost_actual = sum(Cost[:,1])
    total_cost_withoutESS = sum(Cost[:,2])
    total_carbon = sum(Carbon[:,0])
    total_carbon_actual = sum(Carbon[:,1])
    total_carbon_withoutESS = sum(Carbon[:,2])
    res_date = {'price': price[:N_t*N_day], 'profit': profit, 'profit_actual': profit_actual,
    'P_ESS': P_cleared, 'P_ESS_controlled': P_cleared_ctrl, 'P_ESS_passive': battery_passive,
    'P_ESS_actual': P_cleared_actual,
    'SOC_controlled': ESSOC_status_ctrl,
    'P_natural_gas':Pgas[:,0],'P_natural_gas_actual':Pgas[:,1], 'P_renewable_absorbed':absorbed,
    'marginal_price_gas': MargPrice[:,0], 'marginal_price_gas_actual': MargPrice[:,1], 'marginal_price_gas_withoutESS': MargPrice[:,2],
    'cost': Cost[:,0], 'cost_actual': Cost[:,1], 'cost_withoutESS': Cost[:,2],
    'carbon': Carbon[:,0], 'carbon_actual': Carbon[:,1], 'carbon_withoutESS': Carbon[:,2],
    'total_profit': total_profit, 'total_profit_actual': total_profit_actual,
    'total_cost': total_cost, 'total_cost_actual': total_cost_actual, 'total_cost_withoutESS': total_cost_withoutESS,
    'total_carbon': total_carbon, 'total_carbon_actual': total_carbon_actual,
    'total_carbon_withoutESS': total_carbon_withoutESS,
    'total_absorb': sum(absorbed)}

    output_dir = ensure_results_dir()
    df_results = pd.DataFrame(res_date)
    df_results.to_csv(
        output_dir / (
            f'{month}{year}_eta95%_std{int(meanstd)}_'
            f'{COST_MODE}_V5_k{int(k*100)}.csv'
        ),
        index=False,
    )
    return [
        total_profit, total_profit_actual, total_cost, total_cost_actual,
        total_cost_withoutESS, Cost, absorbed, P_cleared, P_cleared_ctrl,
        ESSOC_status_ctrl, Pgas, ncd, dsfunctions, initial_value, following_value,
        total_carbon, total_carbon_actual, total_carbon_withoutESS, Carbon,
    ]


#%% main

def run_one_month(num, export_dsfunctions=False, return_detection_counts=False):
    """Compute a single month (0-based index into monthlist) and return its
    summary dict.  When ``return_detection_counts`` is true, also return the
    month's legacy detection placeholders as a second dict.

    Sets the module-level globals the other functions read, writes the per-month
    CSV / .npy files, and returns the row that goes into the summary table.  This
    is the single source of truth for the per-month pipeline; run_all_months
    calls it, so edits here apply to every month.
    """
    global year, month, monthnum, N_day

    if not isinstance(num, (int, np.integer)):
        raise TypeError(f'month index must be an integer, got {type(num).__name__}')
    num = int(num)
    if num < 0 or num >= len(year_month_list):
        raise IndexError(
            f'month index {num} is outside 0..{len(year_month_list) - 1}'
        )

    year_num, month_num = year_month_list[num]
    year = str(year_num)
    month = monthlist[num]
    monthnum = monthnumlist[num]
    N_day = calendar.monthrange(int(year), int(monthnum))[1]
    price_error_z = load_monthly_price_error_data(year_num, month_num, N_day)

    output_dir = ensure_results_dir()

    data = readdata(N_t * 12, N_t, N_day)
    price_glo, gas_glo, battery_glo, Pdmax_glo, Pcmax_glo, Smax_glo, Smin_glo, Cap_glo, eta_glo, SINI_glo, curtailment_glo, skipped_gas_dates = data
    result = calculate_main(
        meanstd, price_glo, gas_glo, battery_glo, curtailment_glo,
        Pdmax_glo, Pcmax_glo, Smax_glo, Smin_glo, Cap_glo, eta_glo, SINI_glo,
        price_error_z,
    )
    (
        total_profit, total_profit_actual, total_cost, total_cost_actual,
        total_cost_withoutESS, costdetails, absorbed, P_cleared,
        P_cleared_ctrl, ESSOC_status_ctrl, Pgas, ncd, dsfunctions, initial_value,
        following_value, total_carbon, total_carbon_actual,
        total_carbon_withoutESS, carbon_details,
    ) = result

    np.save(output_dir / f'ncd_{month}{year}_{COST_MODE}_V5_k{int(k*100)}.npy', ncd)
    np.save(output_dir / f'Pcleared_{month}{year}_{COST_MODE}_V5_k{int(k*100)}.npy', P_cleared)
    if export_dsfunctions:
        times = pd.date_range(
            f'{year}-{monthnum}-01 00:00:00', periods=len(dsfunctions), freq='h'
        )
        export_data = pd.DataFrame({
            'time': times,
            'dsfunction': [
                json.dumps(dsfunction.tolist(), separators=(',', ':'))
                for dsfunction in dsfunctions
            ],
            'P_ESS': P_cleared_ctrl,
            'SOC': ESSOC_status_ctrl,
        })
        dsfunction_path = (
            output_dir
            / f'dsfunction_{month}{year}_{COST_MODE}_V5_k{int(k*100)}.xlsx'
        )
        export_data.to_excel(
            dsfunction_path,
            sheet_name=f'{month}{year}',
            index=False,
        )
        Marginal_Check(dsfunction_path, Cap_glo*k, eta_glo, SINI_glo)

    totalabsorbed = float(np.sum(absorbed))
    totalgas = float(np.sum(gas_glo))
    rate_gas = totalabsorbed / totalgas
    carbon_reduce = total_carbon_actual - total_carbon
    year_month = f'{year}{monthnum}'
    rate_carbon = (
        carbon_reduce / total_carbon_actual
        if total_carbon_actual != 0
        else 0.0
    )
    summary = {
        'year_month': year_month,
        'k': k,
        'total_profit': total_profit,
        'total_profit_actual': total_profit_actual,
        'total_cost': total_cost,
        'total_cost_actual': total_cost_actual,
        'total_cost_withoutESS': total_cost_withoutESS,
        'total_carbon': total_carbon,
        'total_carbon_actual': total_carbon_actual,
        'total_carbon_withoutESS': total_carbon_withoutESS,
        'total_absorbed': totalabsorbed,
        'total_natural_gas': totalgas,
        'rate_gas': rate_gas,
        'carbon_reduce': carbon_reduce,
        'rate_carbon': rate_carbon,
        'skipped_gas_dates': ','.join(skipped_gas_dates),
    }
    detection_counts = {
        'Month': year_month,
        'Initial_Value': initial_value,
        'following_value': following_value,
    }
    if return_detection_counts:
        return summary, detection_counts
    return summary


def write_detection_summary(detection_rows, summary_path):
    """Write monthly legacy detection placeholders beside a summary."""
    summary_name = summary_path.name
    if summary_name.startswith('summary_'):
        summary_name = summary_name[len('summary_'):]
    detection_path = summary_path.with_name(
        f'simultaneous_charge_discharge_counts_{summary_name}'
    )
    columns = ['Month', 'Initial_Value', 'following_value']
    pd.DataFrame(detection_rows, columns=columns).to_csv(detection_path, index=False)
    print(f'Wrote simultaneous charge/discharge counts: {detection_path}')
    return detection_path


def run_all_months():
    output_dir = ensure_results_dir()

    month_results = [
        run_one_month(num, return_detection_counts=True)
        for num in tqdm(range(len(monthlist)), desc='All months', unit='month')
    ]
    summaries = [summary for summary, _ in month_results]
    detection_rows = [counts for _, counts in month_results]

    start_period = f'{START_YEAR_MONTH[0]}{START_YEAR_MONTH[1]:02d}'
    end_period = f'{END_YEAR_MONTH[0]}{END_YEAR_MONTH[1]:02d}'
    summary_path = output_dir / f'summary_{start_period}_{end_period}_{COST_MODE}_V5_k{int(k*100)}.csv'
    pd.DataFrame(summaries).to_csv(summary_path, index=False)
    write_detection_summary(detection_rows, summary_path)
    return summary_path


def run_may_2025():
    """Run only May 2025 and export its NCD and hourly dsfunctions."""
    may_2025_index = year_month_list.index((2025, 5))
    summary, detection_counts = run_one_month(
        may_2025_index,
        export_dsfunctions=True,
        return_detection_counts=True,
    )
    summary_path = (
        ensure_results_dir()
        / f'summary_202505_{COST_MODE}_V5_k{int(k*100)}.csv'
    )
    pd.DataFrame([summary]).to_csv(summary_path, index=False)
    write_detection_summary([detection_counts], summary_path)
    return summary_path

def run_april_2025():
    """Run only April 2025 and export its NCD and hourly dsfunctions."""
    april_2025_index = year_month_list.index((2025, 4))
    summary, detection_counts = run_one_month(
        april_2025_index,
        export_dsfunctions=True,
        return_detection_counts=True,
    )
    summary_path = (
        ensure_results_dir()
        / f'summary_202504_{COST_MODE}_V5_k{int(k*100)}.csv'
    )
    pd.DataFrame([summary]).to_csv(summary_path, index=False)
    write_detection_summary([detection_counts], summary_path)
    return summary_path



def run_Jan_2025():
    """Run only January 2025 and export its NCD and hourly dsfunctions."""
    jan_2025_index = year_month_list.index((2025, 1))
    summary, detection_counts = run_one_month(
        jan_2025_index,
        export_dsfunctions=True,
        return_detection_counts=True,
    )
    summary_path = (
        ensure_results_dir()
        / f'summary_202501_{COST_MODE}_V5_k{int(k*100)}.csv'
    )
    pd.DataFrame([summary]).to_csv(summary_path, index=False)
    write_detection_summary([detection_counts], summary_path)
    return summary_path

if __name__ == '__main__':
    run_may_2025()
