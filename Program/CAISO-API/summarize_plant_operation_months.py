"""Summarize monthly appearances by Plant ID and reported prime mover.

Input files:
    Month_Agg_Clear/CAISO_NG_YYYY_MM.xlsx

Period:
    2023-01 through 2025-12 (inclusive)

Output columns:
    Plant ID | Prime Mover | Operation Time | Total Operation Time

``Operation Time`` is a text field containing comma-separated YYYY_MM values.
``Total Operation Time`` is the number of distinct months in which the Plant ID
and Prime Mover combination appears.  Exact duplicates of that combination in
one monthly file count once.
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_FOLDER = BASE_DIR / "Month_Agg_Clear"
DEFAULT_OUTPUT_NAME = "CAISO_NG_Plant_Operation_2023_01_to_2025_12.xlsx"
START_YEAR_MONTH = (2023, 1)
END_YEAR_MONTH = (2025, 12)


def iter_year_months(
    start: tuple[int, int], end: tuple[int, int]
) -> Iterable[tuple[int, int]]:
    """Yield (year, month) pairs from start through end, inclusive."""
    year, month = start
    while (year, month) <= end:
        yield year, month
        if month == 12:
            year, month = year + 1, 1
        else:
            month += 1


def normalize_header(value: object) -> str:
    """Normalize a column name for case- and spacing-insensitive matching."""
    return "".join(str(value).split()).casefold() if value is not None else ""


def normalize_plant_id(value: object) -> int | float | str | None:
    """Keep Plant IDs stable and convert integral floats to integers."""
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        stripped = value.strip()
        try:
            number = float(stripped)
        except ValueError:
            return stripped
        return int(number) if number.is_integer() else number
    return value


def plant_sort_key(plant_id: int | float | str) -> tuple[int, object]:
    """Sort numeric Plant IDs numerically, followed by any text IDs."""
    if isinstance(plant_id, (int, float)) and not isinstance(plant_id, bool):
        return 0, plant_id
    return 1, str(plant_id)


PlantId = int | float | str
PlantKey = tuple[PlantId, str]


def collect_operation_months(folder: Path) -> dict[PlantKey, list[str]]:
    """Collect months for each (Plant ID, Prime Mover) combination."""
    operation_months: defaultdict[PlantKey, set[str]] = defaultdict(set)
    missing_files: list[Path] = []

    for year, month in iter_year_months(START_YEAR_MONTH, END_YEAR_MONTH):
        month_label = f"{year}_{month:02d}"
        input_path = folder / f"CAISO_NG_{month_label}.xlsx"
        if not input_path.is_file():
            missing_files.append(input_path)
            continue

        workbook = load_workbook(input_path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rows = worksheet.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                raise ValueError(f"Workbook is empty: {input_path}")

            normalized_header = [normalize_header(value) for value in header]
            try:
                plant_id_column = normalized_header.index("plantid")
            except ValueError as error:
                raise ValueError(
                    f"Column 'Plant Id' was not found in {input_path}"
                ) from error
            try:
                prime_mover_column = normalized_header.index(
                    "reportedprimemover"
                )
            except ValueError as error:
                raise ValueError(
                    "Column 'Reported Prime Mover' was not found in "
                    f"{input_path}"
                ) from error

            # A set prevents an identical Plant ID + Prime Mover combination
            # from being counted more than once in the same month.
            plants_in_this_month: set[PlantKey] = set()
            for row in rows:
                if max(plant_id_column, prime_mover_column) >= len(row):
                    continue
                plant_id = normalize_plant_id(row[plant_id_column])
                prime_mover_value = row[prime_mover_column]
                prime_mover = (
                    str(prime_mover_value).strip()
                    if prime_mover_value is not None
                    else ""
                )
                if plant_id is not None and prime_mover:
                    plants_in_this_month.add((plant_id, prime_mover))

            for plant_key in plants_in_this_month:
                operation_months[plant_key].add(month_label)
        finally:
            workbook.close()

    if missing_files:
        missing_text = "\n".join(f"  - {path}" for path in missing_files)
        raise FileNotFoundError(f"Required monthly files are missing:\n{missing_text}")

    return {
        plant_key: sorted(months)
        for plant_key, months in operation_months.items()
    }


def write_summary(
    operation_months: dict[PlantKey, list[str]], output_path: Path
) -> None:
    """Write the four-column plant operation summary workbook."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Plant Operation Summary"

    headers = (
        "Plant ID",
        "Prime Mover",
        "Operation Time",
        "Total Operation Time",
    )
    worksheet.append(headers)

    for plant_id, prime_mover in sorted(
        operation_months,
        key=lambda key: (plant_sort_key(key[0]), key[1]),
    ):
        months = operation_months[(plant_id, prime_mover)]
        worksheet.append(
            (plant_id, prime_mover, ",".join(months), len(months))
        )
        # Explicit Excel text format as requested.
        worksheet.cell(worksheet.max_row, 3).number_format = "@"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:D{worksheet.max_row}"
    worksheet.column_dimensions["A"].width = 14
    worksheet.column_dimensions["B"].width = 14
    worksheet.column_dimensions["C"].width = 105
    worksheet.column_dimensions["D"].width = 22

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Group CAISO natural-gas records by Plant ID and Reported Prime "
            "Mover, then summarize the months present from 2023-01 through "
            "2025-12."
        )
    )
    parser.add_argument(
        "--folder",
        type=Path,
        default=DEFAULT_FOLDER,
        help=f"folder containing monthly workbooks (default: {DEFAULT_FOLDER})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help=(
            "output workbook path (default: "
            f"<folder>/{DEFAULT_OUTPUT_NAME})"
        ),
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_path = args.output or args.folder / DEFAULT_OUTPUT_NAME
    operation_months = collect_operation_months(args.folder)
    write_summary(operation_months, output_path)
    print(
        f"Completed: {len(operation_months)} Plant ID/Prime Mover "
        f"combinations written to {output_path}"
    )


if __name__ == "__main__":
    main()
