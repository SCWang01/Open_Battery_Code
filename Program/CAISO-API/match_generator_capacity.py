"""Match CAISO plant/prime-mover pairs to EIA generator attributes.

By default, the CAISO plant/prime-mover pairs are collected from the 40 monthly
workbooks from January 2023 through April 2026.  A plant-operation workbook can
still be supplied explicitly with ``--operation-file``.  The program first
searches the small California Operable workbook.  It then searches the
Operable and Retired and Canceled sheets of the 2023, 2024, and 2025 annual
EIA generator workbooks, in that order, only for keys not found earlier.
Remaining misses are searched in the Operating and Retired sheets of the May
2026 EIA-860M workbook, restricted to natural-gas generator rows.  Every
lookup uses Plant ID and Prime Mover as its composite key.  Workbooks are read
in streaming read-only mode to keep memory use low.

All EIA rows with the same Plant Code and Prime Mover are retained, and their
capacity values are serialized as JSON-style arrays in Excel cells.  Thus one
match is written as ``[a]``, repeated matches as ``[a, b, c]``, and no match
as ``[]``.  A missing value inside a matched EIA row is written as ``null`` so
the two arrays remain aligned by EIA row.
"""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path
from typing import Iterable, Iterator, Sequence

from openpyxl import Workbook, load_workbook


DEFAULT_MONTHLY_DIR = Path("Month_Agg_Clear")
DEFAULT_START_YEAR_MONTH = (2023, 1)
DEFAULT_END_YEAR_MONTH = (2026, 4)
DEFAULT_GENERATOR_FILE = Path(
    "Generator_Info/3_1_Generator_Y2025_Early_Release_CA_Operable.xlsx"
)
DEFAULT_ANNUAL_GENERATOR_FILES = (
    Path("Data/3_1_Generator_Y2023.xlsx"),
    Path("Data/3_1_Generator_Y2024.xlsx"),
    Path("Data/3_1_Generator_Y2025.xlsx"),
)
DEFAULT_SUPPLEMENTAL_GENERATOR_FILE = Path("Data/may_generator2026.xlsx")
DEFAULT_OUTPUT_FILE = Path(
    "Generator_Info/CAISO_NG_Plant_Capacity_Minimum_Load_2023_01_to_2026_04.xlsx"
)
ANNUAL_SHEET_NAMES = ("Operable", "Retired and Canceled")
SUPPLEMENTAL_SHEET_NAMES = ("Operating", "Retired")

OUTPUT_HEADERS = (
    "Plant ID",
    "Prime Mover",
    "Nameplate Capacity (MW)",
    "Minimum Load (MW)",
)

PlantId = int | float | str
PlantKey = tuple[PlantId, str]
GeneratorValues = tuple[list[object | None], list[object | None]]


def parse_year_month(value: str) -> tuple[int, int]:
    """Parse an inclusive command-line month in YYYY-MM form."""
    try:
        year_text, month_text = value.split("-", maxsplit=1)
        year, month = int(year_text), int(month_text)
    except (TypeError, ValueError) as error:
        raise argparse.ArgumentTypeError(
            f"invalid month {value!r}; expected YYYY-MM"
        ) from error
    if len(year_text) != 4 or len(month_text) != 2 or not 1 <= month <= 12:
        raise argparse.ArgumentTypeError(
            f"invalid month {value!r}; expected YYYY-MM"
        )
    return year, month


def format_year_month(year_month: tuple[int, int]) -> str:
    """Format a (year, month) tuple for file names and messages."""
    return f"{year_month[0]:04d}-{year_month[1]:02d}"


def iter_year_months(
    start: tuple[int, int], end: tuple[int, int]
) -> Iterable[tuple[int, int]]:
    """Yield all (year, month) values in an inclusive range."""
    year, month = start
    while (year, month) <= end:
        yield year, month
        year, month = (year + 1, 1) if month == 12 else (year, month + 1)


def normalize_header(value: object) -> str:
    """Normalize a header for case- and whitespace-insensitive matching."""
    if value is None:
        return ""
    return "".join(str(value).split()).casefold()


def normalize_plant_id(value: object) -> PlantId | None:
    """Normalize numeric-looking IDs so, for example, 246 and 246.0 match."""
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else value

    stripped = str(value).strip()
    try:
        number = float(stripped)
    except ValueError:
        return stripped
    return int(number) if number.is_integer() else number


def normalize_prime_mover(value: object) -> str:
    """Normalize prime-mover codes for matching."""
    return str(value).strip().casefold() if value is not None else ""


def clean_array_value(value: object) -> object | None:
    """Convert blank worksheet cells to None, which JSON renders as null."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    return value.strip() if isinstance(value, str) else value


def find_header(
    rows: Iterator[tuple[object, ...]],
    required_headers: Sequence[str],
    workbook_path: Path,
) -> tuple[tuple[object, ...], dict[str, int]]:
    """Find a table header row and return it plus normalized column indexes."""
    required = {normalize_header(header) for header in required_headers}
    for row in rows:
        indexes = {
            normalize_header(value): index
            for index, value in enumerate(row)
            if value is not None
        }
        if required.issubset(indexes):
            return row, indexes

    missing = ", ".join(required_headers)
    raise ValueError(
        f"Could not find a header row containing {missing} in {workbook_path}"
    )


def read_operation_pairs(
    operation_path: Path,
) -> list[tuple[PlantId, str]]:
    """Read Plant ID and Prime Mover pairs in source order."""
    if not operation_path.is_file():
        raise FileNotFoundError(f"Operation workbook not found: {operation_path}")

    workbook = load_workbook(operation_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        _, indexes = find_header(
            rows, ("Plant ID", "Prime Mover"), operation_path
        )
        plant_column = indexes[normalize_header("Plant ID")]
        mover_column = indexes[normalize_header("Prime Mover")]

        pairs: list[tuple[PlantId, str]] = []
        for row in rows:
            if max(plant_column, mover_column) >= len(row):
                continue
            plant_id = normalize_plant_id(row[plant_column])
            prime_mover = (
                str(row[mover_column]).strip()
                if row[mover_column] is not None
                else ""
            )
            if plant_id is not None and prime_mover:
                pairs.append((plant_id, prime_mover))
        return pairs
    finally:
        workbook.close()


def read_monthly_operation_pairs(
    monthly_dir: Path,
    start: tuple[int, int],
    end: tuple[int, int],
) -> list[tuple[PlantId, str]]:
    """Collect unique Plant ID/Prime Mover pairs from monthly workbooks."""
    if start > end:
        raise ValueError("start month must not be later than end month")
    if not monthly_dir.is_dir():
        raise FileNotFoundError(
            f"Monthly workbook folder not found: {monthly_dir}"
        )

    pairs: list[tuple[PlantId, str]] = []
    seen_keys: set[PlantKey] = set()
    missing_files: list[Path] = []

    for year, month in iter_year_months(start, end):
        monthly_path = monthly_dir / f"CAISO_NG_{year:04d}_{month:02d}.xlsx"
        if not monthly_path.is_file():
            missing_files.append(monthly_path)
            continue

        workbook = load_workbook(monthly_path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rows = worksheet.iter_rows(values_only=True)
            _, indexes = find_header(
                rows,
                ("Plant Id", "Reported Prime Mover"),
                monthly_path,
            )
            plant_column = indexes[normalize_header("Plant Id")]
            mover_column = indexes[normalize_header("Reported Prime Mover")]

            for row in rows:
                if max(plant_column, mover_column) >= len(row):
                    continue
                plant_id = normalize_plant_id(row[plant_column])
                prime_mover = (
                    str(row[mover_column]).strip()
                    if row[mover_column] is not None
                    else ""
                )
                key = (plant_id, normalize_prime_mover(prime_mover))
                if plant_id is None or not key[1] or key in seen_keys:
                    continue
                seen_keys.add(key)
                pairs.append((plant_id, prime_mover))
        finally:
            workbook.close()

    if missing_files:
        missing_text = "\n".join(f"  - {path}" for path in missing_files)
        raise FileNotFoundError(
            f"Required monthly workbooks are missing:\n{missing_text}"
        )
    return pairs


def collect_generator_values(
    generator_path: Path,
    sheet_numbers: Sequence[int] | None = None,
    requested_keys: set[PlantKey] | None = None,
    sheet_names: Sequence[str] | None = None,
) -> dict[PlantKey, GeneratorValues]:
    """Collect capacity/minimum-load values from selected worksheets.

    ``requested_keys`` limits which rows are retained.  Selected worksheets
    are streamed, and unrelated rows never accumulate in memory.
    """
    if not generator_path.is_file():
        raise FileNotFoundError(f"Generator workbook not found: {generator_path}")

    required_headers = (
        "Plant Code",
        "Prime Mover",
        "Nameplate Capacity (MW)",
        "Minimum Load (MW)",
    )
    values: defaultdict[PlantKey, GeneratorValues] = defaultdict(
        lambda: ([], [])
    )

    workbook = load_workbook(generator_path, read_only=True, data_only=True)
    try:
        if sheet_numbers is not None and sheet_names is not None:
            raise ValueError("Specify worksheet numbers or names, not both")

        if sheet_names is not None:
            worksheets_by_name = {
                worksheet.title.casefold(): worksheet
                for worksheet in workbook.worksheets
            }
            missing_sheets = [
                name
                for name in sheet_names
                if name.casefold() not in worksheets_by_name
            ]
            if missing_sheets:
                raise KeyError(
                    f"Worksheet(s) {missing_sheets} not found in "
                    f"{generator_path}"
                )
            selected_worksheets = [
                worksheets_by_name[name.casefold()] for name in sheet_names
            ]
        else:
            selected_numbers = sheet_numbers or (1,)
            invalid_numbers = [
                number
                for number in selected_numbers
                if number < 1 or number > len(workbook.worksheets)
            ]
            if invalid_numbers:
                raise IndexError(
                    f"Worksheet number(s) {invalid_numbers} do not exist in "
                    f"{generator_path}; workbook has "
                    f"{len(workbook.worksheets)} worksheets"
                )
            selected_worksheets = [
                workbook.worksheets[number - 1]
                for number in selected_numbers
            ]

        for worksheet in selected_worksheets:
            rows = worksheet.iter_rows(values_only=True)
            _, indexes = find_header(
                rows,
                required_headers,
                Path(f"{generator_path} [{worksheet.title}]"),
            )
            plant_column = indexes[normalize_header("Plant Code")]
            mover_column = indexes[normalize_header("Prime Mover")]
            capacity_column = indexes[
                normalize_header("Nameplate Capacity (MW)")
            ]
            minimum_column = indexes[normalize_header("Minimum Load (MW)")]
            last_required_column = max(
                plant_column, mover_column, capacity_column, minimum_column
            )

            for row in rows:
                if last_required_column >= len(row):
                    continue
                plant_id = normalize_plant_id(row[plant_column])
                prime_mover = normalize_prime_mover(row[mover_column])
                if plant_id is None or not prime_mover:
                    continue

                key = (plant_id, prime_mover)
                if requested_keys is not None and key not in requested_keys:
                    continue
                capacities, minimum_loads = values[key]
                capacities.append(clean_array_value(row[capacity_column]))
                minimum_loads.append(clean_array_value(row[minimum_column]))
    finally:
        workbook.close()

    return dict(values)


def collect_annual_generator_values(
    generator_paths: Sequence[Path],
    requested_keys: set[PlantKey],
) -> tuple[dict[PlantKey, GeneratorValues], int]:
    """Search annual files in order, stopping after each composite-key match."""
    collected: dict[PlantKey, GeneratorValues] = {}
    remaining_keys = set(requested_keys)
    matched_key_count = 0

    for generator_path in generator_paths:
        if not remaining_keys:
            break
        values = collect_generator_values(
            generator_path,
            requested_keys=remaining_keys,
            sheet_names=ANNUAL_SHEET_NAMES,
        )
        collected.update(values)
        matched_keys = remaining_keys & values.keys()
        matched_key_count += len(matched_keys)
        remaining_keys -= matched_keys

    return collected, matched_key_count


def collect_supplemental_generator_values(
    generator_path: Path,
    requested_keys: set[PlantKey],
    sheet_names: Sequence[str] = SUPPLEMENTAL_SHEET_NAMES,
) -> dict[PlantKey, GeneratorValues]:
    """Collect NG generator capacities from EIA-860M operating/retired rows.

    The CAISO identifier is matched to EIA ``Plant ID``.  ``Generator ID``
    distinguishes individual generators retained in each output array.  The
    EIA-860M workbook has no minimum-load field, so a ``None`` placeholder is
    appended for every matched capacity to keep both arrays aligned.
    """
    if not requested_keys:
        return {}
    if not generator_path.is_file():
        raise FileNotFoundError(
            f"Supplemental generator workbook not found: {generator_path}"
        )

    required_headers = (
        "Plant ID",
        "Generator ID",
        "Nameplate Capacity (MW)",
        "Energy Source Code",
        "Prime Mover Code",
    )
    values: defaultdict[PlantKey, GeneratorValues] = defaultdict(
        lambda: ([], [])
    )
    seen_generators: set[tuple[PlantId, str, str, str]] = set()

    workbook = load_workbook(generator_path, read_only=True, data_only=True)
    try:
        worksheets_by_name = {
            worksheet.title.casefold(): worksheet
            for worksheet in workbook.worksheets
        }
        missing_sheets = [
            name for name in sheet_names if name.casefold() not in worksheets_by_name
        ]
        if missing_sheets:
            raise KeyError(
                f"Worksheet(s) {missing_sheets} not found in {generator_path}"
            )

        for sheet_name in sheet_names:
            worksheet = worksheets_by_name[sheet_name.casefold()]
            rows = worksheet.iter_rows(values_only=True)
            _, indexes = find_header(
                rows,
                required_headers,
                Path(f"{generator_path} [{worksheet.title}]"),
            )
            plant_column = indexes[normalize_header("Plant ID")]
            generator_column = indexes[normalize_header("Generator ID")]
            capacity_column = indexes[
                normalize_header("Nameplate Capacity (MW)")
            ]
            energy_column = indexes[normalize_header("Energy Source Code")]
            mover_column = indexes[normalize_header("Prime Mover Code")]
            last_required_column = max(
                plant_column,
                generator_column,
                capacity_column,
                energy_column,
                mover_column,
            )

            for row in rows:
                if last_required_column >= len(row):
                    continue
                if normalize_prime_mover(row[energy_column]) != "ng":
                    continue

                plant_id = normalize_plant_id(row[plant_column])
                prime_mover = normalize_prime_mover(row[mover_column])
                if plant_id is None or not prime_mover:
                    continue

                key = (plant_id, prime_mover)
                if key not in requested_keys:
                    continue

                generator_id = normalize_prime_mover(row[generator_column])
                generator_key = (plant_id, generator_id, "ng", prime_mover)
                if not generator_id or generator_key in seen_generators:
                    continue
                seen_generators.add(generator_key)

                capacities, minimum_loads = values[key]
                capacities.append(clean_array_value(row[capacity_column]))
                minimum_loads.append(None)
    finally:
        workbook.close()

    return dict(values)


def serialize_array(values: list[object | None]) -> str:
    """Serialize values as a compact, human-readable JSON array."""
    return json.dumps(values, ensure_ascii=False, separators=(", ", ": "))


def write_output(
    operation_pairs: list[tuple[PlantId, str]],
    generator_values: dict[PlantKey, GeneratorValues],
    output_path: Path,
) -> int:
    """Write the requested four-column workbook and return unmatched count."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Generator Capacity"
    worksheet.append(OUTPUT_HEADERS)

    unmatched_count = 0
    for plant_id, prime_mover in operation_pairs:
        key = (plant_id, normalize_prime_mover(prime_mover))
        capacities, minimum_loads = generator_values.get(key, ([], []))
        if not capacities:
            unmatched_count += 1
        worksheet.append(
            (
                plant_id,
                prime_mover,
                serialize_array(capacities),
                serialize_array(minimum_loads),
            )
        )
        worksheet.cell(worksheet.max_row, 3).number_format = "@"
        worksheet.cell(worksheet.max_row, 4).number_format = "@"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:D{worksheet.max_row}"
    worksheet.column_dimensions["A"].width = 14
    worksheet.column_dimensions["B"].width = 14
    worksheet.column_dimensions["C"].width = 42
    worksheet.column_dimensions["D"].width = 42

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return unmatched_count


def build_capacity_table(
    operation_path: Path,
    generator_path: Path,
    output_path: Path,
    annual_generator_paths: Sequence[Path] = DEFAULT_ANNUAL_GENERATOR_FILES,
    supplemental_generator_path: Path = DEFAULT_SUPPLEMENTAL_GENERATOR_FILE,
) -> tuple[int, int, int, int]:
    """Build the table and return row, annual, supplemental, unmatched counts."""
    operation_pairs = read_operation_pairs(operation_path)
    return build_capacity_table_from_pairs(
        operation_pairs,
        generator_path,
        output_path,
        annual_generator_paths,
        supplemental_generator_path,
    )


def build_capacity_table_from_pairs(
    operation_pairs: list[tuple[PlantId, str]],
    generator_path: Path,
    output_path: Path,
    annual_generator_paths: Sequence[Path] = DEFAULT_ANNUAL_GENERATOR_FILES,
    supplemental_generator_path: Path = DEFAULT_SUPPLEMENTAL_GENERATOR_FILE,
) -> tuple[int, int, int, int]:
    """Build the table from prepared pairs and return matching statistics."""
    generator_values = collect_generator_values(generator_path)

    operation_keys = {
        (plant_id, normalize_prime_mover(prime_mover))
        for plant_id, prime_mover in operation_pairs
    }
    missing_keys = operation_keys - generator_values.keys()
    annual_values, annual_match_count = (
        collect_annual_generator_values(
            annual_generator_paths,
            requested_keys=missing_keys,
        )
        if missing_keys
        else ({}, 0)
    )
    # Annual values are collected only for misses, and each successive year
    # receives only keys not found in an earlier source.
    generator_values.update(annual_values)

    still_missing_keys = missing_keys - annual_values.keys()
    supplemental_values = (
        collect_supplemental_generator_values(
            supplemental_generator_path,
            requested_keys=still_missing_keys,
        )
        if still_missing_keys
        else {}
    )
    generator_values.update(supplemental_values)

    unmatched_count = write_output(
        operation_pairs, generator_values, output_path
    )
    supplemental_match_count = len(
        still_missing_keys & supplemental_values.keys()
    )
    return (
        len(operation_pairs),
        annual_match_count,
        supplemental_match_count,
        unmatched_count,
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Match CAISO Plant ID/Prime Mover pairs to all corresponding EIA "
            "generator capacity and minimum-load rows."
        )
    )
    parser.add_argument(
        "--operation-file",
        type=Path,
        default=None,
        help=(
            "optional CAISO operation workbook; when omitted, pairs are read "
            "from the monthly workbooks"
        ),
    )
    parser.add_argument(
        "--monthly-dir",
        type=Path,
        default=DEFAULT_MONTHLY_DIR,
        help=(
            "folder containing monthly CAISO workbooks "
            f"(default: {DEFAULT_MONTHLY_DIR})"
        ),
    )
    parser.add_argument(
        "--start",
        type=parse_year_month,
        default=DEFAULT_START_YEAR_MONTH,
        metavar="YYYY-MM",
        help="first monthly workbook to include (default: 2023-01)",
    )
    parser.add_argument(
        "--end",
        type=parse_year_month,
        default=DEFAULT_END_YEAR_MONTH,
        metavar="YYYY-MM",
        help="last monthly workbook to include (default: 2026-04)",
    )
    parser.add_argument(
        "--generator-file",
        type=Path,
        default=DEFAULT_GENERATOR_FILE,
        help=f"EIA generator workbook (default: {DEFAULT_GENERATOR_FILE})",
    )
    parser.add_argument(
        "--annual-generator-files",
        type=Path,
        nargs="+",
        default=DEFAULT_ANNUAL_GENERATOR_FILES,
        metavar="FILE",
        help=(
            "annual EIA generator workbooks searched in the supplied order; "
            "both Operable and Retired and Canceled are checked "
            "(default: Data/3_1_Generator_Y2023.xlsx, Y2024.xlsx, Y2025.xlsx)"
        ),
    )
    parser.add_argument(
        "--supplemental-generator-file",
        type=Path,
        default=DEFAULT_SUPPLEMENTAL_GENERATOR_FILE,
        help=(
            "EIA-860M workbook whose Operating and Retired sheets are used "
            "for remaining misses "
            f"(default: {DEFAULT_SUPPLEMENTAL_GENERATOR_FILE})"
        ),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_FILE,
        help=f"output workbook (default: {DEFAULT_OUTPUT_FILE})",
    )
    args = parser.parse_args()
    if args.start > args.end:
        parser.error("--start must not be later than --end")
    return args


def main() -> None:
    args = parse_args()
    if args.operation_file is not None:
        operation_pairs = read_operation_pairs(args.operation_file)
        source_description = str(args.operation_file)
    else:
        operation_pairs = read_monthly_operation_pairs(
            args.monthly_dir, args.start, args.end
        )
        source_description = (
            f"{args.monthly_dir} "
            f"({format_year_month(args.start)} to {format_year_month(args.end)})"
        )

    output_rows, annual_matches, supplemental_matches, unmatched_rows = (
        build_capacity_table_from_pairs(
            operation_pairs,
            args.generator_file,
            args.output,
            args.annual_generator_files,
            args.supplemental_generator_file,
        )
    )
    print(f"Operation pairs collected from: {source_description}")
    print(f"Completed: {output_rows} rows written to {args.output}")
    print(f"Matched rows: {output_rows - unmatched_rows}")
    print(f"Matches recovered from annual workbooks: {annual_matches}")
    print(f"Matches recovered from EIA-860M workbook: {supplemental_matches}")
    print(f"Unmatched rows (written with []): {unmatched_rows}")


if __name__ == "__main__":
    main()
