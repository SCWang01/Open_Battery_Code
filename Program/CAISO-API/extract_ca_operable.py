"""Extract California rows from the Operable sheet of the EIA generator file.

The output workbook contains one worksheet named ``Operable``.  It retains
the introductory rows, the complete header, and every column from source rows
whose State value is CA.  The source workbook is never modified.
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = BASE_DIR / "Data" / "3_1_Generator_Y2025.xlsx"
DEFAULT_OUTPUT = (
    BASE_DIR / "Generator_Info" / "3_1_Generator_Y2025_Early_Release_CA_Operable.xlsx"
)
SHEET_NAME = "Operable"


def extract_ca_operable(input_path: Path, output_path: Path) -> int:
    """Write the Operable rows with State=CA and return the data-row count."""
    if not input_path.is_file():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")

    # Read-only mode avoids loading the large source workbook fully into memory.
    source_workbook = load_workbook(input_path, read_only=True, data_only=False)
    try:
        if SHEET_NAME not in source_workbook.sheetnames:
            raise KeyError(f"Worksheet {SHEET_NAME!r} was not found")

        source_sheet = source_workbook[SHEET_NAME]
        output_workbook = Workbook(write_only=True)
        output_sheet = output_workbook.create_sheet(SHEET_NAME)

        state_column: int | None = None
        california_rows = 0

        for row_number, row in enumerate(
            source_sheet.iter_rows(values_only=True), start=1
        ):
            if state_column is None:
                normalized = [
                    str(value).strip().casefold() if value is not None else ""
                    for value in row
                ]

                # Retain title/description rows before the table header.
                output_sheet.append(row)
                if "state" in normalized:
                    state_column = normalized.index("state")
                continue

            state = row[state_column] if state_column < len(row) else None
            if str(state).strip().upper() == "CA":
                # Append the whole row so every non-State field is retained.
                output_sheet.append(row)
                california_rows += 1

        if state_column is None:
            raise ValueError(
                f"No 'State' column was found in worksheet {SHEET_NAME!r}"
            )

        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_workbook.save(output_path)
        return california_rows
    finally:
        source_workbook.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract State=CA rows from the Operable worksheet."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"source Excel file (default: {DEFAULT_INPUT})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"output Excel file (default: {DEFAULT_OUTPUT})",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    count = extract_ca_operable(args.input, args.output)
    print(f"Completed: {count} California rows written to {args.output}")


if __name__ == "__main__":
    main()
