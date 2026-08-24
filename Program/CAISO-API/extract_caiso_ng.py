"""Extract CAISO (CISO) natural-gas records from EIA-923 workbooks.

The source workbook stores its table header on different rows for different
pages.  Page 4 has no fuel-code column, so its generator rows are identified by
matching (Plant Id, Reported Prime Mover) pairs found in the filtered Page 1
natural-gas data.

With no command-line arguments, the script scans ``Data`` for EIA-923 files
from 2023 through 2026.  The already processed years 2024 and 2025 are skipped,
so the default batch run creates ``Data/CAISO_NG_2023.xlsx`` and
``Data/CAISO_NG_2026.xlsx``.  Supplying ``--source`` retains the original
single-workbook mode.
"""

import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_DATA_DIR = BASE_DIR / "Data"
DEFAULT_START_YEAR = 2023
DEFAULT_END_YEAR = 2026
DEFAULT_SKIP_YEARS = (2024, 2025)
SOURCE_FILE_PREFIX = "EIA923_Schedules_2_3_4_5_"
SOURCE_YEAR_PATTERN = re.compile(r"_M_\d{2}_(\d{4})(?:_|\.xlsx$)", re.IGNORECASE)

PAGE1 = "Page 1 Generation and Fuel Data"
PAGE4 = "Page 4 Generator Data"
PAGE5 = "Page 5 Fuel Receipts and Costs"

BA_COLUMN = "Balancing\nAuthority Code"
PAGE1_FUEL_COLUMN = "Reported\nFuel Type Code"
PAGE4_PRIME_MOVER_COLUMN = "Reported\nPrime Mover"
PAGE5_FUEL_COLUMN = "ENERGY_SOURCE"
PLANT_ID_COLUMN = "Plant Id"

COLUMN_ALIASES = {
    BA_COLUMN: (BA_COLUMN, "BA_CODE"),
}


def normalized(series: pd.Series) -> pd.Series:
    """Return trimmed, upper-case strings without changing the source data."""
    return series.astype("string").str.strip().str.upper()


def field_names(field: str) -> tuple[str, ...]:
    """Return accepted source names for one canonical field."""
    return COLUMN_ALIASES.get(field, (field,))


def find_header_row(source_file: Path, sheet_name: str, required_fields: list[str]) -> int:
    """Find the zero-based table header row, which differs by EIA release type."""
    preview = pd.read_excel(source_file, sheet_name=sheet_name, header=None, nrows=15)
    for row_index, row in preview.iterrows():
        values = {str(value).strip() for value in row.dropna()}
        if all(any(name in values for name in field_names(field)) for field in required_fields):
            return int(row_index)
    raise ValueError(
        f"Could not find a header containing {required_fields} within the "
        f"first 15 rows of {sheet_name}"
    )


def read_eia_page(source_file: Path, sheet_name: str, required_fields: list[str]) -> pd.DataFrame:
    """Read one EIA page, remove the repeated early-release warning column, and unify column names."""
    header_row = find_header_row(source_file, sheet_name, required_fields)
    frame = pd.read_excel(source_file, sheet_name=sheet_name, header=header_row)

    # Column A is a repeated warning sentence rather than a data field.
    first_column = str(frame.columns[0])
    if first_column.startswith("Early release data"):
        frame = frame.iloc[:, 1:].copy()

    rename_map: dict[str, str] = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        if canonical in frame.columns:
            continue
        matches = [alias for alias in aliases if alias in frame.columns]
        if len(matches) == 1:
            rename_map[matches[0]] = canonical
        elif len(matches) > 1:
            raise KeyError(
                f"Multiple candidate columns for {canonical!r} exist in "
                f"{sheet_name}: {matches}"
            )
    if rename_map:
        frame = frame.rename(columns=rename_map)

    return frame.dropna(how="all")


def require_columns(frame: pd.DataFrame, sheet_name: str, columns: list[str]) -> None:
    missing = [column for column in columns if column not in frame.columns]
    if missing:
        raise KeyError(f"{sheet_name} is missing required columns: {missing}")


def format_output(writer: pd.ExcelWriter) -> None:
    """Apply readable formatting to every output worksheet."""
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        worksheet.row_dimensions[1].height = 32
        for column_index, cells in enumerate(worksheet.iter_cols(), start=1):
            values = ["" if cell.value is None else str(cell.value) for cell in cells[:300]]
            width = min(max((len(value) for value in values), default=0) + 2, 45)
            worksheet.column_dimensions[get_column_letter(column_index)].width = max(width, 10)


def extract_workbook(
    source_file: Path,
    output_file: Path,
    data_year: int | None = None,
) -> None:
    """Extract one EIA-923 workbook into one CAISO natural-gas workbook."""
    if not source_file.exists():
        raise FileNotFoundError(f"Source file not found: {source_file}")

    page1 = read_eia_page(source_file, PAGE1, [PLANT_ID_COLUMN, BA_COLUMN, PAGE1_FUEL_COLUMN])
    page4 = read_eia_page(source_file, PAGE4, [PLANT_ID_COLUMN, BA_COLUMN, "Generator Id"])
    page5 = read_eia_page(source_file, PAGE5, ["YEAR", PLANT_ID_COLUMN, BA_COLUMN, PAGE5_FUEL_COLUMN])

    require_columns(page1, PAGE1, [BA_COLUMN, PAGE1_FUEL_COLUMN, PLANT_ID_COLUMN, PAGE4_PRIME_MOVER_COLUMN])
    require_columns(page4, PAGE4, [BA_COLUMN, PLANT_ID_COLUMN, PAGE4_PRIME_MOVER_COLUMN])
    require_columns(page5, PAGE5, [BA_COLUMN, PAGE5_FUEL_COLUMN, PLANT_ID_COLUMN])

    page1_ng = page1.loc[
        normalized(page1[BA_COLUMN]).eq("CISO")
        & normalized(page1[PAGE1_FUEL_COLUMN]).eq("NG")
    ].copy()

    # Page 4 does not contain ENERGY_SOURCE.  Match both plant and prime mover
    # against Page 1 so that only generator types reported with NG are included.
    ng_plant_prime_movers = set(
        zip(
            pd.to_numeric(page1_ng[PLANT_ID_COLUMN], errors="coerce"),
            normalized(page1_ng[PAGE4_PRIME_MOVER_COLUMN]),
        )
    )
    page4_keys = zip(
        pd.to_numeric(page4[PLANT_ID_COLUMN], errors="coerce"),
        normalized(page4[PAGE4_PRIME_MOVER_COLUMN]),
    )
    page4_ng = page4.loc[
        normalized(page4[BA_COLUMN]).eq("CISO")
        & pd.Series(
            [key in ng_plant_prime_movers for key in page4_keys],
            index=page4.index,
        )
    ].copy()

    page5_ng = page5.loc[
        normalized(page5[BA_COLUMN]).eq("CISO")
        & normalized(page5[PAGE5_FUEL_COLUMN]).eq("NG")
    ].copy()

    detected_years = sorted(
        {
            int(year)
            for frame in (page1_ng, page4_ng, page5_ng)
            if "YEAR" in frame.columns
            for year in pd.to_numeric(frame["YEAR"], errors="coerce").dropna().unique()
        }
    )
    if data_year is None:
        if len(detected_years) != 1:
            raise ValueError(
                f"Could not determine a unique data year; detected: {detected_years}"
            )
        data_year = detected_years[0]
    elif detected_years != [data_year]:
        raise ValueError(
            f"The requested year is {data_year}, but the workbook contains: "
            f"{detected_years}"
        )

    summary = pd.DataFrame(
        {
            "Data Year": [data_year, data_year, data_year],
            "Output Sheet": ["Page1_GenFuel", "Page4_Generator", "Page5_FuelReceipts"],
            "Source Sheet": [PAGE1, PAGE4, PAGE5],
            "Filter Rule": [
                "Balancing Authority Code=CISO and Reported Fuel Type Code=NG",
                "Balancing Authority Code=CISO and (Plant Id, Prime Mover) "
                "appears in the Page 1 CISO+NG data",
                "Balancing Authority Code=CISO and ENERGY_SOURCE=NG",
            ],
            "Record Count": [len(page1_ng), len(page4_ng), len(page5_ng)],
            "Plant Count": [
                page1_ng[PLANT_ID_COLUMN].nunique(),
                page4_ng[PLANT_ID_COLUMN].nunique(),
                page5_ng[PLANT_ID_COLUMN].nunique(),
            ],
        }
    )

    output_file.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Filter_Summary", index=False)
        page1_ng.to_excel(writer, sheet_name="Page1_GenFuel", index=False)
        page4_ng.to_excel(writer, sheet_name="Page4_Generator", index=False)
        page5_ng.to_excel(writer, sheet_name="Page5_FuelReceipts", index=False)
        format_output(writer)

    print(f"Data year: {data_year}")
    print(f"Created: {output_file}")
    print(
        f"Page 1: {len(page1_ng)} rows; Page 4: {len(page4_ng)} rows; "
        f"Page 5: {len(page5_ng)} rows"
    )


def main(source_file: Path, output_file: Path, data_year: int | None = None) -> None:
    """Keep the original single-workbook Python API for compatibility."""
    extract_workbook(source_file, output_file, data_year)


def find_source_file(data_dir: Path, year: int) -> Path:
    """Find the unique EIA-923 Schedules 2/3/4/5 workbook for one year."""
    candidates = sorted(
        path
        for path in data_dir.glob(f"{SOURCE_FILE_PREFIX}*.xlsx")
        if not path.name.startswith("~$")
        and (match := SOURCE_YEAR_PATTERN.search(path.name)) is not None
        and int(match.group(1)) == year
    )
    if not candidates:
        raise FileNotFoundError(
            f"No {year} EIA-923 Schedules 2/3/4/5 file was found in the "
            f"Data directory: {data_dir}"
        )
    if len(candidates) > 1:
        names = ", ".join(path.name for path in candidates)
        raise ValueError(
            f"Multiple candidate source files were found for {year} in the "
            f"Data directory. Keep only one or select one with --source: {names}"
        )
    return candidates[0]


def process_year_range(
    data_dir: Path,
    start_year: int,
    end_year: int,
    skip_years: set[int],
) -> list[tuple[int, Path, Path]]:
    """Process a year range and return successfully generated workbooks."""
    if start_year > end_year:
        raise ValueError(
            f"The start year cannot be later than the end year: "
            f"{start_year} > {end_year}"
        )
    if not data_dir.is_dir():
        raise FileNotFoundError(f"Data directory not found: {data_dir}")

    generated: list[tuple[int, Path, Path]] = []
    for year in range(start_year, end_year + 1):
        if year in skip_years:
            print(f"Skipping {year}: data has already been processed")
            continue

        source_file = find_source_file(data_dir, year)
        output_file = data_dir / f"CAISO_NG_{year}.xlsx"
        print(f"\nProcessing {year}: {source_file.name}")
        extract_workbook(source_file, output_file, year)
        generated.append((year, source_file, output_file))

    return generated


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Extract CISO natural-gas data from EIA-923. By default, scan "
            "the 2023-2026 files in Data and skip 2024 and 2025."
        )
    )
    parser.add_argument(
        "--source",
        type=Path,
        default=None,
        help="Single-file mode: explicitly select one EIA-923 source workbook.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Single-file output; defaults to Data/CAISO_NG_YYYY.xlsx.",
    )
    parser.add_argument(
        "--year",
        type=int,
        default=None,
        help="Data year for single-file mode; inferred from YEAR when omitted.",
    )
    parser.add_argument(
        "--data-dir",
        type=Path,
        default=DEFAULT_DATA_DIR,
        help="Source and output directory for batch mode (default: Data).",
    )
    parser.add_argument(
        "--start-year",
        type=int,
        default=DEFAULT_START_YEAR,
        help=f"First year for batch mode (default: {DEFAULT_START_YEAR}).",
    )
    parser.add_argument(
        "--end-year",
        type=int,
        default=DEFAULT_END_YEAR,
        help=f"Last year for batch mode (default: {DEFAULT_END_YEAR}).",
    )
    parser.add_argument(
        "--skip-years",
        type=int,
        nargs="*",
        default=list(DEFAULT_SKIP_YEARS),
        help="Years to skip in batch mode (default: 2024 2025).",
    )
    return parser.parse_args()


if __name__ == "__main__":
    arguments = parse_args()
    if arguments.output is not None and arguments.source is None:
        raise SystemExit("--output can only be used with --source")

    if arguments.source is not None:
        source = arguments.source.resolve()
        output = (
            arguments.output.resolve()
            if arguments.output is not None
            else DEFAULT_DATA_DIR
            / f"CAISO_NG_{arguments.year}.xlsx"
            if arguments.year is not None
            else DEFAULT_DATA_DIR / "CAISO_NG.xlsx"
        )
        extract_workbook(source, output, arguments.year)
    else:
        results = process_year_range(
            data_dir=arguments.data_dir.resolve(),
            start_year=arguments.start_year,
            end_year=arguments.end_year,
            skip_years=set(arguments.skip_years),
        )
        print(f"\nBatch processing complete: created {len(results)} files")
        for year, _, output_file in results:
            print(f"{year}: {output_file}")
