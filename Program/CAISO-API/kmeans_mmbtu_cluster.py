"""Cluster CAISO monthly records by MMBtuPer_Unit using one-dimensional K-means."""

import argparse
from pathlib import Path
import warnings

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_score


warnings.filterwarnings(
    "ignore",
    message="Could not find the number of physical cores.*",
    category=UserWarning,
)


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = BASE_DIR / "Month_Agg" / "CAISO_NG_2024_06.xlsx"
DEFAULT_OUTPUT = BASE_DIR / "Cluster_Result" / "CAISO_NG_2024_06_kmeans_5.xlsx"
FEATURE = "MMBtuPer_Unit"


def format_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.row_dimensions[1].height = 32
        for index, cells in enumerate(worksheet.iter_cols(), start=1):
            values = ["" if cell.value is None else str(cell.value) for cell in cells[:300]]
            width = min(max((len(value) for value in values), default=0) + 2, 42)
            worksheet.column_dimensions[get_column_letter(index)].width = max(width, 10)


def cluster_file(
    input_file: Path,
    output_file: Path,
    clusters: int,
    random_state: int,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if not input_file.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")
    if clusters < 2:
        raise ValueError("The number of clusters must be at least 2")

    frame = pd.read_excel(input_file, dtype={FEATURE: object})
    if FEATURE not in frame.columns:
        raise KeyError(f"Input file is missing column: {FEATURE}")

    numeric_feature = pd.to_numeric(frame[FEATURE], errors="coerce")
    valid = numeric_feature.gt(0)
    if valid.sum() < clusters:
        raise ValueError(
            f"Only {int(valid.sum())} valid records are available; cannot form "
            f"{clusters} clusters"
        )
    if numeric_feature.loc[valid].nunique() < clusters:
        raise ValueError(
            f"Fewer than {clusters} distinct valid feature values are available; "
            f"cannot form {clusters} clusters"
        )

    features = numeric_feature.loc[valid].to_numpy().reshape(-1, 1)
    model = KMeans(
        n_clusters=clusters,
        random_state=random_state,
        n_init=50,
        algorithm="lloyd",
    )
    raw_labels = model.fit_predict(features)
    raw_centers = model.cluster_centers_.ravel()

    # sklearn labels are arbitrary. Re-label from the lowest center to highest.
    ordered_raw_labels = raw_centers.argsort()
    label_map = {int(raw_label): rank + 1 for rank, raw_label in enumerate(ordered_raw_labels)}
    ordered_labels = pd.Series(
        [label_map[int(raw_label)] for raw_label in raw_labels],
        index=frame.index[valid],
        dtype="Int64",
    )
    ordered_centers = {rank + 1: float(raw_centers[raw_label]) for rank, raw_label in enumerate(ordered_raw_labels)}

    result = frame.copy()
    result[f"{FEATURE}_Numeric"] = numeric_feature
    result["KMeans_Cluster"] = pd.Series(pd.NA, index=result.index, dtype="Int64")
    result.loc[valid, "KMeans_Cluster"] = ordered_labels
    result["Cluster_Center"] = result["KMeans_Cluster"].map(ordered_centers)
    result["Cluster_Status"] = "Clustered"
    result.loc[numeric_feature.eq(0), "Cluster_Status"] = "Excluded: zero"
    result.loc[numeric_feature.lt(0), "Cluster_Status"] = "Excluded: negative"
    result.loc[numeric_feature.isna(), "Cluster_Status"] = "Excluded: non-numeric"

    clustered = result.loc[valid].copy()
    summary = (
        clustered.groupby("KMeans_Cluster", observed=True)
        .agg(
            Record_Count=(FEATURE, "size"),
            Cluster_Center=("Cluster_Center", "first"),
            Minimum=(f"{FEATURE}_Numeric", "min"),
            Maximum=(f"{FEATURE}_Numeric", "max"),
            Mean=(f"{FEATURE}_Numeric", "mean"),
            Standard_Deviation=(f"{FEATURE}_Numeric", "std"),
        )
        .reset_index()
    )

    score = silhouette_score(features, raw_labels) if len(set(raw_labels)) > 1 else float("nan")
    model_info = pd.DataFrame(
        {
            "Item": [
                "Input File",
                "Feature",
                "Algorithm",
                "Number of Clusters",
                "Random State",
                "n_init",
                "Total Records",
                "Clustered Records",
                "Excluded Records",
                "Exclusion Rule",
                "Inertia",
                "Silhouette Score",
            ],
            "Value": [
                str(input_file),
                FEATURE,
                "K-means (Lloyd)",
                clusters,
                random_state,
                50,
                len(frame),
                int(valid.sum()),
                int((~valid).sum()),
                "MMBtuPer_Unit <= 0 or non-numeric",
                float(model.inertia_),
                float(score),
            ],
        }
    )

    output_file.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        result.to_excel(writer, sheet_name="Clustered_Data", index=False)
        summary.to_excel(writer, sheet_name="Cluster_Summary", index=False)
        model_info.to_excel(writer, sheet_name="Model_Info", index=False)
        format_workbook(writer)

    return result, summary, model_info


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Cluster monthly data by MMBtuPer_Unit using K-means."
    )
    parser.add_argument(
        "input",
        type=Path,
        nargs="?",
        default=DEFAULT_INPUT,
        help="Input monthly XLSX file.",
    )
    parser.add_argument(
        "--output", type=Path, default=DEFAULT_OUTPUT, help="Cluster-result XLSX."
    )
    parser.add_argument(
        "--clusters", type=int, default=5, help="Number of clusters."
    )
    parser.add_argument(
        "--random-state", type=int, default=42, help="Random seed."
    )
    return parser.parse_args()


if __name__ == "__main__":
    arguments = parse_args()
    data, cluster_summary, info = cluster_file(
        arguments.input.resolve(),
        arguments.output.resolve(),
        arguments.clusters,
        arguments.random_state,
    )
    print(f"Created: {arguments.output.resolve()}")
    print(cluster_summary.to_string(index=False))
    print(
        f"Clustered records: {data['KMeans_Cluster'].notna().sum()}; "
        f"excluded: {data['KMeans_Cluster'].isna().sum()}"
    )
