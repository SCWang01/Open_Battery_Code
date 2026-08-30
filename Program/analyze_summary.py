r"""Create monthly and annual battery-study analysis in an Excel workbook.

Run rule (PowerShell)
---------------------
Run the following command from the ``Program`` directory to analyze the
``Results\Self-scheduling`` summary file and save the generated workbook in
the same results directory::

    python .\analyze_summary.py "..\Results\Self-scheduling\summary_202301_202512_exact_V5_self_scheduling_k20.csv" -o "..\Results\Self-scheduling\analysis_202301_202512.xlsx"

The first path is the input summary CSV. The path after ``-o`` is the output
Excel workbook. Both paths may be changed when analyzing another data set.

The workbook contains two sheets:

* ``Monthly Analysis``: monthly profit, cost, and carbon metrics.
* ``Annual Summary``: complete calendar-year aggregate rates.
"""

from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


PROJECT_ROOT = Path(__file__).resolve().parent.parent
RESULTS_DIR = PROJECT_ROOT / "Results" / "Bidding"
DEFAULT_INPUT = RESULTS_DIR / "summary_202301_202512_exact_V5_k20.csv"

REQUIRED_COLUMNS = {
    "year_month",
    "k",
    "total_profit",
    "total_profit_actual",
    "total_cost",
    "total_cost_actual",
    "carbon_reduce",
    "total_carbon_actual",
}

PERCENT_FORMAT = "0.00%"
NUMBER_FORMAT = "#,##0.00"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


@dataclass(frozen=True)
class MonthlyMetrics:
    """Numeric values read or calculated for one source month."""

    month: str
    k: Decimal
    total_profit: Decimal
    profit_actual: Decimal
    total_cost: Decimal
    cost_actual: Decimal
    carbon_reduce: Decimal
    carbon_actual: Decimal

    @property
    def profit_increment(self) -> Decimal:
        return self.total_profit / self.profit_actual - Decimal("1")

    @property
    def controlled_profit_increment(self) -> Decimal:
        """Profit increment normalized to the controlled k-share baseline."""
        return self.profit_increment / self.k

    @property
    def cost_reduction(self) -> Decimal:
        return Decimal("1") - self.total_cost / self.cost_actual

    @property
    def rate_carbon(self) -> Decimal:
        """Carbon reduction relative to the matching original-mode emission."""
        if self.carbon_actual == 0:
            return Decimal("0")
        return self.carbon_reduce / self.carbon_actual


def parse_decimal(value: str | None, column: str, row_number: int) -> Decimal:
    """Parse a required numeric field and report a useful source-row error."""
    if value is None or not value.strip():
        raise ValueError(f"Row {row_number}: {column!r} is empty.")
    try:
        return Decimal(value.strip())
    except InvalidOperation as exc:
        raise ValueError(
            f"Row {row_number}: {column!r} is not a valid number: {value!r}."
        ) from exc


def validate_month(value: str, row_number: int) -> str:
    """Return a validated YYYYMM identifier."""
    if len(value) != 6 or not value.isdigit():
        raise ValueError(
            f"Row {row_number}: 'year_month' must use YYYYMM format: {value!r}."
        )
    month_number = int(value[4:])
    if month_number < 1 or month_number > 12:
        raise ValueError(f"Row {row_number}: invalid month in 'year_month': {value!r}.")
    return value


def read_summary(input_path: Path) -> list[MonthlyMetrics]:
    """Read and validate the monthly study summary."""
    if not input_path.is_file():
        raise FileNotFoundError(f"Input CSV does not exist: {input_path}")

    metrics: list[MonthlyMetrics] = []
    seen_months: set[str] = set()
    expected_k: Decimal | None = None
    with input_path.open("r", encoding="utf-8-sig", newline="") as source:
        reader = csv.DictReader(source)
        source_columns = set(reader.fieldnames or [])
        missing_columns = sorted(REQUIRED_COLUMNS - source_columns)
        if missing_columns:
            raise ValueError(
                "Input CSV is missing required columns: " + ", ".join(missing_columns)
            )

        for row_number, row in enumerate(reader, start=2):
            month = validate_month((row.get("year_month") or "").strip(), row_number)
            if month in seen_months:
                raise ValueError(f"Row {row_number}: duplicate month {month!r}.")
            seen_months.add(month)

            k = parse_decimal(row.get("k"), "k", row_number)
            if not Decimal("0") < k <= Decimal("1"):
                raise ValueError(
                    f"Row {row_number}: 'k' must be greater than 0 and at most 1."
                )
            if expected_k is None:
                expected_k = k
            elif k != expected_k:
                raise ValueError(
                    f"Row {row_number}: inconsistent 'k' value {k}; "
                    f"expected {expected_k} for every month."
                )

            total_profit = parse_decimal(row.get("total_profit"), "total_profit", row_number)
            profit_actual = parse_decimal(
                row.get("total_profit_actual"), "total_profit_actual", row_number
            )
            total_cost = parse_decimal(row.get("total_cost"), "total_cost", row_number)
            cost_actual = parse_decimal(
                row.get("total_cost_actual"), "total_cost_actual", row_number
            )
            if profit_actual == 0:
                raise ZeroDivisionError(
                    f"Row {row_number}: 'total_profit_actual' is zero."
                )
            if cost_actual == 0:
                raise ZeroDivisionError(f"Row {row_number}: 'total_cost_actual' is zero.")

            metrics.append(
                MonthlyMetrics(
                    month=month,
                    k=k,
                    total_profit=total_profit,
                    profit_actual=profit_actual,
                    total_cost=total_cost,
                    cost_actual=cost_actual,
                    carbon_reduce=parse_decimal(
                        row.get("carbon_reduce"), "carbon_reduce", row_number
                    ),
                    carbon_actual=parse_decimal(
                        row.get("total_carbon_actual"),
                        "total_carbon_actual",
                        row_number,
                    ),
                )
            )

    if not metrics:
        raise ValueError("Input CSV contains no data rows.")
    return sorted(metrics, key=lambda item: item.month)


def select_complete_calendar_years(
    metrics: list[MonthlyMetrics],
) -> list[MonthlyMetrics]:
    """Keep only years containing all twelve months, in chronological order."""
    by_year: dict[str, list[MonthlyMetrics]] = defaultdict(list)
    for item in metrics:
        by_year[item.month[:4]].append(item)

    selected: list[MonthlyMetrics] = []
    for year in sorted(by_year):
        year_metrics = sorted(by_year[year], key=lambda item: item.month)
        if [int(item.month[4:]) for item in year_metrics] == list(range(1, 13)):
            selected.extend(year_metrics)

    if not selected:
        raise ValueError("Input CSV contains no complete calendar year to analyze.")
    return selected


def annual_label(year: str, months: list[str]) -> str:
    """Return the display label for one calendar-year group.

    A complete January–December year is labelled as ``YYYY-(YYYY+1)``
    (e.g. ``"2023-2024"``).  This cross-year format is intentional and
    matches the paper's annual reporting convention.  Incomplete years
    fall back to the ``YYYY_M--YYYY_M`` range form.
    """
    month_numbers = [int(month[4:]) for month in months]
    if month_numbers == list(range(1, 13)):
        return f"{year}-{int(year) + 1}"
    return f"{year}_{month_numbers[0]}--{year}_{month_numbers[-1]}"


def period_label(first_month: str, last_month: str) -> str:
    """Label an arbitrary inclusive month range without leading month zeroes."""
    return (
        f"{first_month[:4]}_{int(first_month[4:])}--"
        f"{last_month[:4]}_{int(last_month[4:])}"
    )


def k_column_suffix(k: Decimal) -> str:
    """Return an identifier-safe percentage suffix, e.g. 0.2 -> ``k20``."""
    percentage = format((k * Decimal("100")).normalize(), "f")
    return "k" + percentage.replace(".", "_")


def aggregate_rates(
    period_metrics: list[MonthlyMetrics],
    period_name: str,
) -> tuple[Decimal, Decimal, Decimal, Decimal]:
    """Calculate weighted whole-fleet, controlled-share, cost, and carbon rates."""
    total_profit = sum(
        (item.total_profit for item in period_metrics), Decimal("0")
    )
    total_profit_actual = sum(
        (item.profit_actual for item in period_metrics), Decimal("0")
    )
    total_cost = sum((item.total_cost for item in period_metrics), Decimal("0"))
    total_cost_actual = sum(
        (item.cost_actual for item in period_metrics), Decimal("0")
    )
    total_carbon_reduce = sum(
        (item.carbon_reduce for item in period_metrics), Decimal("0")
    )
    total_carbon_actual = sum(
        (item.carbon_actual for item in period_metrics), Decimal("0")
    )
    if total_profit_actual == 0:
        raise ZeroDivisionError(f"Period {period_name}: total actual profit is zero.")
    if total_cost_actual == 0:
        raise ZeroDivisionError(f"Period {period_name}: total actual cost is zero.")
    profit_increment = total_profit / total_profit_actual - Decimal("1")
    controlled_profit_increment = (
        (total_profit - total_profit_actual)
        / sum(
            (item.k * item.profit_actual for item in period_metrics),
            Decimal("0"),
        )
    )
    carbon_reduction = (
        total_carbon_reduce / total_carbon_actual
        if total_carbon_actual != 0
        else Decimal("0")
    )
    return (
        profit_increment,
        controlled_profit_increment,
        Decimal("1") - total_cost / total_cost_actual,
        carbon_reduction,
    )


def style_sheet(
    sheet: Worksheet,
    percentage_columns: tuple[int, ...] = (),
    number_columns: tuple[int, ...] = (),
) -> None:
    """Apply compact, readable formatting to one result sheet."""
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 24
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column in percentage_columns:
        for cell in sheet.iter_cols(
            min_col=column, max_col=column, min_row=2, max_row=sheet.max_row
        ):
            for item in cell:
                item.number_format = PERCENT_FORMAT
    for column in number_columns:
        for cell in sheet.iter_cols(
            min_col=column, max_col=column, min_row=2, max_row=sheet.max_row
        ):
            for item in cell:
                item.number_format = NUMBER_FORMAT

    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        sheet.column_dimensions[column_cells[0].column_letter].width = min(width, 28)


def write_workbook(
    metrics: list[MonthlyMetrics],
    output_path: Path,
) -> None:
    """Create the monthly and annual worksheets and save the workbook."""
    workbook = Workbook()
    k_suffix = k_column_suffix(metrics[0].k)
    monthly_sheet = workbook.active
    monthly_sheet.title = "Monthly Analysis"
    monthly_sheet.append(
        [
            "Month",
            "profit increment",
            f"profit_increment_{k_suffix}",
            "cost reduction",
            "carbon reduction",
            "rate carbon",
        ]
    )
    for item in metrics:
        monthly_sheet.append(
            [
                item.month,
                float(item.profit_increment),
                float(item.controlled_profit_increment),
                float(item.cost_reduction),
                float(item.carbon_reduce),
                float(item.rate_carbon),
            ]
        )
    style_sheet(
        monthly_sheet,
        percentage_columns=(2, 3, 4, 6),
        number_columns=(5,),
    )

    annual_sheet = workbook.create_sheet("Annual Summary")
    annual_sheet.append(
        [
            "annual",
            "profit increment rate",
            f"profit_increment_rate_{k_suffix}",
            "cost reduction rate",
            "carbon reduction rate",
        ]
    )
    by_year: dict[str, list[MonthlyMetrics]] = defaultdict(list)
    for item in metrics:
        by_year[item.month[:4]].append(item)

    for year in sorted(by_year):
        year_metrics = sorted(by_year[year], key=lambda item: item.month)
        label = annual_label(year, [item.month for item in year_metrics])
        (
            profit_rate,
            controlled_profit_rate,
            cost_rate,
            carbon_rate,
        ) = aggregate_rates(year_metrics, label)

        annual_sheet.append(
            [
                label,
                float(profit_rate),
                float(controlled_profit_rate),
                float(cost_rate),
                float(carbon_rate),
            ]
        )

    all_months_label = period_label(metrics[0].month, metrics[-1].month)
    (
        profit_rate,
        controlled_profit_rate,
        cost_rate,
        carbon_rate,
    ) = aggregate_rates(metrics, all_months_label)
    annual_sheet.append(
        [
            all_months_label,
            float(profit_rate),
            float(controlled_profit_rate),
            float(cost_rate),
            float(carbon_rate),
        ]
    )
    style_sheet(annual_sheet, percentage_columns=(2, 3, 4, 5))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def analyze_summary(
    input_path: Path,
    output_path: Path | None = None,
) -> Path:
    """Analyze the source CSV and write the requested Excel workbook."""
    metrics = select_complete_calendar_years(read_summary(input_path))
    if output_path is None:
        output_path = RESULTS_DIR / (
            f"analysis_{metrics[0].month}_{metrics[-1].month}.xlsx"
        )
    if output_path.suffix.lower() != ".xlsx":
        raise ValueError(f"Output path must use the .xlsx extension: {output_path}")

    write_workbook(metrics, output_path)
    return output_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Generate monthly and annual battery-study analysis in Excel."
    )
    parser.add_argument(
        "input",
        nargs="?",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"Source summary CSV (default: {DEFAULT_INPUT})",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Output .xlsx; default is analysis_<start>_<end>.xlsx in Results/Bidding.",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    output_path = analyze_summary(args.input, args.output)
    print(f"Created {output_path}")


if __name__ == "__main__":
    main()
