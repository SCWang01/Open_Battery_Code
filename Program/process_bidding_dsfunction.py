"""Process the 36 monthly Bidding dsfunction workbooks."""

from __future__ import annotations

import argparse
import ast
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

try:
    from .add_dsfunction_state_array import STATE_ORDER, add_state_widths
    from .classify_p_ess import (
        PDMAX_HEADER,
        PCMAX_HEADER,
        STATUS_HEADER,
        SUMMARY_ORDER,
        as_float,
        classify_workbook,
        find_column,
        row_has_zero_power_limit,
    )
except ImportError:
    from add_dsfunction_state_array import STATE_ORDER, add_state_widths
    from classify_p_ess import (
        PDMAX_HEADER,
        PCMAX_HEADER,
        STATUS_HEADER,
        SUMMARY_ORDER,
        as_float,
        classify_workbook,
        find_column,
        row_has_zero_power_limit,
    )


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_BIDDING_DIR = PROJECT_ROOT / "Results" / "Bidding"
DEFAULT_SUMMARY = DEFAULT_BIDDING_DIR / "dsfunction_state_summary_202301_202512.xlsx"
MONTHS = {
    "January": 1,
    "February": 2,
    "March": 3,
    "April": 4,
    "May": 5,
    "June": 6,
    "July": 7,
    "August": 8,
    "September": 9,
    "October": 10,
    "November": 11,
    "December": 12,
}
EXPECTED_MONTHS = [(year, month) for year in range(2023, 2026) for month in range(1, 13)]
INPUT_PATTERN = re.compile(
    r"^dsfunction_(January|February|March|April|May|June|July|August|September|October|November|December)(20\d{2})_exact_V5_k20\.xlsx$"
)


def discover_inputs(input_dir: Path) -> list[tuple[int, int, Path]]:
    """Discover exactly one source workbook for each month in the target range."""
    discovered: dict[tuple[int, int], Path] = {}
    for path in input_dir.glob("dsfunction_*_exact_V5_k20.xlsx"):
        match = INPUT_PATTERN.match(path.name)
        if not match:
            continue
        month = (int(match.group(2)), MONTHS[match.group(1)])
        if month in discovered:
            raise ValueError(f"Duplicate source workbook for {month}: {path}")
        discovered[month] = path
    missing = [month for month in EXPECTED_MONTHS if month not in discovered]
    if missing:
        missing_text = ", ".join(f"{year:04d}-{month:02d}" for year, month in missing)
        raise FileNotFoundError(f"Missing 36-month dsfunction inputs: {missing_text}")
    return [(year, month, discovered[(year, month)]) for year, month in EXPECTED_MONTHS]


def output_path_for(input_path: Path, output_dir: Path) -> Path:
    """Build the stable per-month classified-width output name."""
    return output_dir / f"{input_path.stem}_classified_width{input_path.suffix}"


def summarize_output(path: Path) -> dict[str, Any]:
    """Read one generated workbook and return auditable monthly metrics."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = list(next(rows))
        state_col = find_column(headers, STATUS_HEADER) - 1
        pcmax_col = find_column(headers, PCMAX_HEADER) - 1
        pdmax_col = find_column(headers, PDMAX_HEADER) - 1
        width_col = find_column(headers, "dsfunction_state_width") - 1
        state_counts = {state: 0 for state in SUMMARY_ORDER}
        width_totals = {state: 0.0 for state in STATE_ORDER}
        pcmax_values: list[float] = []
        pdmax_values: list[float] = []
        raw_rows = 0
        skipped = 0

        for excel_row, row in enumerate(rows, start=2):
            if all(value is None for value in row):
                continue
            raw_rows += 1
            pcmax = as_float(row[pcmax_col], name=PCMAX_HEADER, excel_row=excel_row)
            pdmax = as_float(row[pdmax_col], name=PDMAX_HEADER, excel_row=excel_row)
            pcmax_values.append(pcmax)
            pdmax_values.append(pdmax)
            state = row[state_col]
            width_value = row[width_col]
            if row_has_zero_power_limit(pcmax, pdmax):
                if state is not None or width_value is not None:
                    raise ValueError(f"Row {excel_row}: zero-limit row must have blank analysis columns.")
                skipped += 1
                continue
            if state not in state_counts:
                raise ValueError(f"Row {excel_row}: unexpected P_ESS_state {state!r}.")
            if not isinstance(width_value, str):
                raise ValueError(f"Row {excel_row}: width output must be text.")
            widths = ast.literal_eval(width_value)
            if not isinstance(widths, (list, tuple)) or len(widths) != len(STATE_ORDER):
                raise ValueError(f"Row {excel_row}: expected {len(STATE_ORDER)} state widths.")
            state_counts[state] += 1
            for state_name, width in zip(STATE_ORDER, widths):
                numeric_width = float(width)
                if numeric_width < 0:
                    raise ValueError(f"Row {excel_row}: negative width for {state_name}.")
                width_totals[state_name] += numeric_width

        result: dict[str, Any] = {
            "raw_rows": raw_rows,
            "analyzed_rows": raw_rows - skipped,
            "skipped_zero_limit_rows": skipped,
            "Pcmax_min": min(pcmax_values),
            "Pcmax_max": max(pcmax_values),
            "Pdmax_min": min(pdmax_values),
            "Pdmax_max": max(pdmax_values),
        }
        result.update({f"state_{state}": state_counts[state] for state in SUMMARY_ORDER})
        result.update({f"width_{state}": width_totals[state] for state in STATE_ORDER})
        return result
    finally:
        workbook.close()


def write_summary(rows: list[dict[str, Any]], output_path: Path) -> None:
    """Write monthly and total metrics to an Excel workbook."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Monthly Summary"
    columns = [
        "year_month", "source_file", "output_file", "raw_rows", "analyzed_rows",
        "skipped_zero_limit_rows", "Pcmax_min", "Pcmax_max", "Pdmax_min", "Pdmax_max",
        *[f"state_{state}" for state in SUMMARY_ORDER],
        *[f"width_{state}" for state in STATE_ORDER],
    ]
    worksheet.append(columns)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in rows:
        worksheet.append([row[column] for column in columns])

    total = {column: "" for column in columns}
    total["year_month"] = "TOTAL"
    for column in ("raw_rows", "analyzed_rows", "skipped_zero_limit_rows", *[f"state_{s}" for s in SUMMARY_ORDER], *[f"width_{s}" for s in STATE_ORDER]):
        total[column] = sum(row[column] for row in rows)
    total["Pcmax_min"] = min(row["Pcmax_min"] for row in rows)
    total["Pcmax_max"] = max(row["Pcmax_max"] for row in rows)
    total["Pdmax_min"] = min(row["Pdmax_min"] for row in rows)
    total["Pdmax_max"] = max(row["Pdmax_max"] for row in rows)
    worksheet.append([total[column] for column in columns])
    for column in range(1, worksheet.max_column + 1):
        worksheet.column_dimensions[chr(64 + column) if column <= 26 else "A"].width = 20
    worksheet.freeze_panes = "A2"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def process_all(input_dir: Path, output_dir: Path, summary_path: Path) -> list[dict[str, Any]]:
    """Process all 36 source workbooks and write the aggregate summary."""
    summaries: list[dict[str, Any]] = []
    for year, month, input_path in discover_inputs(input_dir):
        output_path = output_path_for(input_path, output_dir)
        counts = classify_workbook(input_path, output_path)
        add_state_widths(output_path, output_path)
        summary = summarize_output(output_path)
        summary.update(
            {
                "year_month": f"{year:04d}-{month:02d}",
                "source_file": input_path.name,
                "output_file": output_path.name,
            }
        )
        expected_skipped = counts["skipped_zero_limit_rows"]
        if summary["skipped_zero_limit_rows"] != expected_skipped:
            raise ValueError(f"{input_path.name}: skip count did not reconcile.")
        summaries.append(summary)
        print(
            f"{summary['year_month']}: raw={summary['raw_rows']}, "
            f"analyzed={summary['analyzed_rows']}, skipped={summary['skipped_zero_limit_rows']} -> {output_path.name}"
        )
    write_summary(summaries, summary_path)
    return summaries


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Process 36 months of Bidding dsfunction workbooks.")
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_BIDDING_DIR)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_BIDDING_DIR)
    parser.add_argument("--summary", type=Path, default=DEFAULT_SUMMARY)
    return parser


def main() -> None:
    arguments = build_parser().parse_args()
    summaries = process_all(arguments.input_dir, arguments.output_dir, arguments.summary)
    print(f"Processed months: {len(summaries)}")
    print(f"Summary: {arguments.summary}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        raise SystemExit(f"Error: {exc}") from None
