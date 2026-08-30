"""Classify P_ESS values using row-specific dsfunction and power-limit data."""

from __future__ import annotations

import argparse
import ast
import math
from collections import Counter
from copy import copy
from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = PROJECT_ROOT / "Results" / "Bidding" / "dsfunction_May2025_exact_V5_k20.xlsx"
STATUS_HEADER = "P_ESS_state"
PCMAX_HEADER = "Pcmax"
PDMAX_HEADER = "Pdmax"
REL_TOL = 1e-9
ABS_TOL = 1e-6
MAX_RATE_ABS_TOL = 5e-4
SUMMARY_ORDER = ("MC", "MD", "NU", "DC", "DD", "CC", "CD", "NC")
SKIPPED_ZERO_LIMIT_ROWS = "skipped_zero_limit_rows"


def values_match(left: float, right: float) -> bool:
    """Compare numeric values while allowing spreadsheet round-off."""
    return math.isclose(left, right, rel_tol=REL_TOL, abs_tol=ABS_TOL)


def max_rate_values_match(left: float, right: float) -> bool:
    """Compare power values after the upstream four-decimal rounding."""
    return math.isclose(
        left,
        right,
        rel_tol=REL_TOL,
        abs_tol=MAX_RATE_ABS_TOL,
    )


def as_float(value: Any, *, name: str, excel_row: int) -> float:
    """Convert a worksheet value to a finite float."""
    try:
        result = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Row {excel_row}: {name} is not a valid number: {value!r}") from exc
    if not math.isfinite(result):
        raise ValueError(f"Row {excel_row}: {name} is not finite: {value!r}")
    return result


def parse_dsfunction(value: Any, excel_row: int) -> Sequence[Sequence[Any]]:
    """Parse one dsfunction cell with seven core values and optional SOC."""
    if isinstance(value, str):
        try:
            value = ast.literal_eval(value)
        except (SyntaxError, ValueError) as exc:
            raise ValueError(f"Row {excel_row}: dsfunction cannot be parsed: {exc}") from exc
    if not isinstance(value, (list, tuple)) or not value:
        raise ValueError(f"Row {excel_row}: dsfunction must be a non-empty 2-D array.")
    for segment_number, segment in enumerate(value, start=1):
        if not isinstance(segment, (list, tuple)) or len(segment) not in (7, 8):
            raise ValueError(
                f"Row {excel_row}: dsfunction segment {segment_number} must contain "
                "7 core values and may contain an eighth SOC value."
            )
    return value


def find_column(headers: list[Any], header_name: str) -> int:
    """Find a unique one-based worksheet column by header."""
    matches = [index for index, value in enumerate(headers, start=1) if value == header_name]
    if len(matches) != 1:
        raise ValueError(
            f"Worksheet header {header_name!r} appears {len(matches)} times; exactly one is required."
        )
    return matches[0]


def row_has_zero_power_limit(pcmax: float, pdmax: float) -> bool:
    """Return whether a row has no available charge or discharge capability."""
    return values_match(pcmax, 0.0) or values_match(pdmax, 0.0)


def classify_p_ess(
    p_ess_value: Any,
    ds_row: Sequence[Any],
    excel_row: int,
    *,
    pcmax: float,
    pdmax: float,
) -> str:
    """Classify one segment using signed row-level Pcmax/Pdmax values."""
    p_ess = as_float(p_ess_value, name="P_ESS", excel_row=excel_row)
    if max_rate_values_match(p_ess, pcmax):
        return "MC"
    if max_rate_values_match(p_ess, pdmax):
        return "MD"
    if values_match(p_ess, 0.0):
        return "NU"

    seventh = as_float(ds_row[6], name="dsfunction segment column 7", excel_row=excel_row)
    if values_match(seventh, 1.0):
        return "DC" if p_ess > 0 else "CC"
    if values_match(seventh, -1.0):
        return "DD" if p_ess > 0 else "CD"
    if values_match(seventh, 0.0):
        fourth = as_float(ds_row[3], name="dsfunction segment column 4", excel_row=excel_row)
        fifth = as_float(ds_row[4], name="dsfunction segment column 5", excel_row=excel_row)
        if values_match(fourth, fifth):
            return "DD" if p_ess > 0 else "CC"
        if p_ess > 0:
            return "DC" if fourth > fifth else "DD"
        return "CC" if fourth > fifth else "CD"
    return "NC"


def classify_p_ess_from_dsfunction(
    p_ess_value: Any,
    dsfunction_value: Any,
    excel_row: int,
    *,
    pcmax: float,
    pdmax: float,
) -> str:
    """Match P_ESS to its unique dsfunction segment and classify it."""
    p_ess = as_float(p_ess_value, name="P_ESS", excel_row=excel_row)
    if (
        max_rate_values_match(p_ess, pcmax)
        or max_rate_values_match(p_ess, pdmax)
        or values_match(p_ess, 0.0)
    ):
        return classify_p_ess(p_ess, (), excel_row, pcmax=pcmax, pdmax=pdmax)

    dsfunction = parse_dsfunction(dsfunction_value, excel_row)
    matches: list[Sequence[Any]] = []
    for segment_number, segment in enumerate(dsfunction, start=1):
        candidate = as_float(
            segment[2],
            name=f"dsfunction segment {segment_number}, column 3",
            excel_row=excel_row,
        )
        if values_match(p_ess, candidate):
            matches.append(segment)
    if not matches:
        raise ValueError(
            f"Row {excel_row}: P_ESS={p_ess:g} matched no rows in dsfunction column 3."
        )
    states = {
        classify_p_ess(p_ess, segment, excel_row, pcmax=pcmax, pdmax=pdmax)
        for segment in matches
    }
    if len(states) != 1:
        raise ValueError(
            f"Row {excel_row}: P_ESS={p_ess:g} matched multiple segments with "
            f"different states: {sorted(states)}."
        )
    return states.pop()


def default_output_path(input_path: Path) -> Path:
    """Return the standalone classifier's default output path."""
    return input_path.with_name(f"{input_path.stem}_classified{input_path.suffix}")


def classify_workbook(
    input_path: Path,
    output_path: Path,
    sheet_name: str | None = None,
) -> Counter:
    """Add/update P_ESS_state using each row's Pcmax/Pdmax and save a copy."""
    if input_path.resolve() == output_path.resolve():
        raise ValueError("Classification output must differ from its input workbook.")
    if not input_path.is_file():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    workbook = load_workbook(input_path, data_only=False)
    try:
        if sheet_name is None:
            worksheet = workbook.active
        elif sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            raise ValueError(
                f"Worksheet {sheet_name!r} not found; available worksheets: "
                f"{', '.join(workbook.sheetnames)}"
            )

        headers = [cell.value for cell in worksheet[1]]
        p_ess_col = find_column(headers, "P_ESS")
        find_column(headers, "dsfunction")
        find_column(headers, PCMAX_HEADER)
        find_column(headers, PDMAX_HEADER)
        status_col = p_ess_col + 1
        existing = [index for index, value in enumerate(headers, start=1) if value == STATUS_HEADER]
        if existing and existing != [status_col]:
            raise ValueError(f"Existing {STATUS_HEADER!r} must be immediately right of P_ESS.")
        if not existing:
            worksheet.insert_cols(status_col, amount=1)
            worksheet.cell(row=1, column=status_col).value = STATUS_HEADER

        headers = [cell.value for cell in worksheet[1]]
        dsfunction_col = find_column(headers, "dsfunction")
        p_ess_col = find_column(headers, "P_ESS")
        pcmax_col = find_column(headers, PCMAX_HEADER)
        pdmax_col = find_column(headers, PDMAX_HEADER)
        status_col = find_column(headers, STATUS_HEADER)

        for row in range(1, worksheet.max_row + 1):
            source_cell = worksheet.cell(row=row, column=p_ess_col)
            target_cell = worksheet.cell(row=row, column=status_col)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.hyperlink:
                target_cell.hyperlink = copy(source_cell.hyperlink)
        source_letter = get_column_letter(p_ess_col)
        status_letter = get_column_letter(status_col)
        worksheet.column_dimensions[status_letter].width = max(
            worksheet.column_dimensions[source_letter].width or 13.0,
            float(len(STATUS_HEADER) + 2),
        )
        worksheet.cell(row=1, column=status_col).value = STATUS_HEADER

        counts: Counter = Counter()
        for excel_row in range(2, worksheet.max_row + 1):
            pcmax = as_float(
                worksheet.cell(excel_row, pcmax_col).value,
                name=PCMAX_HEADER,
                excel_row=excel_row,
            )
            pdmax = as_float(
                worksheet.cell(excel_row, pdmax_col).value,
                name=PDMAX_HEADER,
                excel_row=excel_row,
            )
            target_cell = worksheet.cell(row=excel_row, column=status_col)
            if row_has_zero_power_limit(pcmax, pdmax):
                target_cell.value = None
                counts[SKIPPED_ZERO_LIMIT_ROWS] += 1
                continue
            state = classify_p_ess_from_dsfunction(
                worksheet.cell(excel_row, p_ess_col).value,
                worksheet.cell(excel_row, dsfunction_col).value,
                excel_row,
                pcmax=pcmax,
                pdmax=pdmax,
            )
            target_cell.value = state
            target_cell.number_format = "@"
            counts[state] += 1

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return counts
    finally:
        workbook.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Classify each P_ESS row using its signed Pcmax and Pdmax values."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--sheet")
    return parser


def main() -> None:
    arguments = build_parser().parse_args()
    output_path = arguments.output or default_output_path(arguments.input)
    counts = classify_workbook(arguments.input, output_path, arguments.sheet)
    print(f"Done: {output_path}")
    print(f"Total analyzed records: {sum(counts[state] for state in SUMMARY_ORDER)}")
    print(f"Skipped zero-limit rows: {counts[SKIPPED_ZERO_LIMIT_ROWS]}")
    for state in SUMMARY_ORDER:
        print(f"P_ESS_state={state}: {counts[state]}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        raise SystemExit(f"Error: {exc}") from None
