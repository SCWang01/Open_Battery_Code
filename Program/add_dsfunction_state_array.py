"""Add dsfunction state interval-width arrays to classified workbooks."""

from __future__ import annotations

import argparse
import ast
from collections import Counter
from copy import copy
from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    from .classify_p_ess import (
        PCMAX_HEADER,
        PDMAX_HEADER,
        STATUS_HEADER,
        as_float,
        classify_p_ess,
        find_column,
        row_has_zero_power_limit,
        values_match,
    )
except ImportError:
    from classify_p_ess import (
        PCMAX_HEADER,
        PDMAX_HEADER,
        STATUS_HEADER,
        as_float,
        classify_p_ess,
        find_column,
        row_has_zero_power_limit,
        values_match,
    )


STATE_ORDER = ("CC", "CD", "NU", "DC", "DD")
OUTPUT_HEADER = "dsfunction_state_width"
LEGACY_OUTPUT_HEADERS = ("dsfunction_state_share", "dsfunction_state_array")
MB_TOLERANCE = 1e-5
DSFUNCTION_COLUMN_COUNT = 8


def parse_dsfunction(dsfunction_value: Any, excel_row: int) -> Sequence[Sequence[Any]]:
    """Parse one dsfunction cell and require eight values per embedded row."""
    if not isinstance(dsfunction_value, str):
        raise ValueError(f"Row {excel_row}: dsfunction is not text.")
    try:
        dsfunction = ast.literal_eval(dsfunction_value)
    except (SyntaxError, ValueError) as exc:
        raise ValueError(f"Row {excel_row}: dsfunction is not a valid literal.") from exc
    if not isinstance(dsfunction, (list, tuple)) or not dsfunction:
        raise ValueError(f"Row {excel_row}: dsfunction must be a non-empty 2-D array.")
    for segment_number, segment in enumerate(dsfunction, start=1):
        if not isinstance(segment, (list, tuple)) or len(segment) != DSFUNCTION_COLUMN_COUNT:
            raise ValueError(
                f"Row {excel_row}: dsfunction segment {segment_number} must contain "
                f"exactly {DSFUNCTION_COLUMN_COUNT} values."
            )
    return dsfunction


def format_width(value: float) -> str:
    """Format one interval width compactly while retaining precision."""
    if abs(value) < 5e-13:
        return "0"
    if abs(value - 1.0) < 5e-13:
        return "1"
    return f"{value:.10f}".rstrip("0").rstrip(".")


def is_marginal_boundary(soc_value: float) -> bool:
    """Return whether the post-action SOC is at either marginal boundary."""
    return abs(soc_value) < MB_TOLERANCE or abs(soc_value - 1.0) < MB_TOLERANCE


def state_width_array(
    dsfunction_value: Any,
    excel_row: int,
    *,
    pcmax: float,
    pdmax: float,
) -> tuple[str, list[str], list[float]]:
    """Classify segments and sum widths, excluding internally tagged mB segments."""
    dsfunction = parse_dsfunction(dsfunction_value, excel_row)
    states: list[str] = []
    widths = Counter({state: 0.0 for state in STATE_ORDER})
    for segment_number, segment in enumerate(dsfunction, start=1):
        values = [
            as_float(
                value,
                name=f"dsfunction segment {segment_number}, column {column_number}",
                excel_row=excel_row,
            )
            for column_number, value in enumerate(segment, start=1)
        ]
        lower, upper, p_ess, fourth, fifth, _sixth, seventh, eighth = values
        if is_marginal_boundary(eighth):
            state = "mB"
        else:
            state = classify_p_ess(
                p_ess,
                segment,
                excel_row,
                pcmax=pcmax,
                pdmax=pdmax,
            )
            if values_match(seventh, 0.0) and values_match(fourth - fifth, 0.0):
                if p_ess > 0:
                    state = "DD"
                elif p_ess < 0:
                    state = "CC"
        width = upper - lower
        if width <= 0:
            raise ValueError(
                f"Row {excel_row}: dsfunction segment {segment_number} must have a positive interval width."
            )
        states.append(state)
        if state in STATE_ORDER:
            widths[state] += width
    state_widths = [widths[state] for state in STATE_ORDER]
    array = "[" + ",".join(format_width(width) for width in state_widths) + "]"
    return array, states, state_widths


def _ensure_output_column(worksheet: Any, headers: list[Any]) -> int:
    """Find or append the width column without disturbing source columns."""
    existing = [
        index
        for index, value in enumerate(headers, start=1)
        if value in (OUTPUT_HEADER, *LEGACY_OUTPUT_HEADERS)
    ]
    if len(existing) > 1:
        raise ValueError(f"Multiple {OUTPUT_HEADER!r} columns found: {existing}")
    if existing:
        return existing[0]
    output_col = worksheet.max_column + 1
    source_col = find_column(headers, STATUS_HEADER)
    for row in range(1, worksheet.max_row + 1):
        source_cell = worksheet.cell(row=row, column=source_col)
        target_cell = worksheet.cell(row=row, column=output_col)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.hyperlink:
            target_cell.hyperlink = copy(source_cell.hyperlink)
    worksheet.cell(row=1, column=output_col).value = OUTPUT_HEADER
    worksheet.column_dimensions[get_column_letter(output_col)].width = 60
    return output_col


def add_state_widths(
    input_path: Path,
    output_path: Path | None = None,
    sheet_name: str | None = None,
) -> Counter:
    """Add/update width arrays; same-path writes use a temporary workbook."""
    output_path = output_path or input_path
    if not input_path.is_file():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    workbook = load_workbook(input_path, data_only=False)
    try:
        if sheet_name is None:
            worksheet = workbook.active
        elif sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            raise ValueError(f"Worksheet {sheet_name!r} not found.")
        headers = [cell.value for cell in worksheet[1]]
        dsfunction_col = find_column(headers, "dsfunction")
        pcmax_col = find_column(headers, PCMAX_HEADER)
        pdmax_col = find_column(headers, PDMAX_HEADER)
        output_col = _ensure_output_column(worksheet, headers)

        counts: Counter = Counter()
        for excel_row in range(2, worksheet.max_row + 1):
            pcmax = as_float(worksheet.cell(excel_row, pcmax_col).value, name=PCMAX_HEADER, excel_row=excel_row)
            pdmax = as_float(worksheet.cell(excel_row, pdmax_col).value, name=PDMAX_HEADER, excel_row=excel_row)
            cell = worksheet.cell(excel_row, output_col)
            if row_has_zero_power_limit(pcmax, pdmax):
                cell.value = None
                counts["skipped_zero_limit_rows"] += 1
                continue
            value, states, _ = state_width_array(
                worksheet.cell(excel_row, dsfunction_col).value,
                excel_row,
                pcmax=pcmax,
                pdmax=pdmax,
            )
            cell.value = value
            cell.number_format = "@"
            counts.update(states)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        if input_path.resolve() == output_path.resolve():
            temporary = output_path.with_name(f"{output_path.stem}.tmp{output_path.suffix}")
            workbook.save(temporary)
            temporary.replace(output_path)
        else:
            workbook.save(output_path)
        return counts
    finally:
        workbook.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Add dsfunction state widths to a classified workbook.")
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--sheet")
    return parser


def main() -> None:
    arguments = build_parser().parse_args()
    counts = add_state_widths(arguments.input, arguments.output, arguments.sheet)
    print(f"Done: {arguments.output or arguments.input}")
    print(f"State segment counts: {dict(counts)}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        raise SystemExit(f"Error: {exc}") from None
